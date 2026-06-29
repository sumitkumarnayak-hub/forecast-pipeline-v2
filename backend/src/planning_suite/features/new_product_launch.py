###############################################################
# newlaunchv2_full.py  —  New Product Launch Module
#
# Flow (per launch type):
#   Stage 1 – "input"        : select plan level, pick cities/hubs,
#                              download template, upload filled file
#   Stage 2 – "split_review" : (city-level upload) editable hub split
#                              derived from hub_suggestion_latest.parquet
#   Stage 3 – "set_date"     : Monday ≥ T+4 launch date picker
#   Stage 4 – "confirm"      : review + submit  →  mail placeholder
#
# Duplicate rule:
#   • After city-level upload  → check city + product_id
#   • After hub-level upload   → check city + hub + product_id
#   → Duplicate found: show existing submission (from Submission_Log)
#     in editable mode; user updates and re-submits
###############################################################

import pandas as pd
import numpy as np
import gspread
import io
import os
from datetime import datetime, date, timedelta
from google.oauth2.service_account import Credentials
import uuid
import pandera as pa

from planning_suite.config import (
    CLUSTER_MASTER_SHEET_KEY,
    GOOGLE_CREDENTIALS_PATH,
    HUB_LEVEL_PLANNING_SHEET_KEY,
)

WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

CITY_UPLOAD_SCHEMA = pa.DataFrameSchema({
    "city_name": pa.Column(str, coerce=True),
    "product_id": pa.Column(str, coerce=True),
    "product_name": pa.Column(str, coerce=True, nullable=True),
    "category": pa.Column(str, coerce=True, nullable=True),
    "MRP": pa.Column(float, coerce=True, nullable=True),
    **{day: pa.Column(int, coerce=True, nullable=True) for day in WEEKDAYS}
}, strict=False)

HUB_UPLOAD_SCHEMA = pa.DataFrameSchema({
    "city_name": pa.Column(str, coerce=True),
    "hub_name": pa.Column(str, coerce=True),
    "product_id": pa.Column(str, coerce=True),
    "product_name": pa.Column(str, coerce=True, nullable=True),
    "category": pa.Column(str, coerce=True, nullable=True),
    "MRP": pa.Column(float, coerce=True, nullable=True),
    **{day: pa.Column(int, coerce=True, nullable=True) for day in WEEKDAYS}
}, strict=False)

SERVICE_ACCOUNT_FILE = GOOGLE_CREDENTIALS_PATH
SPREADSHEET_ID = HUB_LEVEL_PLANNING_SHEET_KEY
MASTER_FILE_ID = CLUSTER_MASTER_SHEET_KEY
MASTER_SHEET_NAME    = "P-L Master"
OUTPUT_SHEET_NAME    = "Launch_Output"
LOG_SHEET_NAME       = "Submission_Log"

WEEKDAYS    = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
OUTPUTS_DIR = "outputs"
SALIENCE_SHEET_NAME = "Hub level Suggestion"
HUB_SKU_MASTER_SHEET = "Hub Sku Master"

LOG_HEADERS = [
    "Timestamp", "Submission_ID", "Submission_Type",
    "Product ID", "Product Name", "Category",
    "City", "Hub", "MRP", "Start Date",
    "Status", "Rejection_Reason", "Submitted_By",
] + WEEKDAYS


# ──────────────────────────────────────────────────────────────────
# COLUMN-NAME NORMALISER  (handles any casing / spacing from sheets)
# ──────────────────────────────────────────────────────────────────
def _canon(s: str) -> str:
    return s.strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        k = _canon(col)
        if k == "subcategory":   rename[col] = "sub_category"
        elif k == "baseplan":    rename[col] = "Base_plan"
        elif k == "planflag":    rename[col] = "Plan Flag"
        elif k == "productid":   rename[col] = "Product id"
        elif k == "cityname":   rename[col] = "city_name"
        elif k == "hubname":    rename[col] = "hub_name"
    if rename:
        df = df.rename(columns=rename)
    for c in df.select_dtypes(include="object").columns:
        df[c] = df[c].astype(str).str.strip()
    return df


def _resolve_col(df: pd.DataFrame, *candidates: str) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    lookup = {_canon(c): c for c in df.columns}
    for cand in candidates:
        hit = lookup.get(_canon(cand))
        if hit:
            return hit
    return None


def _subcat_col(df: pd.DataFrame) -> str:
    for c in ["sub_category", "sub category", "Sub-category", "Sub Category"]:
        if c in df.columns:
            return c
    for c in df.columns:
        if _canon(c) == "subcategory":
            return c
    return "sub_category"


# ──────────────────────────────────────────────────────────────────
# GOOGLE SHEETS  CONNECTORS
# ──────────────────────────────────────────────────────────────────
def _get_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    return gspread.authorize(creds)


def _open_sheet(spreadsheet_id: str, sheet_name: str):
    client = _get_client()
    sh = client.open_by_key(spreadsheet_id)
    try:
        return sh.worksheet(sheet_name)
    except Exception:
        ws = sh.add_worksheet(sheet_name, rows="5000", cols="50")
        return ws


# ──────────────────────────────────────────────────────────────────
# PRODUCT MASTER  (Google Sheet)
# ──────────────────────────────────────────────────────────────────

def load_product_master() -> pd.DataFrame:
    sheet = _open_sheet(MASTER_FILE_ID, MASTER_SHEET_NAME)
    data  = sheet.get_values("B:N")
    if len(data) < 2:
        return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=data[0])
    df = _normalize_columns(df)
    for c in df.columns:
        if "order" in c.lower() and "type" in c.lower():
            df = df[df[c] == "E"]
            break
    pid_col = next((c for c in ["Product id", "Product ID"] if c in df.columns), None)
    if pid_col:
        df = df.drop_duplicates(subset=[pid_col])
    return df


def _subcat_col_master(df: pd.DataFrame) -> str:
    return _subcat_col(df)


def get_categories(df_master: pd.DataFrame) -> list:
    c = _subcat_col_master(df_master)
    return sorted(df_master[c].dropna().unique().tolist()) if c in df_master.columns else []


def get_products_by_category(df_master: pd.DataFrame, category: str) -> list:
    c = _subcat_col_master(df_master)
    sub = df_master[df_master[c] == category] if c in df_master.columns else df_master
    return sorted(sub["Product Name"].dropna().unique().tolist())


def get_product_id(df_master: pd.DataFrame, product_name: str) -> str:
    row = df_master[df_master["Product Name"] == product_name]
    if row.empty:
        return ""
    for c in ["Product id", "Product ID"]:
        if c in row.columns:
            return str(row[c].iloc[0])
    return ""


def get_product_info(df_master: pd.DataFrame, product_id: str) -> dict:
    for pid_col in ["Product id", "Product ID"]:
        if pid_col not in df_master.columns:
            continue
        row = df_master[df_master[pid_col].astype(str) == str(product_id)]
        if not row.empty:
            sc = _subcat_col_master(df_master)
            return {
                "name":     row["Product Name"].iloc[0] if "Product Name" in row.columns else "",
                "category": row[sc].iloc[0] if sc in row.columns else "",
            }
    return {"name": "", "category": ""}


# ──────────────────────────────────────────────────────────────────
# HUB SALIENCE  (Hub level Suggestion sheet — same as Product)
# ──────────────────────────────────────────────────────────────────

def load_salience_source() -> pd.DataFrame:
    """Load hub-level plan from Hub Level Planning sheet (Product parity)."""
    sheet = _open_sheet(SPREADSHEET_ID, SALIENCE_SHEET_NAME)
    data = sheet.get_values("A:F")
    if not data or len(data) < 2:
        return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=data[0])
    df = _normalize_columns(df)
    if "Base_plan" not in df.columns:
        bp = _resolve_col(df, "Base_plan", "Base plan", "base_plan")
        if bp:
            df = df.rename(columns={bp: "Base_plan"})
    if "Base_plan" not in df.columns:
        return pd.DataFrame()
    df["Base_plan"] = pd.to_numeric(df["Base_plan"], errors="coerce").fillna(0).round()
    df = df[df["Base_plan"] > 0]
    return df.reset_index(drop=True)


def compute_salience_category(df: pd.DataFrame) -> pd.DataFrame:
    """Salience at city × hub × category (Product newlaunchv2_full.compute_salience_category)."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    df = _normalize_columns(df.copy())
    city_col = _resolve_col(df, "city_name", "City Name", "city")
    hub_col = _resolve_col(df, "hub_name", "Hub Name", "hub")
    sc_col = _subcat_col(df)
    if not city_col or not hub_col or sc_col not in df.columns or "Base_plan" not in df.columns:
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    if sc_col != "sub_category":
        df = df.rename(columns={sc_col: "sub_category"})

    hub_tot = (
        df.groupby([city_col, hub_col, "sub_category"], as_index=False)["Base_plan"]
        .sum()
        .rename(columns={city_col: "city_name", hub_col: "hub_name", "Base_plan": "hub_total"})
    )
    city_tot = (
        df.groupby([city_col, "sub_category"], as_index=False)["Base_plan"]
        .sum()
        .rename(columns={city_col: "city_name", "Base_plan": "city_total"})
    )
    merged = pd.merge(hub_tot, city_tot, on=["city_name", "sub_category"], how="left")
    merged["salience"] = np.where(
        merged["city_total"] > 0,
        merged["hub_total"] / merged["city_total"],
        0.0,
    )
    return merged[["city_name", "hub_name", "sub_category", "salience"]]



def load_hub_salience() -> pd.DataFrame:
    """
    Returns city_name, hub_name, sub_category, salience (+ day when available).

    Primary source: Google Sheets **Hub level Suggestion** (Product).
    Fallback: outputs/hub_suggestion_latest.parquet if present and valid.
    """
    raw = load_salience_source()
    if not raw.empty:
        return compute_salience_category(raw)

    path = os.path.join(OUTPUTS_DIR, "hub_suggestion_latest.parquet")
    if not os.path.exists(path):
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    df = pd.read_parquet(path)
    df = _normalize_columns(df)
    sc = _subcat_col(df)
    if sc != "sub_category":
        df = df.rename(columns={sc: "sub_category"})

    city_col = _resolve_col(df, "city_name")
    hub_col = _resolve_col(df, "hub_name")
    if not city_col or not hub_col or "sub_category" not in df.columns:
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    if city_col != "city_name":
        df = df.rename(columns={city_col: "city_name"})
    if hub_col != "hub_name":
        df = df.rename(columns={hub_col: "hub_name"})

    if "Base_plan" in df.columns and "salience" not in df.columns:
        return compute_salience_category(df)

    if "salience" in df.columns:
        cols = ["city_name", "hub_name", "sub_category", "salience"]
        if "day" in df.columns:
            cols.append("day")
        return df[[c for c in cols if c in df.columns]]

    return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])


def get_cities_from_salience(sal_df: pd.DataFrame) -> list:
    if sal_df is None or sal_df.empty:
        return []
    city_col = _resolve_col(sal_df, "city_name", "City Name", "city")
    if not city_col:
        return []
    return sorted(sal_df[city_col].dropna().astype(str).str.strip().unique().tolist())


def _require_salience(sal_df: pd.DataFrame) -> bool:
    if sal_df is not None and not sal_df.empty:
        return True
    print(
        "Could not load hub salience from **Hub level Suggestion** on the Hub Level "
        f"Planning sheet. Open the sheet and confirm tab **{SALIENCE_SHEET_NAME}** has "
        "columns city_name, hub_name, sub category, and Base_plan with data."
    )
    return False


def get_hubs_for_city(sal_df: pd.DataFrame, city: str, category: str = None) -> list:
    if sal_df is None or sal_df.empty:
        return []
    city_col = _resolve_col(sal_df, "city_name", "City Name", "city")
    hub_col = _resolve_col(sal_df, "hub_name", "Hub Name", "hub")
    if not city_col or not hub_col:
        return []
    mask = sal_df[city_col].astype(str).str.strip() == str(city).strip()
    if category:
        sc = _subcat_col(sal_df)
        if sc in sal_df.columns:
            mask &= sal_df[sc].astype(str).str.strip().str.lower() == category.strip().lower()
    return sorted(sal_df.loc[mask, hub_col].dropna().astype(str).str.strip().unique().tolist())


# ──────────────────────────────────────────────────────────────────
# CITY → HUB  SPLIT  (per-day salience allocation)
# ──────────────────────────────────────────────────────────────────
def _weighted_alloc(total: int, hub_sal: dict) -> dict:
    """Allocate 'total' units across hubs by salience; every hub gets ≥ 1."""
    hubs = list(hub_sal.keys())
    n    = len(hubs)
    if n == 0:
        return {}
    if total <= 0:
        return {h: 0 for h in hubs}
    if total < n:
        alloc = {h: 0 for h in hubs}
        for h in hubs[:total]:
            alloc[h] = 1
        return alloc
    alloc     = {h: 1 for h in hubs}
    remaining = total - n
    total_sal = sum(hub_sal.values())
    if total_sal == 0:
        base, rem = divmod(remaining, n)
        for h in hubs:
            alloc[h] += base
        for h in hubs[:rem]:
            alloc[h] += 1
        return alloc
    raw = {h: remaining * hub_sal[h] / total_sal for h in hubs}
    for h in hubs:
        alloc[h] += int(raw[h])
    leftover = remaining - (sum(alloc.values()) - n)
    for h in sorted(hubs, key=lambda x: raw[x] - int(raw[x]), reverse=True)[:leftover]:
        alloc[h] += 1
    return alloc


def split_city_to_hubs(
    city_df: pd.DataFrame,
    sal_df: pd.DataFrame,
    forced_hubs: dict = None,   # {city: [hub, ...]} override from user selection
) -> tuple:
    """
    city_df : rows with city_name, product_id, product_name, category, MRP, Mon-Sun
    sal_df  : hub salience DataFrame from load_hub_salience()
    forced_hubs : if provided, only split to these hubs per city

    Returns (hub_split_df, zero_sal_info)
      hub_split_df : city_name, hub_name, product_id, product_name, category, MRP, Mon-Sun
      zero_sal_info: {city: [hubs with zero salience]}
    """
    rows          = []
    zero_sal_info = {}

    for _, r in city_df.iterrows():
        city     = str(r.get("city_name", "")).strip()
        category = str(r.get("category", "")).strip()
        if not city:
            continue

        # Hubs to split into
        if forced_hubs and city in forced_hubs:
            hubs = forced_hubs[city]
        else:
            hubs = get_hubs_for_city(sal_df, city, category)

        if not hubs:
            continue

        zero_hubs = []
        sc_col = _subcat_col(sal_df)
        per_day_salience = "day" in sal_df.columns
        # Pre-compute per-day allocations
        day_alloc = {}
        for day in WEEKDAYS:
            day_sal = {}
            for hub in hubs:
                mask = (
                    (sal_df["city_name"].astype(str).str.strip() == city) &
                    (sal_df["hub_name"].astype(str).str.strip() == hub) &
                    (sal_df[sc_col].astype(str).str.strip().str.lower() == category.lower())
                )
                if per_day_salience:
                    mask &= sal_df["day"].astype(str).str.strip() == day
                rows_match = sal_df[mask]
                day_sal[hub] = float(rows_match["salience"].iloc[0]) if not rows_match.empty else 0.0

            # Track zero-salience hubs (once, not per day)
            if day == WEEKDAYS[0]:
                zero_hubs = [h for h, s in day_sal.items() if s == 0.0]

            total       = int(float(r.get(day, 0) or 0))
            day_alloc[day] = _weighted_alloc(total, day_sal)

        if zero_hubs:
            zero_sal_info[city] = zero_hubs

        for hub in hubs:
            hub_row = {
                "city_name":    city,
                "hub_name":     hub,
                "product_id":   str(r.get("product_id", "")).strip(),
                "product_name": str(r.get("product_name", "")).strip(),
                "category":     category,
                "MRP":          r.get("MRP", ""),
            }
            for day in WEEKDAYS:
                hub_row[day] = day_alloc[day].get(hub, 0)
            rows.append(hub_row)

    cols = ["city_name", "hub_name", "product_id", "product_name", "category", "MRP"] + WEEKDAYS
    return (pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols), zero_sal_info)


# ──────────────────────────────────────────────────────────────────
# EXCEL  TEMPLATE  BUILDERS
# ──────────────────────────────────────────────────────────────────
CITY_COLS = ["city_name", "product_id", "product_name", "category", "MRP"] + WEEKDAYS
HUB_COLS  = ["city_name", "hub_name", "product_id", "product_name", "category", "MRP"] + WEEKDAYS


def _style_ws(ws):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        blue = PatternFill("solid", fgColor="1A73E8")
        grey = PatternFill("solid", fgColor="F8FAFC")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = blue
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                if i % 2 == 0:
                    cell.fill = grey
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max(13, min(mx + 3, 40))
        ws.row_dimensions[1].height = 20
    except ImportError:
        pass


def build_city_template(cities: list, category: str,
                         product_id: str = "", product_name: str = "") -> bytes:
    """One blank row per city; product_id / product_name / MRP left blank for user to fill."""
    rows = [
        {"city_name": c, "product_id": product_id,
         "product_name": product_name, "category": category,
         "MRP": "", **{d: 0 for d in WEEKDAYS}}
        for c in cities
    ]
    df  = pd.DataFrame(rows, columns=CITY_COLS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="City Plan")
        _style_ws(w.sheets["City Plan"])
    buf.seek(0)
    return buf.getvalue()


def build_hub_template(cities_hubs: dict, category: str,
                        product_id: str = "", product_name: str = "") -> bytes:
    """cities_hubs = {city: [hub, ...]}
    Rows pre-filled with city_name / hub_name / category; product_id, product_name, MRP left blank."""
    rows = [
        {"city_name": city, "hub_name": hub,
         "product_id": product_id, "product_name": product_name,
         "category": category, "MRP": "", **{d: 0 for d in WEEKDAYS}}
        for city, hubs in cities_hubs.items()
        for hub in sorted(hubs)
    ]
    df  = pd.DataFrame(rows, columns=HUB_COLS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Hub Plan")
        _style_ws(w.sheets["Hub Plan"])
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# EXCEL  UPLOAD  PARSERS
# ──────────────────────────────────────────────────────────────────
def _safe_int(val) -> int:
    try:
        return int(round(float(val))) if pd.notna(val) else 0
    except Exception:
        return 0


def parse_city_upload(file) -> tuple:
    """Returns (df, errors).  df columns = CITY_COLS."""
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return pd.DataFrame(), [f"Cannot read file: {e}"]
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in ["city_name", "product_id"] + WEEKDAYS if c not in df.columns]
    if missing:
        return pd.DataFrame(), [f"Missing columns: {missing}"]
    errors = []
    try:
        df = CITY_UPLOAD_SCHEMA.validate(df)
    except Exception as err:
        return pd.DataFrame(), [f"Validation failed: {err}"]

    for day in WEEKDAYS:
        df[day] = df[day].apply(_safe_int)
    df["city_name"]    = df["city_name"].astype(str).str.strip()
    df["product_id"]   = df["product_id"].astype(str).str.strip()
    df["product_name"] = df.get("product_name", "").astype(str).str.strip()
    df["category"]     = df.get("category", "").astype(str).str.strip()
    df["MRP"]          = pd.to_numeric(df.get("MRP", 0), errors="coerce").fillna(0)
    df = df[df["city_name"].str.lower() != "nan"]
    if df.empty:
        errors.append("No valid data rows found.")
    return df[CITY_COLS], errors


def parse_hub_upload(file) -> tuple:
    """Returns (df, errors).  df columns = HUB_COLS."""
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return pd.DataFrame(), [f"Cannot read file: {e}"]
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in ["city_name", "hub_name", "product_id"] + WEEKDAYS if c not in df.columns]
    if missing:
        return pd.DataFrame(), [f"Missing columns: {missing}"]
    errors = []
    try:
        df = HUB_UPLOAD_SCHEMA.validate(df)
    except Exception as err:
        return pd.DataFrame(), [f"Validation failed: {err}"]

    for day in WEEKDAYS:
        df[day] = df[day].apply(_safe_int)
    for col in ["city_name", "hub_name", "product_id", "product_name", "category"]:
        df[col] = df.get(col, "").astype(str).str.strip()
    df["MRP"] = pd.to_numeric(df.get("MRP", 0), errors="coerce").fillna(0)
    df = df[df["city_name"].str.lower() != "nan"]
    if df.empty:
        errors.append("No valid data rows found.")
    return df[HUB_COLS], errors


# ──────────────────────────────────────────────────────────────────
# SUBMISSION  LOG  (Google Sheet)
# ──────────────────────────────────────────────────────────────────
def _ensure_log():
    client = _get_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    try:
        ws = sh.worksheet(LOG_SHEET_NAME)
        existing = ws.row_values(1)
        for col in ["Status", "Rejection_Reason", "Submitted_By", "MRP"]:
            if col not in existing:
                ws.update_cell(1, len(existing) + 1, col)
                existing.append(col)
    except Exception:
        ws = sh.add_worksheet(LOG_SHEET_NAME, rows="5000", cols="50")
        ws.append_row(LOG_HEADERS)


def load_log() -> pd.DataFrame:
    _ensure_log()
    sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
    data  = sheet.get_all_values()
    if len(data) <= 1:
        return pd.DataFrame(columns=LOG_HEADERS)
    df = pd.DataFrame(data[1:], columns=data[0])
    for col, default in [("Status", "Pending"), ("Rejection_Reason", ""),
                          ("Submitted_By", ""), ("MRP", "")]:
        if col not in df.columns:
            df[col] = default
    return df


def _sanitize(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().replace([np.nan, np.inf, -np.inf], "")
    for c in df.columns:
        df[c] = df[c].apply(lambda x: "" if pd.isna(x) else x)
    return df


def _overwrite_rows(sheet_id: str, sheet_name: str,
                    df_new: pd.DataFrame, key_cols: list):
    client = _get_client()
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet(sheet_name)
    data = ws.get_all_values()
    if len(data) <= 1:
        ws.clear()
        ws.append_row(df_new.columns.tolist())
        ws.append_rows(df_new.values.tolist())
        return
    headers = data[0]
    df_old = pd.DataFrame(data[1:], columns=headers)
    missing = [k for k in key_cols if k not in df_old.columns]
    if missing:
        ws.clear()
        ws.append_row(df_new.columns.tolist())
        ws.append_rows(df_new.values.tolist())
        return
    merged = df_old.merge(df_new[key_cols], on=key_cols, how="left", indicator=True)
    keep = merged[merged["_merge"] == "left_only"][headers]
    final_df = pd.concat([keep, df_new], ignore_index=True)
    ws.clear()
    ws.append_row(headers)
    ws.append_rows(final_df.values.tolist())


def save_to_log(rows_df: pd.DataFrame):
    """Append or overwrite rows in Submission_Log."""
    _overwrite_rows(
        SPREADSHEET_ID, LOG_SHEET_NAME,
        _sanitize(rows_df),
        key_cols=["Submission_Type", "Product ID", "City", "Hub"],
    )


def update_submission_status(sub_id: str, status: str, reason: str = ""):
    sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
    data = sheet.get_all_values()
    if len(data) <= 1:
        return
    headers = data[0]
    sid_idx = headers.index("Submission_ID") + 1 if "Submission_ID" in headers else None
    stat_idx = headers.index("Status") + 1 if "Status" in headers else None
    reason_idx = headers.index("Rejection_Reason") + 1 if "Rejection_Reason" in headers else None
    if not sid_idx:
        return
    batch_updates = []
    for i, row in enumerate(data[1:], start=2):
        if len(row) >= sid_idx and row[sid_idx - 1] == sub_id:
            if stat_idx:
                batch_updates.append({
                    "range": f"{_col_letter(stat_idx)}{i}",
                    "values": [[status]],
                })
            if reason_idx and reason:
                batch_updates.append({
                    "range": f"{_col_letter(reason_idx)}{i}",
                    "values": [[reason]],
                })
    if batch_updates:
        sheet.batch_update(batch_updates)


def _col_letter(col_idx: int) -> str:
    """1-based column index to A1 letter(s)."""
    letters = ""
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def gen_sub_id(sub_type: str) -> str:
    prefix = {"New Launch": "NEW", "Expansion": "EXP", "Replacement": "REP"}.get(sub_type, "UNK")
    return f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"


# ──────────────────────────────────────────────────────────────────
# DUPLICATE  DETECTION
# ──────────────────────────────────────────────────────────────────
def check_duplicates_city(df_log: pd.DataFrame, sub_type: str,
                           product_id: str, cities: list) -> pd.DataFrame:
    """Check city + product_id duplicates (after city-level upload)."""
    if df_log.empty:
        return pd.DataFrame()
    active = df_log[~df_log.get("Status", pd.Series()).isin(["Withdrawn", "Voided", "Expired"])]
    mask = (
        (active["Submission_Type"] == sub_type) &
        (active["Product ID"].astype(str) == str(product_id)) &
        (active["City"].isin(cities))
    )
    return active[mask]


def check_duplicates_hub(df_log: pd.DataFrame, sub_type: str,
                          hub_df: pd.DataFrame) -> pd.DataFrame:
    """Check city + hub + product_id duplicates (after hub-level upload) using vectorized inner merge."""
    if df_log.empty or hub_df.empty:
        return pd.DataFrame()
    active = df_log[~df_log.get("Status", pd.Series()).isin(["Withdrawn", "Voided", "Expired"])].copy()
    if active.empty:
        return pd.DataFrame()

    # Standardize types and strings for key columns to ensure robust matches
    active["Product ID"] = active["Product ID"].astype(str).str.strip()
    active["City"] = active["City"].astype(str).str.strip()
    active["Hub"] = active["Hub"].astype(str).str.strip()
    active["Submission_Type"] = active["Submission_Type"].astype(str).str.strip()

    temp_hub = hub_df.copy()
    temp_hub["Product ID"] = temp_hub["product_id"].astype(str).str.strip()
    temp_hub["City"] = temp_hub["city_name"].astype(str).str.strip()
    temp_hub["Hub"] = temp_hub["hub_name"].astype(str).str.strip()
    temp_hub["Submission_Type"] = str(sub_type).strip()

    merged = pd.merge(
        active,
        temp_hub[["Product ID", "City", "Hub", "Submission_Type"]],
        on=["Product ID", "City", "Hub", "Submission_Type"],
        how="inner"
    )
    return merged.drop_duplicates()


# ──────────────────────────────────────────────────────────────────
# DATE  HELPERS
# ──────────────────────────────────────────────────────────────────
def get_earliest_monday(min_days: int = 4) -> date:
    earliest = date.today() + timedelta(days=min_days)
    offset   = (7 - earliest.weekday()) % 7
    return earliest if earliest.weekday() == 0 else earliest + timedelta(days=offset)


# ──────────────────────────────────────────────────────────────────
# EMAIL  PLACEHOLDER
# ──────────────────────────────────────────────────────────────────
def show_email_placeholder(sub_id: str, sub_type: str, product_name: str,
                            hub_df: pd.DataFrame, *, user_id: int | None = None):
    """Send launch notification emails and show delivery status."""
    dates = []
    if "Launch Date" in hub_df.columns:
        dates = sorted(hub_df["Launch Date"].dropna().astype(str).unique().tolist())
    elif "Start Date" in hub_df.columns:
        dates = sorted(hub_df["Start Date"].dropna().astype(str).unique().tolist())
    launch_str = ", ".join(dates) if dates else "—"

    if user_id is None:
        user_id = os.environ.get("user", {}).get("id")

    from planning_suite.services.workflow_notifications import notify_launch_submission

    results = notify_launch_submission(
        sub_id=sub_id,
        sub_type=sub_type,
        product_name=product_name,
        launch_dates=dates,
        user_id=user_id,
    )

    if results.get("skipped"):
        if True:
            print(results.get("reason", "Email notifications were not sent."))
        return

    if True:
        col1, col2 = st.columns(2)
        with col1:
            print(f"""
**To: Planners**
> Subject: New {sub_type} — {product_name}
>
> Launch date(s): **{launch_str}**
> Submission ID: `{sub_id}`
""")
            _show_send_result(results.get("planner", {}))
        with col2:
            print(f"""
**To: Admin**
> Subject: Approval required — {sub_type}: {product_name}
>
> Submission ID: `{sub_id}`
> Launch date(s): **{launch_str}**
""")
            _show_send_result(results.get("admin", {}))


def _show_send_result(result: dict) -> None:
    status = result.get("status", "skipped")
    recipients = result.get("recipients") or []
    if status == "sent":
        st.success(f"Sent to: {', '.join(recipients)}")
    elif status == "failed":
        print(f"Failed: {result.get('error', 'Unknown error')}")
    else:
        print(result.get("error") or "Not sent — check SMTP and recipient list in Settings.")


# ──────────────────────────────────────────────────────────────────
# WIDE → LONG  CONVERTER
# ──────────────────────────────────────────────────────────────────
def wide_to_long(df: pd.DataFrame) -> pd.DataFrame:
    id_cols  = [c for c in df.columns if c not in WEEKDAYS]
    day_cols = [c for c in WEEKDAYS if c in df.columns]
    if not day_cols:
        return df
    long_df = df.melt(id_vars=id_cols, value_vars=day_cols,
                      var_name="Day", value_name="Plan")
    return long_df.reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────
# SHARED:  SUBMIT  HUB-LEVEL  DATA  TO  SHEETS
# ──────────────────────────────────────────────────────────────────
def _submit_hub_df(hub_df: pd.DataFrame, sub_type: str) -> str:
    """
    Normalise, attach metadata, write to Submission_Log + Launch_Output.
    Launch Date is read per-row from the 'Launch Date' column in hub_df.
    Returns the generated Submission_ID.
    """
    df = hub_df.copy()

    # Rename to sheet canonical column names
    df = df.rename(columns={
        "city_name":    "City",
        "hub_name":     "Hub",
        "product_id":   "Product ID",
        "product_name": "Product Name",
        "category":     "Category",
        "Launch Date":  "Start Date",
    })

    # Convert Start Date to string (handles date objects and strings)
    if "Start Date" in df.columns:
        df["Start Date"] = df["Start Date"].apply(
            lambda x: x.strftime("%Y-%m-%d") if hasattr(x, "strftime") else str(x)
        )
    else:
        df["Start Date"] = ""

    for day in WEEKDAYS:
        if day not in df.columns:
            df[day] = 0
    df[WEEKDAYS] = df[WEEKDAYS].fillna(0).astype(int)

    user   = os.environ.get("user", {})
    sub_id = gen_sub_id(sub_type)
    df["Timestamp"]        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df["Submission_ID"]    = sub_id
    df["Submission_Type"]  = sub_type
    df["Status"]           = "Pending"
    df["Rejection_Reason"] = ""
    df["Submitted_By"]     = user.get("username", "")

    log_cols = ["Timestamp", "Submission_ID", "Submission_Type",
                "Product ID", "Product Name", "Category",
                "City", "Hub", "MRP", "Start Date",
                "Status", "Rejection_Reason", "Submitted_By"] + WEEKDAYS
    log_df = df[[c for c in log_cols if c in df.columns]]
    save_to_log(_sanitize(log_df))

    # Launch_Output: long format
    out_wide = df[["Product ID", "Product Name", "Category",
                   "City", "Hub", "Start Date"] + WEEKDAYS]
    out_long = wide_to_long(out_wide)
    _overwrite_rows(SPREADSHEET_ID, "Launch_Output",
                    _sanitize(out_long),
                    key_cols=["Product ID", "City", "Hub", "Day"])
    return sub_id


# ──────────────────────────────────────────────────────────────────
# UI  HELPERS
# ──────────────────────────────────────────────────────────────────
def _stage_bar(stage: str, stages: list):
    """Simple text progress breadcrumb."""
    parts = []
    for s in stages:
        if s == stage:
            parts.append(f"**→ {s}**")
        else:
            parts.append(s)
    print("  ›  ".join(parts))
    print("---")


def _show_zero_sal_warning(zero_sal_info: dict):
    if zero_sal_info:
        hubs_list = ", ".join(
            f"{city}: {', '.join(hubs)}" for city, hubs in zero_sal_info.items()
        )
        print(
            f"⚠️ These hubs have **no salience data** for this category and received "
            f"an **equal split**: {hubs_list}. You can edit the values in the table below."
        )


def _show_diff(existing: pd.DataFrame, new_df: pd.DataFrame, label_existing: str = "Existing"):
    print(f"#### Duplicate detected — comparing **{label_existing}** vs **Your Upload**")
    c1, c2 = st.columns(2)
    show = ["City", "Hub"] + WEEKDAYS
    with c1:
        st.caption(label_existing)
        st.dataframe(existing[[c for c in show if c in existing.columns]],
                     use_container_width=True, hide_index=True)
    with c2:
        st.caption("Your Upload")
        rename_new = new_df.rename(columns={"city_name": "City", "hub_name": "Hub"})
        st.dataframe(rename_new[[c for c in show if c in rename_new.columns]],
                     use_container_width=True, hide_index=True)


# ──────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────
# PAGE  TYPE 1  —  NEW  PRODUCT  LAUNCH
# ─────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────
_T1_STAGES = ["1 · Upload", "2 · Hub Split", "3 · Launch Date", "4 · Confirm & Submit"]
_T1_KEY    = "t1"   # session-state prefix


def _t1_key(k): return f"{_T1_KEY}_{k}"


def _t1_reset():
    for k in [k for k in os.environ if k.startswith(f"{_T1_KEY}_")]:
        del os.environ[k]


def _t1_init():
    defs = {
        "stage": "upload",
        "plan_level": "city",
        "product_id": "", "product_name": "", "category": "",
        "selected_cities": [], "selected_hubs": {},
        "upload_city_df": pd.DataFrame(),
        "hub_split_df": pd.DataFrame(),
        "zero_sal_info": {},
        "dup_existing": pd.DataFrame(),
        "launch_date": None,
        "is_edit": False,
        "edit_sub_id": "",
        "upload_ts": "",
    }
    for k, v in defs.items():
        if _t1_key(k) not in os.environ:
            os.environ[_t1_key(k)] = v


def page_type1():
    print(
        '<div class="page-header">'
        '<div class="page-title">New Product Launch</div>'
        '<div class="page-desc">Completely new product ID entering the system</div>'
        '</div>', unsafe_allow_html=True
    )
    _t1_init()

    sal_df    = load_hub_salience()
    df_master = load_product_master()
    all_cats  = get_categories(df_master)
    all_cities = get_cities_from_salience(sal_df)
    if not _require_salience(sal_df):
        return

    stage = os.environ[_t1_key("stage")]
    _stage_bar(stage.replace("upload", "1 · Upload")
                     .replace("split_review", "2 · Hub Split")
                     .replace("set_date", "3 · Launch Date")
                     .replace("confirm", "4 · Confirm & Submit"),
               _T1_STAGES)

    # ── STAGE: upload ────────────────────────────────────────────
    if stage == "upload":
        _t1_stage_upload(df_master, all_cats, all_cities, sal_df)

    # ── STAGE: split_review ─────────────────────────────────────
    elif stage == "split_review":
        _t1_stage_split()

    # ── STAGE: set_date ─────────────────────────────────────────
    elif stage == "set_date":
        _t1_stage_date()

    # ── STAGE: confirm ──────────────────────────────────────────
    elif stage == "confirm":
        _t1_stage_confirm()

    # Reset button
    print("---")
    if st.button("↺ Start new submission", key="t1_reset_btn"):
        _t1_reset()
        st.rerun()


def _t1_stage_upload(df_master, all_cats, all_cities, sal_df):
    # ── Sub-Category only ────────────────────────────────────────
    cat = st.selectbox("Sub-Category", all_cats, key="t1_cat_sel",
                       index=all_cats.index(os.environ[_t1_key("category")])
                       if os.environ[_t1_key("category")] in all_cats else 0)
    os.environ[_t1_key("category")] = cat

    print("---")

    # ── Plan level + cities/hubs ─────────────────────────────────
    plan_level = st.radio("Plan Level", ["City Level", "Hub Level"],
                          horizontal=True, key="t1_plan_level_sel")
    os.environ[_t1_key("plan_level")] = plan_level

    sel_cities = st.multiselect("Select Cities", all_cities,
                                 default=os.environ[_t1_key("selected_cities")],
                                 key="t1_cities_sel")
    os.environ[_t1_key("selected_cities")] = sel_cities

    sel_hubs = {}
    if plan_level == "Hub Level" and sel_cities:
        st.caption("Select hubs per city (these will be pre-filled in the Hub-Level template):")
        for city in sel_cities:
            # No category filter — show all hubs available for the city
            city_hubs = get_hubs_for_city(sal_df, city)
            chosen = st.multiselect(
                f"{city} — hubs",
                city_hubs,
                default=[h for h in os.environ[_t1_key("selected_hubs")].get(city, [])
                         if h in city_hubs],
                key=f"t1_hubs_{city}",
            )
            sel_hubs[city] = chosen
    os.environ[_t1_key("selected_hubs")] = sel_hubs

    print("---")

    # ── Template download (show only the relevant template) ──────
    if sel_cities:
        print("#### Download Template")
        if plan_level == "City Level":
            city_bytes = build_city_template(sel_cities, cat)
            st.download_button(
                "⬇ City-Level Template",
                data=city_bytes,
                file_name=f"city_template_{cat}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t1_dl_city", use_container_width=True,
            )
            st.caption("Pre-filled: city_name · category   |   Fill in: product_id · product_name · MRP · Mon–Sun")
        else:
            hubs_for_tmpl = sel_hubs if any(sel_hubs.values()) \
                else {c: get_hubs_for_city(sal_df, c) for c in sel_cities}
            hub_bytes = build_hub_template(hubs_for_tmpl, cat)
            st.download_button(
                "⬇ Hub-Level Template",
                data=hub_bytes,
                file_name=f"hub_template_{cat}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t1_dl_hub", use_container_width=True,
            )
            st.caption("Pre-filled: city_name · hub_name · category   |   Fill in: product_id · product_name · MRP · Mon–Sun")

    # ── Upload ───────────────────────────────────────────────────
    print("#### Upload Filled Template")
    uploaded = st.file_uploader("Choose .xlsx file", type=["xlsx"],
                                 key="t1_uploader", label_visibility="collapsed")

    if uploaded and sel_cities:
        # Auto-detect type from plan level selection; no manual radio needed
        up_type = "City-Level" if plan_level == "City Level" else "Hub-Level"

        if st.button("Process Upload →", type="primary", key="t1_process_btn"):
            if up_type == "City-Level":
                df_up, errs = parse_city_upload(uploaded)
            else:
                df_up, errs = parse_hub_upload(uploaded)

            if errs:
                for e in errs:
                    print(e)
                return
            if df_up.empty:
                print("No valid rows parsed from the file.")
                return

            # ── Show upload timestamp ────────────────────────────
            import datetime as _dt
            os.environ[_t1_key("upload_ts")] = _dt.datetime.now().strftime("%d %b %Y  %H:%M:%S")

            # ── Duplicate check (use product_id from uploaded file) ──
            pid_from_file = df_up["product_id"].iloc[0] if "product_id" in df_up.columns else ""
            df_log = load_log()
            if up_type == "City-Level":
                cities_in_file = df_up["city_name"].unique().tolist()
                dup = check_duplicates_city(df_log, "New Launch", pid_from_file, cities_in_file)
            else:
                dup = check_duplicates_hub(df_log, "New Launch", df_up)

            if not dup.empty:
                _show_diff(dup, df_up, "Existing Submission")
                print("⚠️ A previous submission exists for this product + city/hub combination. "
                           "The table below shows the existing data. Edit and re-submit to overwrite.")
                os.environ[_t1_key("dup_existing")]  = dup
                os.environ[_t1_key("is_edit")]       = True
                os.environ[_t1_key("edit_sub_id")]   = dup["Submission_ID"].iloc[0]

            if up_type == "City-Level":
                hub_df, zero_info = split_city_to_hubs(df_up, sal_df, forced_hubs=None)
                if hub_df.empty:
                    print("Could not generate hub split. Check salience data for selected category.")
                    return
                os.environ[_t1_key("upload_city_df")] = df_up
                os.environ[_t1_key("hub_split_df")]   = hub_df
                os.environ[_t1_key("zero_sal_info")]  = zero_info
                os.environ[_t1_key("stage")]          = "split_review"
            else:
                os.environ[_t1_key("hub_split_df")]  = df_up
                os.environ[_t1_key("zero_sal_info")] = {}
                os.environ[_t1_key("stage")]         = "set_date"
            st.rerun()

    # ── Show upload timestamp if available ───────────────────────
    if os.environ.get(_t1_key("upload_ts")):
        st.caption(f"Last uploaded: {os.environ[_t1_key('upload_ts')]}")


def _t1_stage_split():
    st.subheader("Hub-Level Plan Split (from city upload)")
    zero_info = os.environ[_t1_key("zero_sal_info")]
    _show_zero_sal_warning(zero_info)

    st.caption("Review the auto-generated hub split. You can edit values directly in the table.")
    edited = st.data_editor(
        os.environ[_t1_key("hub_split_df")],
        num_rows="dynamic", use_container_width=True, key="t1_split_editor",
        column_config={
            "city_name":  st.column_config.TextColumn("City",    disabled=True),
            "hub_name":   st.column_config.TextColumn("Hub",     disabled=True),
            "product_id": st.column_config.TextColumn("Prod ID", disabled=True),
        }
    )
    os.environ[_t1_key("hub_split_df")] = edited

    col_back, col_sync, col_next = st.columns([1, 2, 4])
    with col_back:
        if st.button("← Back", key="t1_split_back"):
            os.environ[_t1_key("stage")] = "upload"
            st.rerun()
    with col_next:
        if st.button("Confirm Split & Set Launch Date →", type="primary", key="t1_split_next"):
            os.environ[_t1_key("stage")] = "set_date"
            st.rerun()


def _t1_stage_date():
    st.subheader("Set Launch Dates")
    min_date = date.today() + timedelta(days=4)
    print(
        f"Minimum launch date: **{min_date.strftime('%d %b %Y')}** (today + 4 days). "
        "Each row can have a different launch date — edit the **Launch Date** column directly."
    )

    hub_df = os.environ[_t1_key("hub_split_df")].copy()
    if "Launch Date" not in hub_df.columns:
        hub_df["Launch Date"] = min_date

    display_cols = ["city_name", "hub_name", "product_id"] + WEEKDAYS + ["Launch Date"]
    display_cols = [c for c in display_cols if c in hub_df.columns]

    edited = st.data_editor(
        hub_df[display_cols],
        num_rows="fixed",
        use_container_width=True,
        key="t1_date_editor",
        column_config={
            "city_name":    st.column_config.TextColumn("City",       disabled=True),
            "hub_name":     st.column_config.TextColumn("Hub",        disabled=True),
            "product_id":   st.column_config.TextColumn("Product ID", disabled=True),
            **{d: st.column_config.NumberColumn(d, disabled=True) for d in WEEKDAYS if d in hub_df.columns},
            "Launch Date":  st.column_config.DateColumn(
                "Launch Date",
                min_value=min_date,
                format="DD-MMM-YYYY",
            ),
        },
    )

    errors = []
    for _, row in edited.iterrows():
        d = row.get("Launch Date")
        if d is None:
            errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: Launch Date is missing.")
        else:
            d_obj = d if isinstance(d, date) else pd.to_datetime(d).date()
            if d_obj < min_date:
                errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: {d_obj} must be ≥ {min_date} (T+4).")
    for e in errors:
        print(e)

    col_back, col_next = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t1_date_back"):
            os.environ[_t1_key("stage")] = (
                "split_review"
                if not os.environ[_t1_key("upload_city_df")].empty
                else "upload"
            )
            st.rerun()
    with col_next:
        if st.button("Review & Submit →", type="primary", key="t1_date_next", disabled=bool(errors)):
            merged = hub_df.copy()
            merged["Launch Date"] = edited["Launch Date"].values
            os.environ[_t1_key("hub_split_df")] = merged
            os.environ[_t1_key("stage")] = "confirm"
            st.rerun()


def _t1_stage_confirm():
    st.subheader("Review & Submit")
    hub_df  = os.environ[_t1_key("hub_split_df")]
    is_edit = os.environ[_t1_key("is_edit")]
    pid     = os.environ[_t1_key("product_id")]
    pname   = os.environ[_t1_key("product_name")]

    print(f"**Product:** `{pid}` — {pname}  \n**Rows:** {len(hub_df)} hub entries")
    if is_edit:
        print("This will **overwrite** the existing submission.")
    st.dataframe(hub_df, use_container_width=True, hide_index=True)

    col_back, col_submit = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t1_confirm_back"):
            os.environ[_t1_key("stage")] = "set_date"
            st.rerun()
    with col_submit:
        if st.button("Submit →", type="primary", key="t1_submit_btn"):
            with st.spinner("Saving…"):
                sub_id = _submit_hub_df(hub_df, "New Launch")
            st.success(f"Submitted! Submission ID: **{sub_id}**")
            show_email_placeholder(sub_id, "New Launch", pname, hub_df)
            os.environ[_t1_key("stage")] = "upload"
            _t1_reset()


# ──────────────────────────────────────────────────────────────────
# PAGE  TYPE 2  —  PRODUCT  EXPANSION
# ──────────────────────────────────────────────────────────────────
_T2_KEY = "t2"


def _t2_key(k): return f"{_T2_KEY}_{k}"


def _t2_reset():
    for k in [k for k in os.environ if k.startswith(f"{_T2_KEY}_")]:
        del os.environ[k]


def _t2_init():
    defs = {
        "stage": "upload", "plan_level": "city",
        "product_id": "", "product_name": "", "category": "",
        "selected_cities": [], "selected_hubs": {},
        "upload_city_df": pd.DataFrame(),
        "hub_split_df": pd.DataFrame(), "zero_sal_info": {},
        "dup_existing": pd.DataFrame(), "launch_date": None,
        "is_edit": False, "edit_sub_id": "",
        "upload_ts": "",
    }
    for k, v in defs.items():
        if _t2_key(k) not in os.environ:
            os.environ[_t2_key(k)] = v


def page_type2():
    print(
        '<div class="page-header">'
        '<div class="page-title">Product Expansion</div>'
        '<div class="page-desc">Existing product ID — expanding into new cities / hubs</div>'
        '</div>', unsafe_allow_html=True
    )
    _t2_init()
    sal_df     = load_hub_salience()
    df_master  = load_product_master()
    all_cities = get_cities_from_salience(sal_df)
    if not _require_salience(sal_df):
        return

    stage = os.environ[_t2_key("stage")]
    _stage_bar(stage.replace("upload", "1 · Upload")
                     .replace("split_review", "2 · Hub Split")
                     .replace("set_date", "3 · Launch Date")
                     .replace("confirm", "4 · Confirm & Submit"),
               _T1_STAGES)

    if stage == "upload":
        _t2_stage_upload(df_master, all_cities, sal_df)
    elif stage == "split_review":
        _t2_stage_split()
    elif stage == "set_date":
        _t2_stage_date()
    elif stage == "confirm":
        _t2_stage_confirm()

    print("---")
    if st.button("↺ Start new submission", key="t2_reset_btn"):
        _t2_reset()
        st.rerun()


def _t2_stage_upload(df_master, all_cities, sal_df):
    pid_col   = next((c for c in ["Product id", "Product ID"] if c in df_master.columns), None)
    all_pids  = sorted(df_master[pid_col].astype(str).unique().tolist()) if pid_col else []

    c1, c2 = st.columns(2)
    with c1:
        sel_pid = st.selectbox("Existing Product ID", all_pids, key="t2_pid_sel",
                               index=all_pids.index(os.environ[_t2_key("product_id")])
                               if os.environ[_t2_key("product_id")] in all_pids else 0)
    info  = get_product_info(df_master, sel_pid)
    with c2:
        st.text_input("Product Name", value=info["name"], disabled=True, key="t2_pname_disp")
    st.caption(f"Category: **{info['category']}**")
    os.environ[_t2_key("product_id")]   = sel_pid
    os.environ[_t2_key("product_name")] = info["name"]
    os.environ[_t2_key("category")]     = info["category"]

    print("---")
    plan_level = st.radio("Plan Level", ["City Level", "Hub Level"],
                          horizontal=True, key="t2_plan_level_sel")
    os.environ[_t2_key("plan_level")] = plan_level
    sel_cities = st.multiselect("Select Cities (new cities for this product)", all_cities,
                                 default=os.environ[_t2_key("selected_cities")],
                                 key="t2_cities_sel")
    os.environ[_t2_key("selected_cities")] = sel_cities

    sel_hubs = {}
    if plan_level == "Hub Level" and sel_cities:
        st.caption("Select expansion hubs per city:")
        for city in sel_cities:
            # No category filter — show all hubs available for the city
            city_hubs = get_hubs_for_city(sal_df, city)
            chosen = st.multiselect(
                f"{city} — hubs", city_hubs,
                default=[h for h in os.environ[_t2_key("selected_hubs")].get(city, [])
                         if h in city_hubs],
                key=f"t2_hubs_{city}",
            )
            sel_hubs[city] = chosen
    os.environ[_t2_key("selected_hubs")] = sel_hubs

    print("---")
    if sel_cities:
        print("#### Download Template")
        if plan_level == "City Level":
            st.download_button(
                "⬇ City-Level Template",
                data=build_city_template(sel_cities, info["category"], sel_pid, info["name"]),
                file_name=f"city_expansion_{sel_pid}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t2_dl_city", use_container_width=True,
            )
            st.caption("Pre-filled: city_name · product_id · product_name · category   |   Fill in: MRP · Mon–Sun")
        else:
            hubs_for_tmpl = sel_hubs if any(sel_hubs.values()) \
                else {c: get_hubs_for_city(sal_df, c) for c in sel_cities}
            st.download_button(
                "⬇ Hub-Level Template",
                data=build_hub_template(hubs_for_tmpl, info["category"], sel_pid, info["name"]),
                file_name=f"hub_expansion_{sel_pid}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t2_dl_hub", use_container_width=True,
            )
            st.caption("Pre-filled: city_name · hub_name · product_id · product_name · category   |   Fill in: MRP · Mon–Sun")

    print("#### Upload Filled Template")
    uploaded = st.file_uploader("Choose .xlsx file", type=["xlsx"],
                                 key="t2_uploader", label_visibility="collapsed")
    if uploaded and sel_cities and sel_pid:
        # Auto-detect type from plan level selection
        up_type = "City-Level" if plan_level == "City Level" else "Hub-Level"
        if st.button("Process Upload →", type="primary", key="t2_process_btn"):
            df_up, errs = parse_city_upload(uploaded) if up_type == "City-Level" \
                else parse_hub_upload(uploaded)
            if errs:
                for e in errs: print(e)
                return
            if df_up.empty:
                print("No valid rows parsed.")
                return

            # ── Show upload timestamp ────────────────────────────
            import datetime as _dt
            os.environ[_t2_key("upload_ts")] = _dt.datetime.now().strftime("%d %b %Y  %H:%M:%S")

            df_log = load_log()
            dup = check_duplicates_city(df_log, "Expansion", sel_pid,
                                         df_up["city_name"].unique().tolist()) \
                  if up_type == "City-Level" \
                  else check_duplicates_hub(df_log, "Expansion", df_up)
            if not dup.empty:
                _show_diff(dup, df_up)
                print("Previous submission found — editing mode.")
                os.environ[_t2_key("dup_existing")] = dup
                os.environ[_t2_key("is_edit")]      = True
            if up_type == "City-Level":
                hub_df, zero_info = split_city_to_hubs(df_up, sal_df)
                if hub_df.empty:
                    print("No hub salience data. Check category.")
                    return
                os.environ[_t2_key("upload_city_df")] = df_up
                os.environ[_t2_key("hub_split_df")]   = hub_df
                os.environ[_t2_key("zero_sal_info")]  = zero_info
                os.environ[_t2_key("stage")]          = "split_review"
            else:
                os.environ[_t2_key("hub_split_df")]  = df_up
                os.environ[_t2_key("zero_sal_info")] = {}
                os.environ[_t2_key("stage")]         = "set_date"
            st.rerun()

    # ── Show upload timestamp if available ───────────────────────
    if os.environ.get(_t2_key("upload_ts")):
        st.caption(f"Last uploaded: {os.environ[_t2_key('upload_ts')]}")


def _t2_stage_split():
    st.subheader("Hub-Level Plan Split")
    _show_zero_sal_warning(os.environ[_t2_key("zero_sal_info")])
    edited = st.data_editor(
        os.environ[_t2_key("hub_split_df")],
        num_rows="dynamic", use_container_width=True, key="t2_split_editor",
        column_config={
            "city_name":  st.column_config.TextColumn("City",    disabled=True),
            "hub_name":   st.column_config.TextColumn("Hub",     disabled=True),
            "product_id": st.column_config.TextColumn("Prod ID", disabled=True),
        }
    )
    os.environ[_t2_key("hub_split_df")] = edited
    col_back, col_next = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t2_split_back"):
            os.environ[_t2_key("stage")] = "upload"
            st.rerun()
    with col_next:
        if st.button("Set Launch Date →", type="primary", key="t2_split_next"):
            os.environ[_t2_key("stage")] = "set_date"
            st.rerun()


def _t2_stage_date():
    st.subheader("Set Launch Dates")
    min_date = date.today() + timedelta(days=4)
    print(
        f"Minimum launch date: **{min_date.strftime('%d %b %Y')}** (today + 4 days). "
        "Each row can have a different launch date — edit the **Launch Date** column directly."
    )

    hub_df = os.environ[_t2_key("hub_split_df")].copy()
    if "Launch Date" not in hub_df.columns:
        hub_df["Launch Date"] = min_date

    display_cols = ["city_name", "hub_name", "product_id"] + WEEKDAYS + ["Launch Date"]
    display_cols = [c for c in display_cols if c in hub_df.columns]

    edited = st.data_editor(
        hub_df[display_cols],
        num_rows="fixed",
        use_container_width=True,
        key="t2_date_editor",
        column_config={
            "city_name":    st.column_config.TextColumn("City",       disabled=True),
            "hub_name":     st.column_config.TextColumn("Hub",        disabled=True),
            "product_id":   st.column_config.TextColumn("Product ID", disabled=True),
            **{d: st.column_config.NumberColumn(d, disabled=True) for d in WEEKDAYS if d in hub_df.columns},
            "Launch Date":  st.column_config.DateColumn(
                "Launch Date",
                min_value=min_date,
                format="DD-MMM-YYYY",
            ),
        },
    )

    errors = []
    for _, row in edited.iterrows():
        d = row.get("Launch Date")
        if d is None:
            errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: Launch Date is missing.")
        else:
            d_obj = d if isinstance(d, date) else pd.to_datetime(d).date()
            if d_obj < min_date:
                errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: {d_obj} must be ≥ {min_date} (T+4).")
    for e in errors:
        print(e)

    col_back, col_next = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t2_date_back"):
            os.environ[_t2_key("stage")] = "split_review" \
                if not os.environ[_t2_key("upload_city_df")].empty else "upload"
            st.rerun()
    with col_next:
        if st.button("Review & Submit →", type="primary", key="t2_date_next", disabled=bool(errors)):
            merged = hub_df.copy()
            merged["Launch Date"] = edited["Launch Date"].values
            os.environ[_t2_key("hub_split_df")] = merged
            os.environ[_t2_key("stage")] = "confirm"
            st.rerun()


def _t2_stage_confirm():
    st.subheader("Review & Submit")
    hub_df = os.environ[_t2_key("hub_split_df")]
    pid    = os.environ[_t2_key("product_id")]
    pname  = os.environ[_t2_key("product_name")]
    if os.environ[_t2_key("is_edit")]:
        print("This will **overwrite** the existing submission.")
    print(f"**Product:** `{pid}` — {pname}  \n**Rows:** {len(hub_df)} hub entries")
    st.dataframe(hub_df, use_container_width=True, hide_index=True)
    col_back, col_submit = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t2_confirm_back"):
            os.environ[_t2_key("stage")] = "set_date"
            st.rerun()
    with col_submit:
        if st.button("Submit →", type="primary", key="t2_submit_btn"):
            with st.spinner("Saving…"):
                sub_id = _submit_hub_df(hub_df, "Expansion")
            st.success(f"Expansion submitted! ID: **{sub_id}**")
            show_email_placeholder(sub_id, "Expansion", pname, hub_df)
            _t2_reset()


# ──────────────────────────────────────────────────────────────────
# PAGE  TYPE 3  —  PRODUCT  REPLACEMENT
# ──────────────────────────────────────────────────────────────────
_T3_KEY = "t3"


def _t3_key(k): return f"{_T3_KEY}_{k}"


def _t3_reset():
    for k in [k for k in os.environ if k.startswith(f"{_T3_KEY}_")]:
        del os.environ[k]


def _t3_init():
    defs = {
        "stage": "setup",
        "old_pid": "", "old_name": "",
        "new_pid": "", "new_name": "", "category": "",
        "split_pct": 100,
        "selected_cities": [],
        "hub_split_df": pd.DataFrame(),
        "upload_city_df": pd.DataFrame(),
        "zero_sal_info": {},
        "launch_date": None,
    }
    for k, v in defs.items():
        if _t3_key(k) not in os.environ:
            os.environ[_t3_key(k)] = v


def page_type3():
    print(
        '<div class="page-header">'
        '<div class="page-title">Product Replacement</div>'
        '<div class="page-desc">Replace an existing product with a new SKU</div>'
        '</div>', unsafe_allow_html=True
    )
    _t3_init()
    sal_df    = load_hub_salience()
    df_master = load_product_master()
    all_cats  = get_categories(df_master)
    all_cities = get_cities_from_salience(sal_df)
    if not _require_salience(sal_df):
        return
    stage = os.environ[_t3_key("stage")]

    stages_rep = ["1 · Old & New SKU", "2 · Plan Upload", "3 · Hub Split",
                  "4 · Launch Date", "5 · Confirm"]
    _stage_bar(stage.replace("setup", "1 · Old & New SKU")
                     .replace("upload", "2 · Plan Upload")
                     .replace("split_review", "3 · Hub Split")
                     .replace("set_date", "4 · Launch Date")
                     .replace("confirm", "5 · Confirm"),
               stages_rep)

    if stage == "setup":
        _t3_stage_setup(df_master, all_cats, all_cities)
    elif stage == "upload":
        _t3_stage_upload(sal_df)
    elif stage == "split_review":
        _t3_stage_split()
    elif stage == "set_date":
        _t3_stage_date()
    elif stage == "confirm":
        _t3_stage_confirm()

    print("---")
    if st.button("↺ Start new submission", key="t3_reset_btn"):
        _t3_reset()
        st.rerun()


def _t3_stage_setup(df_master, all_cats, all_cities):
    st.subheader("Step 1 — Old SKU & New SKU")
    c1, c2 = st.columns(2)
    with c1:
        print("**Old SKU (being replaced)**")
        old_cat  = st.selectbox("Old Category", all_cats, key="t3_old_cat")
        old_prods = get_products_by_category(df_master, old_cat)
        old_name  = st.selectbox("Old Product Name", old_prods, key="t3_old_name")
        old_pid   = get_product_id(df_master, old_name)
        st.text_input("Old Product ID", old_pid, disabled=True, key="t3_old_pid_disp")
    with c2:
        print("**New SKU (replacement)**")
        new_cat  = st.selectbox("New Category", all_cats, key="t3_new_cat")
        new_prods = get_products_by_category(df_master, new_cat)
        new_name  = st.selectbox("New Product Name", new_prods, key="t3_new_name")
        new_pid   = get_product_id(df_master, new_name)
        st.text_input("New Product ID", new_pid, disabled=True, key="t3_new_pid_disp")

    print("---")
    split_pct = st.slider("% Plan going to **New SKU**", 0, 100, 100, key="t3_split_pct")
    st.caption(f"New SKU: **{split_pct}%** · Old SKU: **{100 - split_pct}%**")

    sel_cities = st.multiselect("Select Cities", all_cities,
                                 default=os.environ[_t3_key("selected_cities")],
                                 key="t3_cities_sel")

    if st.button("Next: Upload Plan →", type="primary", key="t3_setup_next"):
        if not old_pid or not new_pid:
            print("Select both old and new products.")
        elif not sel_cities:
            print("Select at least one city.")
        else:
            os.environ[_t3_key("old_pid")]         = old_pid
            os.environ[_t3_key("old_name")]        = old_name
            os.environ[_t3_key("new_pid")]         = new_pid
            os.environ[_t3_key("new_name")]        = new_name
            os.environ[_t3_key("category")]        = new_cat
            os.environ[_t3_key("split_pct")]       = split_pct
            os.environ[_t3_key("selected_cities")] = sel_cities
            os.environ[_t3_key("stage")]           = "upload"
            st.rerun()


def _t3_stage_upload(sal_df):
    st.subheader("Step 2 — Upload Plan for New SKU")
    pid    = os.environ[_t3_key("new_pid")]
    pname  = os.environ[_t3_key("new_name")]
    cat    = os.environ[_t3_key("category")]
    cities = os.environ[_t3_key("selected_cities")]
    pct    = os.environ[_t3_key("split_pct")]

    print(f"Replacing **{os.environ[_t3_key('old_name')]}** "
            f"→ **{pname}** ({pct}% of plan goes to new SKU)")

    dc1, dc2 = st.columns(2)
    with dc1:
        st.download_button(
            "⬇ City-Level Template (New SKU)",
            data=build_city_template(cities, cat, pid, pname),
            file_name=f"replacement_city_{pid}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="t3_dl_city", use_container_width=True,
        )
    with dc2:
        all_hubs = {c: get_hubs_for_city(sal_df, c) for c in cities}
        st.download_button(
            "⬇ Hub-Level Template (New SKU)",
            data=build_hub_template(all_hubs, cat, pid, pname),
            file_name=f"replacement_hub_{pid}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="t3_dl_hub", use_container_width=True,
        )

    uploaded = st.file_uploader("Upload filled template (.xlsx)", type=["xlsx"],
                                 key="t3_uploader", label_visibility="collapsed")
    if uploaded:
        up_type = st.radio("Template type", ["City-Level", "Hub-Level"],
                           horizontal=True, key="t3_up_type")
        col_back, col_proc = st.columns([1, 6])
        with col_back:
            if st.button("← Back", key="t3_upload_back"):
                os.environ[_t3_key("stage")] = "setup"
                st.rerun()
        with col_proc:
            if st.button("Process →", type="primary", key="t3_proc_btn"):
                df_up, errs = parse_city_upload(uploaded) if up_type == "City-Level" \
                    else parse_hub_upload(uploaded)
                if errs:
                    for e in errs: print(e)
                    return
                # Apply split %
                for day in WEEKDAYS:
                    df_up[day] = (df_up[day] * pct / 100).round().astype(int)

                if up_type == "City-Level":
                    hub_df, zero_info = split_city_to_hubs(df_up, sal_df)
                    if hub_df.empty:
                        print("Could not generate hub split.")
                        return
                    os.environ[_t3_key("upload_city_df")] = df_up
                    os.environ[_t3_key("hub_split_df")]   = hub_df
                    os.environ[_t3_key("zero_sal_info")]  = zero_info
                    os.environ[_t3_key("stage")]          = "split_review"
                else:
                    os.environ[_t3_key("hub_split_df")]  = df_up
                    os.environ[_t3_key("zero_sal_info")] = {}
                    os.environ[_t3_key("stage")]         = "set_date"
                st.rerun()
    else:
        if st.button("← Back", key="t3_upload_back_nofile"):
            os.environ[_t3_key("stage")] = "setup"
            st.rerun()


def _t3_stage_split():
    st.subheader("Hub Split — New SKU")
    _show_zero_sal_warning(os.environ[_t3_key("zero_sal_info")])
    edited = st.data_editor(
        os.environ[_t3_key("hub_split_df")],
        num_rows="dynamic", use_container_width=True, key="t3_split_editor",
        column_config={
            "city_name":  st.column_config.TextColumn("City",    disabled=True),
            "hub_name":   st.column_config.TextColumn("Hub",     disabled=True),
            "product_id": st.column_config.TextColumn("Prod ID", disabled=True),
        }
    )
    os.environ[_t3_key("hub_split_df")] = edited
    col_back, col_next = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t3_split_back"):
            os.environ[_t3_key("stage")] = "upload"
            st.rerun()
    with col_next:
        if st.button("Set Launch Date →", type="primary", key="t3_split_next"):
            os.environ[_t3_key("stage")] = "set_date"
            st.rerun()


def _t3_stage_date():
    st.subheader("Set Launch Dates")
    min_date = date.today() + timedelta(days=4)
    print(
        f"Minimum launch date: **{min_date.strftime('%d %b %Y')}** (today + 4 days). "
        "Each row can have a different launch date — edit the **Launch Date** column directly."
    )

    hub_df = os.environ[_t3_key("hub_split_df")].copy()
    if "Launch Date" not in hub_df.columns:
        hub_df["Launch Date"] = min_date

    display_cols = ["city_name", "hub_name", "product_id"] + WEEKDAYS + ["Launch Date"]
    display_cols = [c for c in display_cols if c in hub_df.columns]

    edited = st.data_editor(
        hub_df[display_cols],
        num_rows="fixed",
        use_container_width=True,
        key="t3_date_editor",
        column_config={
            "city_name":    st.column_config.TextColumn("City",       disabled=True),
            "hub_name":     st.column_config.TextColumn("Hub",        disabled=True),
            "product_id":   st.column_config.TextColumn("Product ID", disabled=True),
            **{d: st.column_config.NumberColumn(d, disabled=True) for d in WEEKDAYS if d in hub_df.columns},
            "Launch Date":  st.column_config.DateColumn(
                "Launch Date",
                min_value=min_date,
                format="DD-MMM-YYYY",
            ),
        },
    )

    errors = []
    for _, row in edited.iterrows():
        d = row.get("Launch Date")
        if d is None:
            errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: Launch Date is missing.")
        else:
            d_obj = d if isinstance(d, date) else pd.to_datetime(d).date()
            if d_obj < min_date:
                errors.append(f"{row.get('city_name','')}/{row.get('hub_name','')}: {d_obj} must be ≥ {min_date} (T+4).")
    for e in errors:
        print(e)

    col_back, col_next = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t3_date_back"):
            os.environ[_t3_key("stage")] = "split_review" \
                if not os.environ[_t3_key("upload_city_df")].empty else "upload"
            st.rerun()
    with col_next:
        if st.button("Review & Submit →", type="primary", key="t3_date_next", disabled=bool(errors)):
            merged = hub_df.copy()
            merged["Launch Date"] = edited["Launch Date"].values
            os.environ[_t3_key("hub_split_df")] = merged
            os.environ[_t3_key("stage")] = "confirm"
            st.rerun()


def _t3_stage_confirm():
    st.subheader("Review & Submit Replacement")
    hub_df   = os.environ[_t3_key("hub_split_df")]
    pct      = os.environ[_t3_key("split_pct")]
    new_pid  = os.environ[_t3_key("new_pid")]
    new_name = os.environ[_t3_key("new_name")]
    old_name = os.environ[_t3_key("old_name")]
    print(
        f"**Replacing:** {old_name}  \n"
        f"**With:** `{new_pid}` — {new_name} ({pct}% of plan)  \n"
        f"**Rows:** {len(hub_df)} hub entries"
    )
    st.dataframe(hub_df, use_container_width=True, hide_index=True)
    col_back, col_submit = st.columns([1, 6])
    with col_back:
        if st.button("← Back", key="t3_confirm_back"):
            os.environ[_t3_key("stage")] = "set_date"
            st.rerun()
    with col_submit:
        if st.button("Submit Replacement →", type="primary", key="t3_submit_btn"):
            with st.spinner("Saving…"):
                sub_id = _submit_hub_df(hub_df, "Replacement")
            st.success(f"Replacement submitted! ID: **{sub_id}**")
            show_email_placeholder(sub_id, "Replacement", new_name, hub_df)
            _t3_reset()


# ──────────────────────────────────────────────────────────────────
# PAGE:  SUBMISSION  HISTORY
# ──────────────────────────────────────────────────────────────────
def page_history():
    print(
        '<div class="page-header">'
        '<div class="page-title">Submission History</div>'
        '<div class="page-desc">Track, approve, reject, or withdraw submissions</div>'
        '</div>', unsafe_allow_html=True
    )
    df_log = load_log()
    if df_log.empty:
        print("No submissions yet.")
        return

    user     = os.environ.get("user", {})
    is_admin = user.get("role", "") == "admin"

    # ── Filters ──────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        types = st.multiselect("Type", ["New Launch", "Expansion", "Replacement"],
                               default=["New Launch", "Expansion", "Replacement"])
    with c2:
        statuses = sorted(df_log["Status"].dropna().unique().tolist()) \
                   if "Status" in df_log.columns else []
        sel_status = st.multiselect("Status", statuses, default=statuses)
    with c3:
        pids = sorted(df_log["Product ID"].dropna().unique().tolist())
        sel_pid = st.multiselect("Product ID", pids, default=pids)

    mask = df_log["Submission_Type"].isin(types) & df_log["Product ID"].isin(sel_pid)
    if sel_status and "Status" in df_log.columns:
        mask &= df_log["Status"].isin(sel_status)
    df_f = df_log[mask].copy()

    # SLA flags
    now = datetime.now()
    if "Timestamp" in df_f.columns and "Start Date" in df_f.columns:
        df_f["SLA"] = ""
        for idx, row in df_f[df_f.get("Status", pd.Series()) == "Pending"].iterrows():
            try:
                ts     = pd.to_datetime(row["Timestamp"])
                launch = pd.to_datetime(row["Start Date"]).date()
                if launch < now.date():
                    df_f.at[idx, "Status"] = "Expired"
                    df_f.at[idx, "SLA"]    = "🔴 EXPIRED"
                elif (now - ts).total_seconds() / 3600 > 48:
                    df_f.at[idx, "SLA"] = "⚠️ OVERDUE"
            except Exception:
                pass

    disp_cols = [c for c in ["Submission_ID", "Submission_Type", "Product ID",
                              "Product Name", "City", "Hub", "Start Date",
                              "Status", "SLA", "Rejection_Reason", "Submitted_By"]
                 if c in df_f.columns]
    st.dataframe(df_f[disp_cols], use_container_width=True, hide_index=True)
    st.caption(f"{len(df_f)} rows")

    print("---")
    sub_ids = sorted(df_f["Submission_ID"].dropna().unique().tolist())
    if not sub_ids:
        return
    sel_id   = st.selectbox("Select Submission ID", sub_ids, key="hist_sel_id")
    sel_rows = df_log[df_log["Submission_ID"] == sel_id]
    sel_stat = sel_rows["Status"].iloc[0] if "Status" in sel_rows.columns and not sel_rows.empty else "Pending"
    sel_reason = sel_rows["Rejection_Reason"].iloc[0] if "Rejection_Reason" in sel_rows.columns and not sel_rows.empty else ""

    if sel_reason:
        print(f"Rejection reason: **{sel_reason}**")

    ca, cb, cc, cd = st.columns(4)
    with ca:
        if st.button("🚫 Withdraw", disabled=(sel_stat != "Pending"), key="hist_withdraw"):
            update_submission_status(sel_id, "Withdrawn")
            st.cache_data.clear()
            st.success("Withdrawn.")
            st.rerun()
    with cb:
        if st.button("✅ Approve", disabled=(not is_admin or sel_stat != "Pending"), key="hist_approve"):
            update_submission_status(sel_id, "Approved")
            st.cache_data.clear()
            st.success("Approved. Product will be added to Baseline on launch date.")
            st.rerun()
    with cc:
        if is_admin and sel_stat == "Pending":
            if True:
                reason = st.text_area("Rejection Reason", key="hist_reason")
                if st.button("Confirm Reject", key="hist_reject_confirm"):
                    if not reason.strip():
                        print("Enter a reason.")
                    else:
                        update_submission_status(sel_id, "Rejected", reason.strip())
                        st.cache_data.clear()
                        print("Rejected.")
                        st.rerun()
    with cd:
        if st.button("🔴 Void", disabled=(not is_admin or sel_stat != "Approved"), key="hist_void"):
            update_submission_status(sel_id, "Voided")
            st.cache_data.clear()
            print("Voided.")
            st.rerun()


# ──────────────────────────────────────────────────────────────────
# SESSION  STATE  INIT  (called on every rerun from app.py)
# ──────────────────────────────────────────────────────────────────
for _k, _v in {
    "submission_type": None, "submission_id": None,
    "split_ready": False,    "hub_split_df": pd.DataFrame(),
    "split_ready2": False,   "hub_split_df2": pd.DataFrame(),
    "replace_ready": False,  "replace_df": pd.DataFrame(),
}.items():
    if _k not in os.environ:
        os.environ[_k] = _v

# ──────────────────────────────────────────────────────────────────
# STANDALONE  (guarded — not executed when imported from app.py)
# ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    _page = st.sidebar.radio(
        "Navigate",
        ["New Product Launch", "Product Expansion", "Product Replacement", "Submission History"],
        key="sidebar_nav",
    )
    if _page == "New Product Launch":
        page_type1()
    elif _page == "Product Expansion":
        page_type2()
    elif _page == "Product Replacement":
        page_type3()
    else:
        page_history()
