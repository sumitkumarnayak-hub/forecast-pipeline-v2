"""
Optimized Baseline Generator Module
Keeps the baseline output contract identical while adding scalable raw-data
loading and validation hooks for automated pipeline work.
"""
import streamlit as st
import pandas as pd
import numpy as np
import pandera as pa
import polars as pl
from datetime import datetime, timedelta
from pydantic import BaseModel, Field, field_validator
from planning_suite.services.baseline_io import P_MASTER_READ_RANGE, p_master_enrichment_maps
from planning_suite.services.helpers import generate_run_id, format_number, normalize_base_plan_columns
from planning_suite.services.workflow_notifications import (
    format_error_from_parameters,
    notify_autopilot_run_finished,
    notify_baseline_approved,
    notify_baseline_run_finished,
)
from planning_suite.db.engine import Database
from planning_suite.services.google_sheets import GoogleSheetsManager
from planning_suite.services.sheets_session import get_sheets_manager
from planning_suite.core.permissions import (
    MANUAL_BASELINE_PAGES,
    PAGE_APPROVE_BASELINE,
    PAGE_AUTO_PILOT,
    PAGE_CONFIGURE_PARAMS,
    PAGE_GENERATE_BASELINE,
    PAGE_LOAD_RAW_DATA,
    PAGE_REVIEW_BASELINE,
    can_approve,
    can_write,
)
from planning_suite.ui.nav import request_main_nav
from planning_suite.services.pipeline_state import clear_baseline_approval_cache
from planning_suite.config import (
    BASELINE_APPROVAL_JSON,
    BASELINE_OUTPUTS_FOLDER,
    DEMAND_PLANNING_MASTERS_SHEET_URL,
    DP_LOGICS_SHEET_URL,
    DP_LOGICS_SHEET_KEY,
    GOOGLE_CREDENTIALS_PATH,
    PIPELINE_PARAMS_HUB_CHANGES_TAB,
    PIPELINE_PARAMS_SHEET_URL,
    PROJECT_ROOT,
    RDS_6W_PATH,
    VALIDATION_SHEET_URL,
    RAW_ACTUALS_FOLDER,
    DP_LOGICS_FOLDER,
    FF_MASTERS_XLSX,
)
import os
import glob
import pyreadr
import trino


RAW_DATA_COLUMNS_TO_KEEP = [
    "city_name", "product_id", "hub_name", "process_dt", "sales", "group_flag", "group_instances",
    "grp_r7_plan", "grp_r7_inv", "grp_r7_plan_rev", "grp_r7_inv_rev",
    "grp_BasePlan", "grp_BaseRev",
    "r7_plan", "r7_inv", "r7_plan_rev", "r7_inv_rev",
    "BasePlan", "flag", "instances",
]

RAW_DATA_SCHEMA = pa.DataFrameSchema(
    {
        "city_name": pa.Column(str, required=True),
        "product_id": pa.Column(required=True),
        "hub_name": pa.Column(str, required=True),
        "process_dt": pa.Column(pa.DateTime, required=True, coerce=True),
        "sales": pa.Column(required=True),
        "group_flag": pa.Column(required=True),
        "group_instances": pa.Column(required=True),
        "grp_r7_plan": pa.Column(required=True),
        "grp_r7_inv": pa.Column(required=True),
        "grp_r7_plan_rev": pa.Column(required=True),
        "grp_r7_inv_rev": pa.Column(required=True),
        "grp_BasePlan": pa.Column(required=True),
        "grp_BaseRev": pa.Column(required=True),
        "r7_plan": pa.Column(required=True),
        "r7_inv": pa.Column(required=True),
        "r7_plan_rev": pa.Column(required=True),
        "r7_inv_rev": pa.Column(required=True),
        "BasePlan": pa.Column(required=True),
        "flag": pa.Column(required=True),
        "instances": pa.Column(required=True),
    },
    strict=False,
)


class RawDataDateRange(BaseModel):
    """Validated raw-data pull window."""

    start_date: pd.Timestamp = Field(...)
    end_date: pd.Timestamp = Field(...)

    model_config = {"arbitrary_types_allowed": True}

    @field_validator("start_date", "end_date", mode="before")
    @classmethod
    def _coerce_timestamp(cls, value):
        return pd.to_datetime(value).normalize()

    @field_validator("end_date")
    @classmethod
    def _validate_window(cls, value, info):
        start = info.data.get("start_date")
        if start is not None and start > value:
            raise ValueError("Start date must be before end date.")
        if start is not None and (value - start).days + 1 > 7:
            raise ValueError("Select a maximum of 7 days.")
        return value


DP_LOGICS_WORKSHEET_NAMES = [
    "City_Cat",
    "SellThroughFactor",
    "City_drops",
    "Percentile",
    "Avl_Flag",
]

OPT_PILOT_STEP_KEY = "opt_pilot_step"
OPT_PILOT_STATUS_KEY = "opt_pilot_status"
OPT_PILOT_LOGS_KEY = "opt_pilot_logs"
OPT_PILOT_RUN_ID_KEY = "opt_pilot_run_id"
OPT_PILOT_RUN_NAME_KEY = "opt_pilot_run_name"


from planning_suite.core.dataframe import clean_sheet_df


def _load_p_master_df(sheets_manager: GoogleSheetsManager) -> tuple[pd.DataFrame | None, str]:
    """Load P Master from local Excel first, then Google Sheets with retries."""
    import time

    if os.path.exists(FF_MASTERS_XLSX):
        try:
            local_df = pd.read_excel(FF_MASTERS_XLSX, sheet_name="P Master")
            if local_df is not None and not local_df.empty:
                return clean_sheet_df(local_df), "local Product_Masters.xlsx"
        except Exception:
            pass

    for attempt in range(3):
        remote_df = sheets_manager.read_worksheet_uncached(
            "demand_planning_masters", "product_master", P_MASTER_READ_RANGE,
        )
        if remote_df is not None and not remote_df.empty:
            return clean_sheet_df(remote_df), "Google Sheets"
        if attempt < 2:
            time.sleep(1.5 * (attempt + 1))
    return None, ""


def _load_avl_flag_df(sheets_manager: GoogleSheetsManager) -> pd.DataFrame | None:
    """Load Avl_Flag from local DP Logics cache first, then Google Sheets."""
    local_path = os.path.join(DP_LOGICS_FOLDER, "Avl_Flag.xlsx")
    if os.path.exists(local_path):
        try:
            local_df = pd.read_excel(local_path)
            if local_df is not None and not local_df.empty:
                return local_df
        except Exception:
            pass
    return sheets_manager.read_worksheet_uncached("hub_level_planning", "avl_flag", "A:F")


class OptimizedBaselineGenerator:
    """Handles optimized baseline generation workflow."""
    
    def __init__(self):
        self.db = Database()
        self._pipeline_sheets: GoogleSheetsManager | None = None

    def use_pipeline_sheets(self, sheets: GoogleSheetsManager | None) -> None:
        """Pin one GoogleSheetsManager for an entire Auto-Pilot run."""
        self._pipeline_sheets = sheets

    @property
    def sheets_manager(self) -> GoogleSheetsManager:
        """Shared Sheets client for manual tabs and Auto-Pilot (one auth per session)."""
        if self._pipeline_sheets is not None:
            return self._pipeline_sheets
        return get_sheets_manager()

    def _clear_baseline_approval(self) -> None:
        """Remove baseline approval from session, disk, and database."""
        st.session_state.baseline_approved = False
        st.session_state.pop("baseline_approved_at", None)
        st.session_state.pop("baseline_approved_by", None)
        try:
            os.remove(str(BASELINE_APPROVAL_JSON))
        except OSError:
            pass
        self.db.revoke_baseline_approvals()
        clear_baseline_approval_cache()

    def _prepare_demo_filter_dataset(self, base_env: dict) -> tuple[str, dict, str | None]:
        """
        Apply admin demo city/hub filters before starting a baseline run.

        Returns (active_dataset_path, updated_env, error_message).
        """
        import copy

        _env = copy.copy(base_env)
        _demo_city_run = st.session_state.get("demo_city_filter", "All Cities")
        _demo_hubs_run = st.session_state.get("demo_hub_filter") or []
        _base_ds = os.path.abspath(os.path.join("outputs", "active_dataset.parquet"))
        _active_ds = _base_ds

        _needs_demo_filter = (
            (_demo_city_run and _demo_city_run != "All Cities") or bool(_demo_hubs_run)
        )
        if not _needs_demo_filter:
            return _active_ds, _env, None

        try:
            _full_df = pd.read_parquet(_base_ds)
            _filtered_df = _full_df.copy()

            if _demo_city_run and _demo_city_run != "All Cities":
                if "city_name" not in _filtered_df.columns:
                    st.warning("'city_name' column not found — skipping city filter.")
                else:
                    _filtered_df = _filtered_df[
                        _filtered_df["city_name"] == _demo_city_run
                    ].reset_index(drop=True)
                    if _filtered_df.empty:
                        _cities = sorted(_full_df["city_name"].unique().tolist())
                        return _active_ds, _env, (
                            f"No rows for city **{_demo_city_run}**. "
                            f"Available cities: {_cities}"
                        )
                    _env["BASELINE_DEMO_CITY"] = _demo_city_run

            if _demo_hubs_run:
                if "hub_name" not in _filtered_df.columns:
                    st.warning("'hub_name' column not found — skipping hub filter.")
                else:
                    _pre_hub_count = len(_filtered_df)
                    _filtered_df = _filtered_df[
                        _filtered_df["hub_name"].isin(_demo_hubs_run)
                    ].reset_index(drop=True)
                    if _filtered_df.empty:
                        _avail_hubs = sorted(_full_df["hub_name"].unique().tolist())
                        return _active_ds, _env, (
                            f"No rows for selected hub(s) **{_demo_hubs_run}** "
                            f"(after city filter). Available hubs: {_avail_hubs}"
                        )
                    st.info(
                        f"Hub filter: {_pre_hub_count:,} → **{len(_filtered_df):,} rows** "
                        f"({len(_demo_hubs_run)} hub(s): {', '.join(_demo_hubs_run)})"
                    )

            _city_part = (
                _demo_city_run.replace(" ", "_")
                if _demo_city_run != "All Cities" else "AllCities"
            )
            _hub_part = (
                "_".join(sorted(_demo_hubs_run))[:50] if _demo_hubs_run else "AllHubs"
            )
            _demo_ds = os.path.abspath(
                os.path.join("outputs", f"active_dataset_demo_{_city_part}_{_hub_part}.parquet")
            )
            _filtered_df.to_parquet(_demo_ds, index=False)
            _active_ds = _demo_ds

            _filter_desc = _demo_city_run if _demo_city_run != "All Cities" else "All Cities"
            if _demo_hubs_run:
                _filter_desc += f" · {len(_demo_hubs_run)} hub(s)"
            st.info(
                f"Demo Filter: passing **{len(_filtered_df):,} rows** "
                f"({_filter_desc}) to baseline script."
            )
        except Exception as exc:
            st.warning(f"Could not apply demo filter: {exc} — running on full dataset.")

        return _active_ds, _env, None

    @staticmethod
    def _render_page_header(title: str, description: str, *, step: int | None = None) -> None:
        step_line = f"Step {step} of {len(MANUAL_BASELINE_PAGES)} — " if step else ""
        st.markdown(
            f"""
        <div class="page-header">
            <div class="page-title">{title}</div>
            <div class="page-desc">{step_line}{description}</div>
        </div>
        """,
            unsafe_allow_html=True,
        )

    @staticmethod
    def _render_manual_step_footer(current_page: str) -> None:
        if current_page not in MANUAL_BASELINE_PAGES:
            return
        idx = MANUAL_BASELINE_PAGES.index(current_page)
        if idx >= len(MANUAL_BASELINE_PAGES) - 1:
            return
        next_page = MANUAL_BASELINE_PAGES[idx + 1]
        st.markdown("---")
        if st.button(f"Continue to {next_page} →", key=f"manual_next_{idx}", use_container_width=True):
            request_main_nav(next_page)

    def display_autopilot_page(self, user: dict) -> None:
        """Auto-Pilot — full 6-step automated pipeline."""
        user_id = user["id"]
        role = user.get("role", "viewer")
        read_only = not can_write(role)

        self._render_page_header(
            PAGE_AUTO_PILOT,
            "Run the full 6-step forecast pipeline (master sync → engine → email)",
        )

        if read_only:
            st.info("You have read-only access. Pipeline runs are disabled for your role.")

        from planning_suite.ui.components.autopilot_runner import render_autopilot_runner

        render_autopilot_runner(user_id, read_only=read_only)

    def display_load_raw_data_page(self, user: dict) -> None:
        user_id = user["id"]
        self._render_page_header(
            "Load Raw Data",
            "Fetch weekly actuals from RDS and build the active dataset",
            step=1,
        )
        self.load_raw_data(user_id)
        self._render_manual_step_footer(PAGE_LOAD_RAW_DATA)

    def display_configure_parameters_page(self, user: dict) -> None:
        self._render_page_header(
            "Configure Parameters",
            "Pipeline toggles and DP Logics worksheet sync",
            step=2,
        )
        self.configure_parameters()
        self._render_manual_step_footer(PAGE_CONFIGURE_PARAMS)

    def display_generate_baseline_page(self, user: dict) -> None:
        user_id = user["id"]
        read_only = not can_write(user.get("role", "viewer"))
        self._render_page_header(
            "Generate Baseline",
            "Run the baseline engine and write Summary output",
            step=3,
        )
        if read_only:
            st.info("You have read-only access. Baseline runs are disabled for your role.")
        self.generate_baseline(user_id, read_only=read_only)
        self._render_manual_step_footer(PAGE_GENERATE_BASELINE)

    def display_review_baseline_page(self, user: dict) -> None:
        user_id = user["id"]
        self._render_page_header(
            "Review & Validate",
            "Inspect summary output and base-plan comparison",
            step=4,
        )
        self.review_baseline(user_id)
        self._render_manual_step_footer(PAGE_REVIEW_BASELINE)

    def display_approve_baseline_page(self, user: dict) -> None:
        user_id = user["id"]
        read_only = not can_write(user.get("role", "viewer"))
        self._render_page_header(
            "Approve Baseline",
            "Lock the baseline and unlock Final Plan (admin approval)",
            step=5,
        )
        if read_only:
            st.info("You have read-only access. Approval is disabled for your role.")
        self.approve_baseline(user_id, user=user, read_only=read_only)

    def display_baseline_page(self, user):
        """Legacy entry — redirects to Load Raw Data."""
        request_main_nav(PAGE_LOAD_RAW_DATA)

    @staticmethod
    def _latest_baseline_summary_path() -> str | None:
        """Most recently written Summary_*.xlsx in BASELINE_OUTPUTS_FOLDER."""
        if not os.path.isdir(BASELINE_OUTPUTS_FOLDER):
            return None
        candidates = sorted(
            glob.glob(os.path.join(BASELINE_OUTPUTS_FOLDER, "Summary_*.xlsx")),
            key=os.path.getmtime,
            reverse=True,
        )
        return candidates[0] if candidates else None

    def _opt_pilot_output_paths_for_step(
        self,
        step_idx: int,
        metrics: dict | None = None,
    ) -> list[dict]:
        """Configured output locations for each Auto-Pilot step."""
        metrics = metrics or {}
        paths: list[dict] = []

        def _local(label: str, path: str | None) -> None:
            if not path:
                return
            paths.append({
                "label": label,
                "path": path,
                "kind": "local",
                "exists": os.path.exists(path),
            })

        def _sheet(label: str, url: str | None, note: str = "") -> None:
            if not url:
                return
            display = f"{url} ({note})" if note else url
            paths.append({
                "label": label,
                "path": display,
                "kind": "sheet",
                "exists": None,
            })

        if step_idx == 0:
            _local("Product Masters Excel", metrics.get("Excel path") or FF_MASTERS_XLSX)
            _local("Master sync state", str(PROJECT_ROOT / "outputs" / "master_sync_state.json"))

        elif step_idx == 1:
            _sheet("Hub_Changes (read)", PIPELINE_PARAMS_SHEET_URL, PIPELINE_PARAMS_HUB_CHANGES_TAB)
            _sheet("P-H Master (write)", DEMAND_PLANNING_MASTERS_SHEET_URL, "P-H Master tab")
            if int(metrics.get("P-H rows inserted") or 0) > 0:
                _local("Product Masters Excel (refreshed)", FF_MASTERS_XLSX)

        elif step_idx == 2:
            week = metrics.get("Week")
            if week is not None:
                week_file = os.path.join(RAW_ACTUALS_FOLDER, f"Raw_Actuals_Wk{week}.parquet")
                _local("Weekly raw actuals", week_file)
            _local(
                "Active dataset (engine input)",
                metrics.get("Active dataset")
                or os.path.abspath(os.path.join("outputs", "active_dataset.parquet")),
            )

        elif step_idx == 3:
            for ws_name in DP_LOGICS_WORKSHEET_NAMES:
                _local(ws_name, os.path.join(DP_LOGICS_FOLDER, f"{ws_name}.xlsx"))
                parquet_path = os.path.join(DP_LOGICS_FOLDER, f"{ws_name}.parquet")
                if os.path.exists(parquet_path):
                    _local(f"{ws_name} (parquet sidecar)", parquet_path)

        elif step_idx == 4:
            summary = metrics.get("Summary file") or self._latest_baseline_summary_path()
            _local("Baseline Summary Excel", summary)
            _local("Previous baseline cache", os.path.abspath(os.path.join("outputs", "prev_baseline_latest.parquet")))
            _sheet("Hub level Suggestion (write)", DP_LOGICS_SHEET_URL, "Hub level Suggestion tab")
            _sheet("Validation comparison (write)", VALIDATION_SHEET_URL, "Hub SKU Day, City Category, City Level")
            log_glob = sorted(
                glob.glob(os.path.join(DP_LOGICS_FOLDER, "Hub_level_Suggestion_log_*.xlsx")),
                key=os.path.getmtime,
                reverse=True,
            )
            if log_glob:
                _local("Hub suggestion backup log", log_glob[0])

        elif step_idx == 5:
            paths.append({
                "label": "Run audit",
                "path": "pipeline_runs · pipeline_step_logs · pipeline_run_log_lines (database)",
                "kind": "local",
                "exists": None,
            })

        return paths

    @staticmethod
    def _render_opt_pilot_output_paths(output_paths: list[dict]) -> None:
        if not output_paths:
            return
        st.markdown("**📁 Output paths**")
        for item in output_paths:
            label = item.get("label", "Output")
            path = item.get("path", "")
            kind = item.get("kind", "local")
            exists = item.get("exists")
            if kind == "sheet":
                url = path.split(" (", 1)[0].strip()
                note = path[len(url):].strip() if path.startswith(url) and len(path) > len(url) else ""
                if note:
                    st.markdown(f"- **{label}:** [{url}]({url}) `{note}`")
                else:
                    st.markdown(f"- **{label}:** [{url}]({url})")
            elif exists is True:
                st.markdown(f"- **{label}:** `{path}` ✅")
            elif exists is False:
                st.markdown(f"- **{label}:** `{path}` ⚠️ *not found yet*")
            else:
                st.markdown(f"- **{label}:** `{path}`")

    def _opt_pilot_all_output_paths_reference(self) -> list[dict]:
        """Static reference list of where each step writes (from .env)."""
        return [
            {"step": "Step 1", "label": "Product Masters Excel", "path": FF_MASTERS_XLSX},
            {"step": "Step 2", "label": "P-H Master (new products)", "path": DEMAND_PLANNING_MASTERS_SHEET_URL},
            {"step": "Step 3", "label": "Raw actuals folder", "path": RAW_ACTUALS_FOLDER},
            {"step": "Step 3", "label": "Active dataset", "path": os.path.abspath(os.path.join("outputs", "active_dataset.parquet"))},
            {"step": "Step 4", "label": "DP Logics folder", "path": DP_LOGICS_FOLDER},
            {"step": "Step 5", "label": "Baseline outputs folder", "path": BASELINE_OUTPUTS_FOLDER},
            {"step": "Step 5", "label": "Hub level Suggestion", "path": DP_LOGICS_SHEET_URL},
            {"step": "Step 5", "label": "Validation sheet", "path": VALIDATION_SHEET_URL},
        ]

    def _opt_pilot_steps_config(self) -> list[dict]:
        return [
            {
                "name": "Step 1: Master Data Sync & Validation",
                "desc": "Read Google Sheets masters, run Polars validation, export to Product_Masters.xlsx.",
                "icon": "📋",
            },
            {
                "name": "Step 2: New Product Launch (P-H Master)",
                "desc": "Auto-discover new products in P Master and append P-H Master rows for all active hubs.",
                "icon": "🚀",
            },
            {
                "name": "Step 3: Pull Raw Data",
                "desc": "Fetch the latest week of raw actuals from RDS cache and update the active Parquet dataset.",
                "icon": "📥",
            },
            {
                "name": "Step 4: Sync Config Parameters",
                "desc": "Sync DP Logics worksheets (City_Cat, STF, Percentile, Avl_Flag, etc.) to local Excel.",
                "icon": "⚙️",
            },
            {
                "name": "Step 5: Run Baseline Engine",
                "desc": "Execute optimized_baseline_avail_correction.py on the active dataset.",
                "icon": "🧮",
            },
            {
                "name": "Step 6: Email Notification",
                "desc": "Send success notification when all prior steps complete.",
                "icon": "📧",
            },
        ]

    def _opt_pilot_execute_step(
        self,
        step_idx: int,
        user_id: int,
        *,
        run_id: str | None = None,
        run_name: str | None = None,
        sheets_manager: GoogleSheetsManager | None = None,
    ) -> dict:
        """Run a single Auto-Pilot step. Returns a log dict for the UI."""
        import subprocess
        import sys

        sheets = sheets_manager or self.sheets_manager

        def _pilot_run_id() -> str:
            if run_id:
                return run_id
            return st.session_state.get(OPT_PILOT_RUN_ID_KEY, generate_run_id())

        def _pilot_run_name(default: str = "Auto-Pilot Pipeline") -> str:
            if run_name:
                return run_name
            return st.session_state.get(OPT_PILOT_RUN_NAME_KEY, default)

        if step_idx == 0:
            from planning_suite.automation.master_data_sync import run_master_data_excel_sync
            from planning_suite.config import FF_MASTERS_XLSX
            from planning_suite.core.validations.master_rules import VALIDATION_VERSION

            result = run_master_data_excel_sync(
                FF_MASTERS_XLSX, user_id, db=self.db, sheets_manager=sheets,
            )
            if result.validation_errors:
                raise RuntimeError(
                    f"Master sync blocked: {len(result.validation_errors)} validation error(s) "
                    f"({VALIDATION_VERSION})."
                )
            if not result.success:
                raise RuntimeError(result.error or "Master data sync failed.")

            return {
                "text": "Google Sheets synced successfully to Excel backend. All rules passed.",
                "metrics": {
                    "P Master Rows": result.p_rows,
                    "P-H Master Rows": result.ph_rows,
                    "HTT Rows": result.htt_rows,
                    "Hub Mapping Rows": result.hub_rows,
                    "Excel path": result.excel_path,
                    "File size (KB)": result.file_size_kb,
                    "Validation Status": "All checks passed",
                },
            }

        if step_idx == 1:
            from planning_suite.automation.new_product_launch_sync import run_new_product_launch_sync_cli

            result = run_new_product_launch_sync_cli(user_id, db=self.db, sheets=sheets)
            if not result.success:
                raise RuntimeError(result.error or "New product launch sync failed.")

            metrics = {
                "New products": result.products_found,
                "P-H rows inserted": result.rows_inserted,
                "Duplicates skipped": result.duplicates_skipped,
            }
            if result.masters_re_synced:
                metrics["P-H rows after re-sync"] = result.ph_rows_after
            if result.products_synced:
                metrics["Product IDs"] = ", ".join(result.products_synced[:10])
            if result.products_found == 0:
                return {
                    "text": "No new products in P Master — step skipped.",
                    "metrics": metrics,
                }
            return {
                "text": (
                    f"New product P-H Master sync complete "
                    f"({result.rows_inserted} row(s) for {len(result.products_synced)} product(s))."
                ),
                "metrics": metrics,
            }

        if step_idx == 2:
            from planning_suite.services.raw_actuals_cache import (
                resolve_raw_actuals_for_week,
                write_week_parquet,
            )

            end_date = pd.to_datetime("today").normalize() - timedelta(days=1)
            start_date = end_date - timedelta(days=6)

            def _fetch() -> pd.DataFrame | None:
                return self._fetch_raw_data_from_rds(
                    start_date, end_date, product_parity=True, sheets_manager=sheets,
                )

            df, iso_week, from_cache = resolve_raw_actuals_for_week(
                start_date, RAW_ACTUALS_FOLDER, _fetch,
            )
            week_run_name = f"Auto-Pilot Wk{iso_week}"
            try:
                st.session_state[OPT_PILOT_RUN_NAME_KEY] = week_run_name
            except Exception:
                pass

            if df is None or df.empty:
                raise ValueError("Failed to fetch raw actuals from Trino/RDS.")

            if not from_cache:
                write_week_parquet(df, iso_week, RAW_ACTUALS_FOLDER)

            active_ds = os.path.join("outputs", "active_dataset.parquet")
            dedup_keys = [c for c in ["hub_name", "product_id", "process_dt"] if c in df.columns]
            if dedup_keys:
                df = df.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)
            df.to_parquet(active_ds, index=False)

            cache_note = " (cached parquet — skipped RDS pull)" if from_cache else ""
            return {
                "text": f"Raw actuals loaded and active dataset updated{cache_note}.",
                "metrics": {
                    "Week": iso_week,
                    "Run name": week_run_name,
                    "Rows": len(df),
                    "Start": str(start_date.date()),
                    "End": str(end_date.date()),
                    "Active dataset": active_ds,
                    "Source": "cache" if from_cache else "RDS",
                },
            }

        if step_idx == 3:
            import os as _os

            os.makedirs(DP_LOGICS_FOLDER, exist_ok=True)
            _skip_dp = _os.getenv("AUTOPILOT_SKIP_DP_SHEETS_IF_FRESH_HOURS", "").strip()
            _max_local = float(_skip_dp) if _skip_dp else None
            sync_results = sheets.sync_dp_logics_worksheets_to_folder(
                DP_LOGICS_FOLDER,
                DP_LOGICS_WORKSHEET_NAMES,
                allow_local_fallback=True,
                parallel=True,
                max_local_age_hours=_max_local,
            )
            local_used = [ws for ws, info in sync_results.items() if info.get("status") == "local"]
            metrics = {
                ws: f"{info.get('rows', 0):,} rows ({info.get('source', '')})"
                for ws, info in sync_results.items()
            }
            log = {"text": "Configuration worksheets synced.", "metrics": metrics}
            if local_used:
                log["warning"] = (
                    "Google Sheets unavailable — used local files for: " + ", ".join(local_used)
                )
            from planning_suite.config import FF_MASTERS_XLSX
            from planning_suite.services.baseline_io import refresh_all_engine_sidecars
            sidecar_status = refresh_all_engine_sidecars(DP_LOGICS_FOLDER, FF_MASTERS_XLSX)
            if sidecar_status:
                log.setdefault("metrics", {}).update(
                    {f"Sidecar {k}": v for k, v in sidecar_status.items()}
                )
            return log

        if step_idx == 4:
            self._ensure_previous_baseline_for_engine()
            script_path = str(PROJECT_ROOT / "scripts" / "optimized_baseline_avail_correction.py")
            env = os.environ.copy()
            env["BASELINE_USE_ACTIVE_DATASET"] = "1"
            env["BASELINE_ACTIVE_DATASET_PATH"] = os.path.abspath(
                os.path.join("outputs", "active_dataset.parquet")
            )
            pipeline_params = sheets.read_pipeline_params()
            apply_hub = pipeline_params.get("apply_hub_changes", True)
            env["BASELINE_APPLY_HUB_CHANGES"] = "1" if apply_hub else "0"
            result = subprocess.run(
                [sys.executable, script_path],
                capture_output=True,
                text=True,
                env=env,
                timeout=1200,
            )
            if result.returncode != 0:
                detail = (result.stderr or result.stdout or "").strip()
                raise RuntimeError(
                    f"Engine failed with code {result.returncode}:\n{detail[-8000:]}"
                )
            stdout = (result.stdout or "").strip()
            summary_path = None
            for line in stdout.splitlines():
                if "Summary saved to:" in line:
                    summary_path = line.split("Summary saved to:", 1)[1].strip()
                    break
            if not summary_path:
                summary_path = self._latest_baseline_summary_path()
            return {
                "text": "Baseline engine completed successfully.",
                "metrics": {
                    "Script": script_path,
                    "Exit code": 0,
                    "Summary file": summary_path or "(not found — check BASELINE_OUTPUTS_FOLDER)",
                },
            }

        if step_idx == 5:
            notify_autopilot_run_finished(
                run_id=_pilot_run_id(),
                run_name=_pilot_run_name(),
                status="completed",
                user_id=user_id,
                db=self.db,
            )
            return {"text": "Success notification sent."}

        raise ValueError(f"Unknown Auto-Pilot step index: {step_idx}")

    def load_raw_data(self, user_id):
        """Load raw sales data from RDS file — one week at a time, saved to repository"""

        REPOSITORY_FOLDER = RAW_ACTUALS_FOLDER

        st.subheader("📥 Get Raw Sales Data")

        # ── Repository Status ──────────────────────────────────────────────
        st.markdown("### 🗂️ Repository Status")

        saved_week_files = []
        saved_weeks = []
        if os.path.exists(REPOSITORY_FOLDER):
            # Support both parquet (fast) and legacy xlsx files
            all_files = os.listdir(REPOSITORY_FOLDER)
            parquet_files = sorted([f for f in all_files if f.startswith("Raw_Actuals_Wk") and f.endswith(".parquet")])
            xlsx_files    = sorted([f for f in all_files if f.startswith("Raw_Actuals_Wk") and f.endswith(".xlsx")])
            # Prefer parquet; fall back to xlsx if no parquet for that week
            parquet_weeks = {int(f.replace("Raw_Actuals_Wk", "").replace(".parquet", "")) for f in parquet_files}
            xlsx_only     = [f for f in xlsx_files if int(f.replace("Raw_Actuals_Wk", "").replace(".xlsx", "")) not in parquet_weeks]
            saved_week_files = parquet_files + xlsx_only
            saved_weeks = sorted([
                int(f.replace("Raw_Actuals_Wk", "").replace(".parquet", "").replace(".xlsx", ""))
                for f in saved_week_files
            ])

        if saved_weeks:
            st.success(f"✅ **{len(saved_weeks)} week(s) saved in repository:** Wk {', '.join(str(w) for w in saved_weeks)}")

            # Build summary table with sales totals per week
            summary_rows = []
            for f in saved_week_files:
                wk = int(f.replace("Raw_Actuals_Wk", "").replace(".parquet", "").replace(".xlsx", ""))
                fpath = os.path.join(REPOSITORY_FOLDER, f)
                try:
                    if f.endswith(".parquet"):
                        _cols_avail = pd.read_parquet(fpath).columns.tolist()
                        _read_cols  = [c for c in ["sales", "final_sales"] if c in _cols_avail]
                        wk_df = pd.read_parquet(fpath, columns=_read_cols) if _read_cols else pd.read_parquet(fpath, columns=["sales"])
                    else:
                        _preview_cols = pd.read_excel(fpath, nrows=0).columns.tolist()
                        _read_cols    = [c for c in ["sales", "final_sales"] if c in _preview_cols]
                        wk_df = pd.read_excel(fpath, usecols=_read_cols if _read_cols else None)
                    sales_col = "sales" if "sales" in wk_df.columns else ("Sales (qty)" if "Sales (qty)" in wk_df.columns else None)
                    total_sales     = wk_df[sales_col].sum() if sales_col else None
                    total_net_sales = wk_df["final_sales"].sum() if "final_sales" in wk_df.columns else 0
                    rows_count      = len(wk_df)
                    summary_rows.append({
                        "Week":               f"Wk {wk}",
                        "Rows":               f"{rows_count:,}",
                        "Total Sales":        f"{total_sales:,.0f}"     if total_sales     is not None else "—",
                        "Sales (w/o Liq)":    f"{total_net_sales:,.0f}",
                        "Format": "parquet" if f.endswith(".parquet") else "xlsx"
                    })
                except Exception:
                    summary_rows.append({"Week": f"Wk {wk}", "Rows": "—", "Total Sales": "—", "Sales (w/o Liq)": "—", "Format": f.split(".")[-1]})

            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)
        else:
            st.info("📭 Repository is empty. Fetch a week to get started.")

        st.markdown("---")

        # ── Date Selection (max 7 days = 1 week) ──────────────────────────
        st.markdown("### 📅 Select Date Range (1 Week Max)")

        col1, col2 = st.columns(2)

        # Load parameters from Google Sheets
        from planning_suite.services.google_sheets import GoogleSheetsManager
        sheets_manager = GoogleSheetsManager()
        sheets_params = sheets_manager.read_pipeline_params()

        with col1:
            default_start_str = sheets_params.get("start_date")
            default_start = pd.to_datetime(default_start_str) if default_start_str else (pd.to_datetime("today").normalize() - timedelta(days=7))
            start_date = st.date_input(
                "Start Date",
                value=default_start,
                help="Select the start date of the week"
            )
            start_date = pd.to_datetime(start_date)

        with col2:
            default_end_str = sheets_params.get("end_date")
            default_end = pd.to_datetime(default_end_str) if default_end_str else (pd.to_datetime("today").normalize() - timedelta(days=1))
            end_date = st.date_input(
                "End Date",
                value=default_end,
                help="Select the end date (max 7 days from start)"
            )
            end_date = pd.to_datetime(end_date)

        # Check if user updated dates in the UI and write back to Google Sheets
        new_start_str = start_date.strftime("%Y-%m-%d")
        new_end_str = end_date.strftime("%Y-%m-%d")
        if new_start_str != default_start_str or new_end_str != default_end_str:
            with st.spinner("Saving dates to Google Sheet..."):
                sheets_manager.write_pipeline_params({
                    "start_date": new_start_str,
                    "end_date": new_end_str
                })
            st.success("✅ Saved new dates to Google Sheet!")
            st.rerun()

        # Validation
        if start_date > end_date:
            st.error("❌ Start date must be before end date.")
            return

        days_diff = (end_date - start_date).days + 1
        if days_diff > 7:
            st.error(f"❌ Selected range is **{days_diff} days**. Please select a maximum of **7 days** (1 week).")
            return

        # Derive ISO week number from start date
        iso_week = start_date.isocalendar()[1]
        st.info(f"📊 Selected range: **{days_diff} days** — ISO Week **{iso_week}** ({start_date.date()} to {end_date.date()})")

        col1, col2 = st.columns(2)
        with col1:
            st.info("**Source File:** 6w_v3.rds")
            st.caption("Location: Planning Database → 6w Rolling Data")
        with col2:
            st.warning("**Excluded Hubs:** INDORE, KKD, RAIPUR, NAGDRM, VDR")
            if iso_week in saved_weeks:
                st.warning(f"⚠️ Week {iso_week} already exists in repository — fetching will **overwrite** it.")

        also_save_csv = st.checkbox("📄 Also save CSV copy (for reference)", value=False,
                                    help="Saves an additional Raw_Actuals_Wk{N}.csv alongside the parquet file")
        use_cached_week = st.checkbox(
            "⚡ Use cached week parquet if available (skip RDS pull)",
            value=True,
            help="Same optimization as Auto-Pilot Step 3 — loads Raw_Actuals_Wk{N}.parquet when it exists.",
        )

        st.markdown("---")

        # ── Fetch Button ───────────────────────────────────────────────────
        if st.button("🔄 Fetch Raw Data", type="primary", use_container_width=True):
            with st.spinner(f"Fetching {start_date.date()} → {end_date.date()} (Wk {iso_week}) from RDS file..."):
                try:
                    from planning_suite.services.raw_actuals_cache import (
                        resolve_raw_actuals_for_week,
                        write_week_parquet,
                    )

                    def _fetch():
                        return self._fetch_raw_data_from_rds(
                            start_date, end_date, sheets_manager=self.sheets_manager,
                        )

                    df, iso_week, from_cache = resolve_raw_actuals_for_week(
                        start_date,
                        REPOSITORY_FOLDER,
                        _fetch,
                        force_refresh=not use_cached_week,
                    )

                    if df is not None and not df.empty:
                        os.makedirs(REPOSITORY_FOLDER, exist_ok=True)
                        week_file = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{iso_week}.parquet")
                        already_exists = os.path.exists(week_file)
                        if not from_cache:
                            write_week_parquet(df, iso_week, REPOSITORY_FOLDER)
                        elif not already_exists:
                            write_week_parquet(df, iso_week, REPOSITORY_FOLDER)

                        if from_cache:
                            st.info(f"⚡ Loaded week {iso_week} from cached parquet (skipped RDS pull).")
                        elif already_exists:
                            st.success(f"✅ Week {iso_week} overwritten: `{week_file}`")
                        else:
                            st.success(f"✅ Week {iso_week} saved: `{week_file}`")

                        if also_save_csv:
                            csv_file = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{iso_week}.csv")
                            df.to_csv(csv_file, index=False)
                            st.success(f"📄 CSV copy saved: `{csv_file}`")

                        st.rerun()
                    else:
                        st.error(f"❌ No data found for the selected date range in the RDS file.")

                except Exception as e:
                    st.error(f"❌ Error fetching data: {str(e)}")
                    import traceback
                    with st.expander("🔍 Error Details"):
                        st.code(traceback.format_exc())

        # ── Bulk Pull: Past 10 Weeks ───────────────────────────────────────
        st.markdown("---")
        with st.expander("⏳ One-Time Bulk Pull — Past 10 Weeks", expanded=False):
            st.caption("Use this to populate the repository with historical data. Pulls each of the past 10 weeks separately and saves them as individual Excel files.")

            today = pd.to_datetime("today").normalize()
            # Build week ranges: each Mon–Sun going back 10 weeks
            bulk_weeks = []
            for i in range(1, 11):
                week_end   = today - timedelta(days=today.weekday() + 1) - timedelta(weeks=i - 1)
                week_start = week_end - timedelta(days=6)
                iso_wk     = week_start.isocalendar()[1]
                bulk_weeks.append((iso_wk, week_start, week_end))

            st.markdown("**Weeks that will be pulled:**")
            preview_data = [
                {
                    "ISO Week": w[0],
                    "Start": w[1].date(),
                    "End":   w[2].date(),
                    "Already Saved": "✅ Yes" if w[0] in saved_weeks else "⬜ No"
                }
                for w in bulk_weeks
            ]
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

            bulk_also_csv = st.checkbox("📄 Also save CSV copy for each week (for reference)", value=False,
                                        key="bulk_csv_checkbox",
                                        help="Saves Raw_Actuals_Wk{N}.csv alongside each parquet file")

            if st.button("🚀 Pull All 10 Weeks & Save", type="primary", use_container_width=True):
                os.makedirs(REPOSITORY_FOLDER, exist_ok=True)
                try:
                    # Use parquet cache — fast after first load
                    full_df = self._load_rds_cached()

                    HUBS_TO_EXCLUDE = ['INDORE', 'KKD', 'RAIPUR', 'NAGDRM', 'VDR']
                    progress_bar = st.progress(0)
                    status_text  = st.empty()

                    for idx, (iso_wk, wk_start, wk_end) in enumerate(bulk_weeks):
                        status_text.text(f"Saving Week {iso_wk} ({wk_start.date()} → {wk_end.date()})...")

                        week_df = full_df[
                            (full_df['process_dt'] >= wk_start) &
                            (full_df['process_dt'] <= wk_end) &
                            (~full_df['hub_name'].isin(HUBS_TO_EXCLUDE))
                        ].copy()

                        # Derive simple_flag columns from raw RDS data
                        if all(c in week_df.columns for c in ["flag", "instances", "group_flag", "group_instances", "r7_inv"]):
                            week_df["plan_sum"] = week_df.groupby(
                                ["hub_name", "process_dt", "product_id"]
                            )["r7_inv"].transform("sum")
                            week_df["simple_flag_when_SP_0"]          = np.where(week_df["plan_sum"] == 0, week_df["group_flag"],      week_df["flag"])
                            week_df["simple_instances_when_SP_0"]     = np.where(week_df["plan_sum"] == 0, week_df["group_instances"], week_df["instances"])
                            week_df["simple_group_flag_when_SP_0"]      = week_df["group_flag"]
                            week_df["simple_group_instances_when_SP_0"] = week_df["group_instances"]
                            week_df["week"] = week_df["process_dt"].dt.isocalendar().week.astype(int)
                            week_df["day"]  = week_df["process_dt"].dt.strftime("%a")
                            week_df.drop(columns=["plan_sum"], inplace=True)

                        # Fetch liquidation and compute final_sales for this week
                        liq_wk = self._fetch_liquidation_data(
                            pd.Timestamp(wk_start), pd.Timestamp(wk_end)
                        )
                        if not liq_wk.empty:
                            week_df["product_id"] = week_df["product_id"].astype(str)
                            week_df = week_df.merge(
                                liq_wk, on=["hub_name", "product_id", "process_dt"], how="left"
                            )
                            week_df["packets_sold"] = pd.to_numeric(
                                week_df["packets_sold"], errors="coerce"
                            ).fillna(0)
                        else:
                            week_df["packets_sold"] = 0
                        week_df["final_sales"] = np.maximum(
                            week_df["sales"] - week_df["packets_sold"], 0
                        )

                        # Pre-save hygiene: remove PAW/OFF hubs + deduplicate keys
                        if "hub_name" in week_df.columns:
                            _hub_mask = week_df["hub_name"].astype(str).str.upper().str.startswith(("PAW", "OFF"))
                            week_df = week_df[~_hub_mask].copy()
                        _dedup_keys = [c for c in ["hub_name", "product_id", "process_dt"] if c in week_df.columns]
                        if _dedup_keys:
                            week_df = week_df.drop_duplicates(subset=_dedup_keys, keep="first").reset_index(drop=True)

                        week_file = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{iso_wk}.parquet")
                        already_exists = os.path.exists(week_file)
                        week_df.to_parquet(week_file, index=False)

                        if bulk_also_csv:
                            csv_file = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{iso_wk}.csv")
                            week_df.to_csv(csv_file, index=False)
                            label = ("overwritten" if already_exists else "saved") + " + CSV"
                        else:
                            label = "overwritten" if already_exists else "saved"

                        st.caption(f"✅ Wk {iso_wk}: {len(week_df):,} rows {label}")
                        progress_bar.progress((idx + 1) / len(bulk_weeks))

                    status_text.empty()
                    progress_bar.empty()
                    st.success("🎉 All 10 weeks saved to repository!")
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Error during bulk pull: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

        # ── Load All Weeks from Repository into raw_data ───────────────────
        st.markdown("---")
        st.markdown("### 📦 Load Repository into Session")
        st.caption("Combines all saved week files and loads them as the active dataset for baseline generation.")

        if saved_weeks:
            col1, col2 = st.columns([2, 1])
            with col1:
                weeks_to_load = st.multiselect(
                    "Select weeks to load",
                    options=saved_weeks,
                    default=saved_weeks,
                    help="Choose which week files to include in main_df"
                )
            with col2:
                st.write("")
                st.write("")
                load_btn = st.button("📂 Load Selected Weeks", type="secondary", use_container_width=True)

            if load_btn:
                if not weeks_to_load:
                    st.warning("⚠️ Please select at least one week.")
                else:
                    with st.spinner("Loading week files..."):
                        try:
                            # Only these columns in the active dataset
                            FINAL_COLS = [
                                "process_dt",
                                "Sub-category",
                                "week",
                                "day",
                                "product_id",
                                "product_name",
                                "sku class prod",
                                "city_name",
                                "hub_name",
                                "sales",
                                "final_sales",
                                "simple_flag_when_SP_0",
                                "simple_instances_when_SP_0",
                                "simple_group_flag_when_SP_0",
                                "simple_group_instances_when_SP_0"
                            ]

                            all_dfs = []
                            for wk in sorted(weeks_to_load):
                                parquet_path = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{wk}.parquet")
                                xlsx_path    = os.path.join(REPOSITORY_FOLDER, f"Raw_Actuals_Wk{wk}.xlsx")
                                if os.path.exists(parquet_path):
                                    wk_df = pd.read_parquet(parquet_path)
                                elif os.path.exists(xlsx_path):
                                    wk_df = pd.read_excel(xlsx_path)
                                else:
                                    st.warning(f"⚠️ File not found for Wk {wk}, skipping.")
                                    continue

                                if "final_sales" not in wk_df.columns:
                                    wk_df["final_sales"] = 0

                                # Keep only the required columns (skip missing ones gracefully)
                                cols_present = [c for c in FINAL_COLS if c in wk_df.columns]
                                all_dfs.append(wk_df[cols_present])

                            combined_df = pd.concat(all_dfs, ignore_index=True)

                            # ── Demo City + Hub Filter ────────────────────────
                            _demo_city  = st.session_state.get("demo_city_filter", "All Cities")
                            _demo_hubs  = st.session_state.get("demo_hub_filter") or []

                            if _demo_city and _demo_city != "All Cities":
                                if "city_name" in combined_df.columns:
                                    _rows_before_city = len(combined_df)
                                    combined_df = combined_df[
                                        combined_df["city_name"] == _demo_city
                                    ].reset_index(drop=True)
                                    st.info(
                                        f"Demo Filter — city **{_demo_city}**: "
                                        f"{len(combined_df):,} rows retained "
                                        f"(from {_rows_before_city:,} total)"
                                    )
                                else:
                                    st.warning("Demo Filter: 'city_name' column not found — city filter not applied.")

                            if _demo_hubs:
                                if "hub_name" in combined_df.columns:
                                    _rows_before_hub = len(combined_df)
                                    combined_df = combined_df[
                                        combined_df["hub_name"].isin(_demo_hubs)
                                    ].reset_index(drop=True)
                                    st.info(
                                        f"Demo Filter — {len(_demo_hubs)} hub(s) "
                                        f"({', '.join(_demo_hubs)}): "
                                        f"{len(combined_df):,} rows retained "
                                        f"(from {_rows_before_hub:,} total)"
                                    )
                                else:
                                    st.warning("Demo Filter: 'hub_name' column not found — hub filter not applied.")

                            # Remove PAW/OFF hubs
                            if "hub_name" in combined_df.columns:
                                _hub_mask  = combined_df["hub_name"].astype(str).str.upper().str.startswith(("PAW", "OFF"))
                                _hub_count = _hub_mask.sum()
                                if _hub_count:
                                    combined_df = combined_df[~_hub_mask].reset_index(drop=True)
                                    st.info("🗑️ Removed {:,} rows where hub_name starts with 'PAW' or 'OFF'".format(_hub_count))

                            # Deduplicate at hub × product_id × date level
                            _rows_before = len(combined_df)
                            _dedup_keys  = [c for c in ["hub_name", "product_id", "process_dt"] if c in combined_df.columns]
                            if _dedup_keys:
                                combined_df = combined_df.drop_duplicates(subset=_dedup_keys, keep="first").reset_index(drop=True)
                            _dupes = _rows_before - len(combined_df)
                            if _dupes:
                                st.info(f"🗑️ Removed {_dupes:,} duplicate rows (hub × product_id × date)")

                            # Save to disk — keeps session_state tiny, prevents white-screen crashes
                            _ACTIVE_DS = os.path.join("outputs", "active_dataset.parquet")
                            os.makedirs("outputs", exist_ok=True)
                            combined_df.to_parquet(_ACTIVE_DS, index=False)

                            # Store only lightweight metadata in session_state
                            st.session_state.raw_data_loaded   = True
                            st.session_state.raw_data_rows     = len(combined_df)
                            st.session_state.raw_data_cols     = combined_df.columns.tolist()
                            st.session_state.raw_data_file     = f"Repository Wk {', '.join(str(w) for w in sorted(weeks_to_load))}"

                            st.success(f"✅ Loaded {len(combined_df):,} rows | {len(combined_df.columns)} columns | {len(weeks_to_load)} week(s): Wk {', '.join(str(w) for w in sorted(weeks_to_load))}")
                            st.caption(f"Columns: {', '.join(combined_df.columns.tolist())}")

                        except Exception as e:
                            st.error(f"❌ Error loading repository: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())

        # ── Summary of currently loaded data ──────────────────────────────
        _ACTIVE_DS = os.path.join("outputs", "active_dataset.parquet")
        _ds_ready  = st.session_state.get("raw_data_loaded") or os.path.exists(_ACTIVE_DS)

        # Auto-detect on first load (parquet exists from a previous session)
        if os.path.exists(_ACTIVE_DS) and not st.session_state.get("raw_data_loaded"):
            st.session_state.raw_data_loaded = True

        if _ds_ready:
            st.markdown("---")
            st.markdown("### 📊 Currently Loaded Data")

            # Show pre-computed metadata instantly (no file read)
            _rows = st.session_state.get("raw_data_rows")
            _file = st.session_state.get("raw_data_file", "local parquet cache")
            if _rows:
                st.metric("Total Rows", f"{_rows:,}")
            st.caption(f"Source: {_file}")

            # Load from disk only for the detailed sections (inside expanders)
            with st.expander("📊 Dataset Statistics & Quality", expanded=True):
                with st.spinner("Loading dataset stats..."):
                    df = pd.read_parquet(_ACTIVE_DS)

                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Total Rows", f"{len(df):,}")
                with col2:
                    st.metric("Unique Hubs", f"{df['hub_name'].nunique():,}" if 'hub_name' in df.columns else "—")
                with col3:
                    st.metric("Unique Cities", f"{df['city_name'].nunique():,}" if 'city_name' in df.columns else "—")
                with col4:
                    st.metric("Total Sales", f"{df['sales'].sum():,.0f}" if 'sales' in df.columns else "—")
                with col5:
                    st.metric("Days", f"{df['process_dt'].nunique():,}" if 'process_dt' in df.columns else "—")

                # Weekly summary
                st.markdown("#### 📅 Weekly Sales Summary")
                weekly_summary = self._calculate_weekly_summary(df)
                _fmt = {'Total Sales': '{:,.0f}'}
                if 'Sales (w/o Liq)' in weekly_summary.columns:
                    _fmt['Sales (w/o Liq)'] = '{:,.0f}'
                st.dataframe(
                    weekly_summary.style.format(_fmt),
                    use_container_width=True
                )

                # Data quality
                st.markdown("#### ⚠️ Data Quality Alerts")
                alerts = self._check_data_quality(df)
                if alerts['has_issues']:
                    st.error(f"⚠️ Found {alerts['total_issues']} data quality issues!")
                    if alerts['zero_flags']['count'] > 0:
                        with st.expander(f"🚩 Zero Availability Flags ({alerts['zero_flags']['count']} city-days)", expanded=True):
                            st.dataframe(alerts['zero_flags']['details'], use_container_width=True, height=300)
                else:
                    st.success("✅ No data quality issues detected!")

                del df  # free memory immediately after display

            with st.expander("📋 View Raw Data (First 100 rows)"):
                with st.spinner("Loading preview…"):
                    st.dataframe(pd.read_parquet(_ACTIVE_DS).head(100), use_container_width=True)
        else:
            st.info("👆 Fetch a week and click 'Load Selected Weeks' to activate the dataset.")
    
    def _load_rds_cached(self):
        """
        Load 6w_v3.rds, using a local Parquet cache for speed.
        - First call: reads RDS (slow, ~1-2 min), saves as local parquet cache.
        - Subsequent calls: reads parquet cache (fast, seconds).
        - Cache is refreshed automatically if the RDS file is newer than the cache.
        """
        rds_path   = RDS_6W_PATH
        cache_path = os.path.join("outputs", "rds_cache.parquet")
        os.makedirs("outputs", exist_ok=True)

        # Check if a valid cache exists and is newer than the RDS file
        use_cache = False
        if os.path.exists(cache_path):
            try:
                rds_mtime   = os.path.getmtime(rds_path)
                cache_mtime = os.path.getmtime(cache_path)
                use_cache   = cache_mtime >= rds_mtime
            except Exception:
                use_cache = False

        if use_cache:
            st.info("⚡ Loading from local cache (parquet)...")
            df = pd.read_parquet(cache_path)
            st.success(f"✅ Cache loaded: {len(df):,} rows")
        else:
            st.info("📂 Reading RDS file (first time or file updated — this takes ~1-2 min)...")
            result = pyreadr.read_r(rds_path)
            df = next(iter(result.values()))
            df['process_dt'] = pd.to_datetime(df['process_dt'])
            df.to_parquet(cache_path, index=False)
            st.success(f"✅ RDS loaded & cached locally: {len(df):,} rows (future reads will be instant)")

        df['process_dt'] = pd.to_datetime(df['process_dt'])
        return df

    def _validate_raw_input_schema(self, df: pd.DataFrame) -> pd.DataFrame:
        """Validate raw RDS columns without changing downstream data values."""
        return RAW_DATA_SCHEMA.validate(df, lazy=True)

    def _filter_raw_data_polars(
        self,
        df: pd.DataFrame,
        start_date: pd.Timestamp,
        end_date: pd.Timestamp,
    ) -> tuple[pd.DataFrame, int]:
        """Use Polars for date, hub, and column filtering at large-row scale."""
        hubs_to_exclude = ["INDORE", "KKD", "RAIPUR", "NAGDRM", "VDR"]
        missing_cols = [c for c in RAW_DATA_COLUMNS_TO_KEEP if c not in df.columns]
        if missing_cols:
            raise ValueError(f"Raw RDS data is missing required columns: {missing_cols}")

        lf = pl.from_pandas(df, include_index=False).lazy()
        date_start = pd.Timestamp(start_date).to_pydatetime()
        date_end = pd.Timestamp(end_date).to_pydatetime()

        in_range = lf.filter(
            (pl.col("process_dt") >= date_start)
            & (pl.col("process_dt") <= date_end)
        )
        before_exclusion = in_range.select(pl.len()).collect().item()

        filtered = (
            in_range
            .filter(~pl.col("hub_name").is_in(hubs_to_exclude))
            .filter(
                pl.col("product_id").is_not_null()
                & (pl.col("product_id").cast(pl.Utf8).str.strip_chars() != "")
                & pl.col("hub_name").is_not_null()
                & (pl.col("hub_name").cast(pl.Utf8).str.strip_chars() != "")
                & pl.col("process_dt").is_not_null()
            )
            .select(RAW_DATA_COLUMNS_TO_KEEP)
            .collect()
            .to_pandas()
        )
        excluded_count = before_exclusion - len(filtered)
        filtered["process_dt"] = pd.to_datetime(filtered["process_dt"])
        return filtered, excluded_count

    def _fetch_liquidation_data(self, start_date, end_date):
        """
        Fetch liquidation packets_sold from Trino for a given date range.
        Returns a DataFrame with columns: hub_name, product_id, process_dt, packets_sold.
        Returns an empty DataFrame on any connection/query failure.
        """
        Start = start_date.strftime("%Y-%m-%d")
        End   = end_date.strftime("%Y-%m-%d")
        try:
            conn = trino.dbapi.connect(
                host="trino.internal.dp.licious.com",
                port=80,
                user="default",
                catalog="hive",
                schema="planning",
                http_scheme="http",
            )
            cursor = conn.cursor()
            query = f"""
SELECT
    fnl4.dt,
    fnl4.hubid,
    map.hub_name,
    map.city_name,
    fnl4.productid,
    fnl4.productname,
    fnl4.liq_discount_perc,
    fnl4.packets_sold,
    fnl4.gross_revenue AS "gross_revenue (mrp)"
FROM (
    SELECT
        dt,
        hubid,
        productid,
        productname,
        liq_discount_perc,
        SUM(productqty) AS packets_sold,
        SUM(mrpproductpricef) AS gross_revenue
    FROM (
        SELECT
            *,
            ROUND(
                (mrpproductpricef - productdiscountf)
                * 100.00 / mrpproductpricef,
                0
            ) AS liq_discount_perc
        FROM (
            SELECT
                *,
                CASE
                    WHEN pormotionlevers_string LIKE '%"type":"LIQUIDATION"%'
                    THEN 1 ELSE 0
                END AS flag
            FROM (
                SELECT
                    *,
                    array_join(
                        transform(
                            promotionlevers,
                            x -> format(
                                '{{"leverid":"%s","type":"%s"}}',
                                x.leverid,
                                x.type
                            )
                        ),
                        ',',
                        '[]'
                    ) AS pormotionlevers_string
                FROM b2c_supplychain.order_item_events_fact
                WHERE status != 'Rejected'
                  AND (
                        yr > year(current_date - interval '84' day)
                     OR (
                            yr  = year(current_date - interval '84' day)
                        AND mon >= month(current_date - interval '84' day)
                        )
                      )
            ) fnl
        ) fnl2
        WHERE flag = 1
    ) fnl3
    WHERE CAST(dt AS DATE)
          BETWEEN
            CAST(date_parse('{Start}', '%Y-%m-%d') AS DATE)
        AND
            CAST(date_parse('{End}', '%Y-%m-%d') AS DATE)
    GROUP BY
        dt,
        hubid,
        productid,
        productname,
        liq_discount_perc
) fnl4
LEFT JOIN pipeline.city_mapping_ba map
    ON CAST(map.hub_id AS VARCHAR) = fnl4.hubid
ORDER BY
    1, 2, 5, 8 DESC
"""
            cursor.execute(query)
            rows    = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            liq_df  = pd.DataFrame(rows, columns=columns)
            liq_df["process_dt"] = pd.to_datetime(liq_df["dt"], errors="coerce")
            if "productid" in liq_df.columns and "product_id" not in liq_df.columns:
                liq_df.rename(columns={"productid": "product_id"}, inplace=True)
            liq_df["packets_sold"] = pd.to_numeric(liq_df["packets_sold"], errors="coerce").fillna(0)
            liq_df["product_id"]   = liq_df["product_id"].astype(str)
            liq_df = liq_df.groupby(
                ["hub_name", "product_id", "process_dt"], as_index=False
            )["packets_sold"].sum()
            return liq_df[["hub_name", "product_id", "process_dt", "packets_sold"]]
        except Exception as e:
            st.warning(f"⚠️ Could not fetch liquidation data from Trino: {e}. final_sales will equal sales.")
            return pd.DataFrame(columns=["hub_name", "product_id", "process_dt", "packets_sold"])

    def _fetch_raw_data_from_rds(
        self,
        start_date,
        end_date,
        *,
        product_parity: bool = False,
        sheets_manager: GoogleSheetsManager | None = None,
    ):
        """
        Fetch raw data for a specific date range (max 7 days) from the RDS cache.

        When ``product_parity=True`` (Auto-Pilot Step 3), match Product/baseline_generator:
        RDS-only flags, no Trino liquidation, no PAW/OFF hub filter, no SKU enrichment.
        """
        date_range = RawDataDateRange(start_date=start_date, end_date=end_date)
        start_date = date_range.start_date
        end_date = date_range.end_date

        df = self._load_rds_cached()

        st.info(f"📅 Filtering data: {start_date.date()} to {end_date.date()} ({(end_date - start_date).days + 1} days)")

        filtered_df, excluded_count = self._filter_raw_data_polars(df, start_date, end_date)

        if filtered_df.empty:
            st.error(f"❌ No data found for {start_date.date()} → {end_date.date()} in the RDS file.")
            return None

        self._validate_raw_input_schema(filtered_df)
        
        # Exclude specific hubs
        HUBS_TO_EXCLUDE = ['INDORE', 'KKD', 'RAIPUR', 'NAGDRM', 'VDR']
        before_exclusion = len(filtered_df)
        filtered_df = filtered_df[~filtered_df['hub_name'].isin(HUBS_TO_EXCLUDE)]
        excluded_count = before_exclusion - len(filtered_df)
        
        if excluded_count > 0:
            st.info(f"🚫 Excluded {excluded_count:,} records from hubs: {', '.join(HUBS_TO_EXCLUDE)}")
        
        # Select columns
        columns_to_keep = [
            'city_name', 'product_id', 'hub_name', 'process_dt', 'sales', 'group_flag', 'group_instances',
            'grp_r7_plan', 'grp_r7_inv', 'grp_r7_plan_rev', 'grp_r7_inv_rev',
            'grp_BasePlan', 'grp_BaseRev',
            'r7_plan', 'r7_inv', 'r7_plan_rev', 'r7_inv_rev',
            'BasePlan', 'flag', 'instances'
        ]
        
        filtered_df = filtered_df[columns_to_keep]
        
        # Calculate weighted metrics
        st.info("⚙️ Calculating weighted metrics...")
        filtered_df['wgt_flag'] = filtered_df['flag'] * filtered_df['r7_plan_rev']
        filtered_df['wgt_instances'] = filtered_df['instances'] * filtered_df['r7_plan_rev']
        
        filtered_df['new_grp_flag'] = np.where(
            filtered_df['r7_plan'] == 0,
            0,
            filtered_df['group_flag'] * filtered_df['grp_r7_plan_rev']
        )
        
        filtered_df['new_grp_instances'] = np.where(
            filtered_df['r7_plan'] == 0,
            0,
            filtered_df['group_instances'] * filtered_df['grp_r7_plan_rev']
        )
        
        # Merge with Product Master (Avl_Flag)
        st.info("🔗 Merging with Product Master...")
        gsm = sheets_manager or self.sheets_manager
        P_Master = _load_avl_flag_df(gsm)
        
        if P_Master is None or P_Master.empty:
            st.warning("⚠️ Could not load Avl_Flag master. Proceeding without it.")
            merged_df = filtered_df.copy()
            # Add placeholder column if not present
            if 'Anchor ID' not in merged_df.columns:
                merged_df['Anchor ID'] = merged_df['product_id']
        else:
            merged_df = filtered_df.merge(P_Master, on="product_id", how="left")
            # Fill missing Anchor IDs with product_id
            if 'Anchor ID' in merged_df.columns:
                merged_df['Anchor ID'] = merged_df['Anchor ID'].fillna(merged_df['product_id'])
            else:
                merged_df['Anchor ID'] = merged_df['product_id']
        
        # Calculate plan_sum
        merged_df['plan_sum'] = merged_df.groupby(['hub_name', 'process_dt', 'Anchor ID'])['r7_inv'].transform('sum')
        
        # Calculate simple flags
        st.info("📊 Calculating availability flags...")
        merged_df['simple_flag_when_SP_0'] = np.where(
            merged_df['plan_sum'] == 0,
            merged_df['group_flag'],
            merged_df['flag']
        )
        
        merged_df['simple_instances_when_SP_0'] = np.where(
            merged_df['plan_sum'] == 0,
            merged_df['group_instances'],
            merged_df['instances']
        )
        
        merged_df['simple_group_flag_when_SP_0'] = np.where(
            merged_df['plan_sum'] == 0,
            merged_df['group_flag'],
            merged_df['group_flag']
        )
        
        merged_df['simple_group_instances_when_SP_0'] = np.where(
            merged_df['plan_sum'] == 0,
            merged_df['group_instances'],
            merged_df['group_instances']
        )
        
        # Remove duplicates
        merged_df = merged_df.drop_duplicates(
            subset=["city_name", "hub_name", "product_id", "process_dt"]
        )

        if product_parity:
            merged_df["process_dt"] = pd.to_datetime(merged_df["process_dt"], errors="coerce")
            merged_df["week"] = merged_df["process_dt"].dt.isocalendar().week.astype(int)
            merged_df["day"] = merged_df["process_dt"].dt.strftime("%a")
            product_cols = [
                "city_name", "hub_name", "product_id", "process_dt", "week", "day", "sales",
                "simple_flag_when_SP_0", "simple_instances_when_SP_0",
                "simple_group_flag_when_SP_0", "simple_group_instances_when_SP_0",
            ]
            final_df = merged_df[[c for c in product_cols if c in merged_df.columns]].copy()
            st.success(
                f"✅ Product-parity raw data: {len(final_df):,} rows × {len(final_df.columns)} columns"
            )
            return final_df

        # ── Fetch liquidation and compute final_sales ──────────────────────
        st.info("🧾 Fetching liquidation data from Trino...")
        liq_df = self._fetch_liquidation_data(start_date, end_date)
        if not liq_df.empty:
            merged_df["product_id"] = merged_df["product_id"].astype(str)
            merged_df = merged_df.merge(
                liq_df, on=["hub_name", "product_id", "process_dt"], how="left"
            )
            merged_df["packets_sold"] = pd.to_numeric(merged_df["packets_sold"], errors="coerce").fillna(0)
        else:
            merged_df["packets_sold"] = 0
        merged_df["final_sales"] = np.maximum(merged_df["sales"] - merged_df["packets_sold"], 0)

        # ── Derive week and day from process_dt ───────────────────────────
        st.info("📅 Deriving week and day columns...")
        merged_df['week'] = merged_df['process_dt'].dt.isocalendar().week.astype(int)
        merged_df['day']  = merged_df['process_dt'].dt.strftime('%a')

        # ── Enrich with SKU Class Prod, product_name, Sub-category ───────
        st.info("🔗 Loading P Master for SKU enrichment...")
        try:
            p_master_df, p_master_source = _load_p_master_df(sheets_manager)
            if p_master_df is not None and not p_master_df.empty:
                sku_map, name_map, category_map, dup_ids = p_master_enrichment_maps(p_master_df)

                merged_df["sku class prod"] = merged_df["product_id"].map(sku_map)
                merged_df["product_name"]   = merged_df["product_id"].map(name_map)
                merged_df["Sub-category"]   = merged_df["product_id"].map(category_map)
                msg = (
                    f"SKU enrichment applied from P Master ({len(sku_map):,} products, "
                    f"{p_master_source})"
                )
                if dup_ids:
                    msg += f" — {dup_ids} duplicate product id(s) in P Master kept first row only."
                st.success(f"✅ {msg}")
            else:
                raise ValueError("P Master returned empty")
        except Exception as e:
            st.warning(f"⚠️ Could not load P Master: {e}. SKU columns will be empty.")
            merged_df["sku class prod"] = None
            merged_df["product_name"]   = None
            merged_df["Sub-category"]   = None

        # ── Select only the final lean columns ────────────────────────────
        final_df = merged_df[[
            "process_dt",
            "Sub-category",
            "week",
            "day",
            "product_id",
            "product_name",
            "sku class prod",
            "city_name",
            "hub_name",
            "sales",
            "final_sales",
            "simple_flag_when_SP_0",
            "simple_instances_when_SP_0",
            "simple_group_flag_when_SP_0",
            "simple_group_instances_when_SP_0"
        ]]

        # ── Pre-save hygiene: remove PAW/OFF hubs + deduplicate keys ─────
        if "hub_name" in final_df.columns:
            _hub_mask = final_df["hub_name"].astype(str).str.upper().str.startswith(("PAW", "OFF"))
            final_df = final_df[~_hub_mask].copy()
        _dedup_keys = [c for c in ["hub_name", "product_id", "process_dt"] if c in final_df.columns]
        if _dedup_keys:
            final_df = final_df.drop_duplicates(subset=_dedup_keys, keep="first").reset_index(drop=True)

        st.success(f"✅ Data processing complete! {len(final_df):,} rows × {len(final_df.columns)} columns")
        return final_df
    
    def _calculate_weekly_summary(self, df):
        """
        Calculate weekly sales summary
        """
        df_copy = df.copy()
        df_copy['process_dt'] = pd.to_datetime(df_copy['process_dt'])
        
        # Add week info
        df_copy['year'] = df_copy['process_dt'].dt.year
        df_copy['week'] = df_copy['process_dt'].dt.isocalendar().week
        df_copy['week_start'] = df_copy['process_dt'] - pd.to_timedelta(df_copy['process_dt'].dt.dayofweek, unit='D')
        df_copy['week_end'] = df_copy['week_start'] + pd.Timedelta(days=6)
        
        if "final_sales" not in df_copy.columns:
            df_copy["final_sales"] = 0

        # Group by week
        agg_dict = {'sales': 'sum', 'process_dt': 'nunique', 'final_sales': 'sum'}

        weekly = df_copy.groupby(['year', 'week', 'week_start', 'week_end']).agg(agg_dict).reset_index()

        base_cols = ['Year', 'Week', 'Week Start', 'Week End', 'Total Sales', 'Days']
        weekly.columns = base_cols + ['Sales (w/o Liq)']

        # Format dates
        weekly['Week Range'] = weekly.apply(
            lambda row: f"{row['Week Start'].strftime('%b %d')} - {row['Week End'].strftime('%b %d')}",
            axis=1
        )

        weekly = weekly[['Week', 'Week Range', 'Total Sales', 'Sales (w/o Liq)']]
        return weekly.sort_values('Week', ascending=False)
    
    def _load_rds_cached_baseline(self):
        """
        Load 6w_v3.rds for baseline fetch.
        Reuses the same parquet cache as _load_rds_cached() so the RDS is
        never read twice — if raw data was already cached, this is instant.
        """
        # Use the shared cache built by _load_rds_cached()
        df = self._load_rds_cached()

        # Add Week/day columns if not already present in the cache
        if "Week" not in df.columns:
            df["process_dt"] = pd.to_datetime(df["process_dt"], errors="coerce")
            df = df.dropna(subset=["process_dt"])
            df["Week"] = df["process_dt"].dt.isocalendar().week.astype(int)
        if "day" not in df.columns:
            df["day"] = df["process_dt"].dt.strftime("%a")

        return df

    def _ensure_previous_baseline_for_engine(self) -> str:
        """
        Ensure outputs/prev_baseline_latest.parquet exists with a BasePlan column.
        Auto-Pilot and the baseline engine require this file before running.
        """
        latest_path = os.path.join("outputs", "prev_baseline_latest.parquet")
        os.makedirs("outputs", exist_ok=True)

        if os.path.exists(latest_path):
            cached = normalize_base_plan_columns(pd.read_parquet(latest_path))
            if "BasePlan" in cached.columns and not cached.empty:
                if "BasePlan" not in pd.read_parquet(latest_path).columns:
                    cached.to_parquet(latest_path, index=False)
                return latest_path

        sheets_manager = GoogleSheetsManager()
        params = sheets_manager.read_pipeline_params()
        now = datetime.now()
        try:
            target_week = int(params.get("target_week", now.isocalendar()[1]))
        except (TypeError, ValueError):
            target_week = now.isocalendar()[1]
        try:
            target_year = int(params.get("target_year", now.year))
        except (TypeError, ValueError):
            target_year = now.year

        prev_baseline = self._fetch_previous_baseline(target_week, target_year)
        if prev_baseline is None or prev_baseline.empty:
            raise ValueError(
                f"Previous baseline not found for Week {target_week} / {target_year}. "
                "Check RDS cache (6w_v3) and pipeline target_week/target_year parameters."
            )

        prev_baseline = normalize_base_plan_columns(prev_baseline)
        if "BasePlan" not in prev_baseline.columns:
            raise ValueError(
                "Previous baseline loaded but has no BasePlan column. "
                f"Available columns: {prev_baseline.columns.tolist()}"
            )

        week_cache = os.path.join(
            "outputs", "prev_baseline_cache",
            f"prev_baseline_wk{target_week}_yr{target_year}.parquet",
        )
        os.makedirs(os.path.dirname(week_cache), exist_ok=True)
        prev_baseline.to_parquet(week_cache, index=False)
        prev_baseline.to_parquet(latest_path, index=False)
        return latest_path

    def _fetch_previous_baseline(self, week_number, year_number=None):
        """
        Fetch previous baseline data for a specific week (and optionally year).
        Uses parquet-cached 6w_v3.rds for speed.
        """
        try:
            full_df = self._load_rds_cached_baseline()
            st.info(f"✓ Loaded {len(full_df):,} total records | Filtering Week {week_number}" +
                    (f" / {year_number}" if year_number else "") + "…")

            week_mask = full_df["Week"] == int(week_number)

            if year_number is not None:
                # Extract year from process_dt and apply filter
                year_mask = full_df["process_dt"].dt.year == int(year_number)
                Baseline_df = full_df[week_mask & year_mask].copy()

                # Warn if the year column produced no rows but the week alone had data
                if Baseline_df.empty:
                    _week_only = full_df[week_mask]
                    if not _week_only.empty:
                        _available_years = sorted(_week_only["process_dt"].dt.year.unique().tolist())
                        st.warning(
                            f"No data found for Week {week_number} in year {year_number}. "
                            f"Available years for this week: {_available_years}"
                        )
                    else:
                        st.warning(f"No data found for Week {week_number} in baseline RDS")
                    return None
            else:
                Baseline_df = full_df[week_mask].copy()
                if Baseline_df.empty:
                    st.warning(f"No data found for Week {week_number} in baseline RDS")
                    return None

            _year_label = f" / {year_number}" if year_number else ""
            st.info(f"✓ Found {len(Baseline_df):,} records for Week {week_number}{_year_label}")

            # Map Sub-category, product_name, sku class prod from local Product_Masters.xlsx
            _P_MASTER_PATH = FF_MASTERS_XLSX
            if os.path.exists(_P_MASTER_PATH):
                _pm_raw = pd.read_excel(_P_MASTER_PATH)

                # Normalise column names for flexible matching
                _col_map = {c: c.strip() for c in _pm_raw.columns}
                _pm_raw.rename(columns=_col_map, inplace=True)

                # Detect actual column names (handles "SKU Class" vs "SKU Class Prod" etc.)
                _id_col   = next((c for c in _pm_raw.columns if c.strip().lower() == "product id"), None)
                _cat_col  = next((c for c in _pm_raw.columns if "sub" in c.lower() and "cat" in c.lower()), None)
                _name_col = next((c for c in _pm_raw.columns if "product" in c.lower() and "name" in c.lower()), None)
                _sku_col  = next((c for c in _pm_raw.columns if "sku" in c.lower()), None)

                if not _id_col:
                    st.warning("⚠️ 'Product id' column not found in Product_Masters.xlsx")
                    for col in ["Sub-category", "product_name", "sku class prod"]:
                        Baseline_df[col] = ""
                else:
                    _pm = _pm_raw[[c for c in [_id_col, _cat_col, _name_col, _sku_col] if c]].copy()
                    _pm.rename(columns={
                        _id_col:   "product_id",
                        _cat_col:  "Sub-category",
                        _name_col: "product_name",
                        _sku_col:  "sku class prod",
                    }, inplace=True)
                    _pm = _pm.dropna(subset=["product_id"]).drop_duplicates(subset=["product_id"])

                    # Drop any existing mapped columns before merge to avoid _x/_y suffixes
                    Baseline_df.drop(columns=["Sub-category", "product_name", "sku class prod"], errors="ignore", inplace=True)
                    Baseline_df = Baseline_df.merge(_pm, on="product_id", how="left")

                    _mapped = Baseline_df["Sub-category"].notna().sum()
                    st.info(f"✓ Sub-category mapped for {_mapped:,} / {len(Baseline_df):,} rows from Product_Masters.xlsx")
            else:
                st.warning(f"⚠️ Product_Masters.xlsx not found at: {_P_MASTER_PATH}")
                for col in ["Sub-category", "product_name", "sku class prod"]:
                    Baseline_df[col] = ""

            Baseline_df = normalize_base_plan_columns(Baseline_df)
            
            # Select final columns
            st.info("📋 Selecting final columns...")
            final_cols = [
                "process_dt",
                "Sub-category",
                "Week",
                "day",
                "product_id",
                "product_name",
                "city_name",
                "hub_name",
                "BasePlan",
                "sku class prod"
            ]
            
            # Only keep columns that exist
            final_cols = [col for col in final_cols if col in Baseline_df.columns]
            Baseline_df = Baseline_df[final_cols]

            # Deduplicate at hub × product_id × date level
            _before_dedup = len(Baseline_df)
            _dedup_keys   = [c for c in ["hub_name", "product_id", "process_dt"] if c in Baseline_df.columns]
            if _dedup_keys:
                Baseline_df = Baseline_df.drop_duplicates(subset=_dedup_keys, keep="first").reset_index(drop=True)
            _dupes_removed = _before_dedup - len(Baseline_df)
            if _dupes_removed:
                st.info(f"🗑️ Removed {_dupes_removed:,} duplicate rows (hub × product_id × date)")

            st.success(f"✅ Previous baseline ready: {len(Baseline_df):,} records")

            # Display basic statistics including sum of BasePlan
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Products", f"{Baseline_df['product_id'].nunique():,}")
            with col2:
                st.metric("Hubs", f"{Baseline_df['hub_name'].nunique():,}")
            with col3:
                _bp_sum = Baseline_df['BasePlan'].sum() if 'BasePlan' in Baseline_df.columns else 0
                st.metric("Sum of BasePlan", f"{_bp_sum:,.0f}")
            
            # Show preview
            with st.expander("📊 Preview Baseline Data"):
                st.dataframe(
                    Baseline_df.head(50),
                    use_container_width=True,
                    height=300
                )
            
            return Baseline_df
        
        except FileNotFoundError:
            st.error(f"❌ Baseline RDS file not found. Please check the file path.")
            return None
        except Exception as e:
            st.error(f"❌ Error fetching previous baseline: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return None
    
    def _sync_specific_range(self, sheets_manager, worksheet_name, range_notation):
        """
        Sync a specific range from a worksheet
        """
        try:
            # Determine which spreadsheet based on worksheet name
            if worksheet_name == "Hub Sku Master":
                sheet_category = "hub_level_planning"
                worksheet_key = "hub_sku_master"
            else:
                sheet_category = "hub_level_planning"
                worksheet_key = worksheet_name.lower().replace(" ", "_")
            
            # Read the specific range
            df = sheets_manager.read_worksheet_to_df(
                sheet_category,
                worksheet_key,
                range_notation
            )
            
            return df
        except Exception as e:
            st.error(f"Error syncing {worksheet_name} [{range_notation}]: {str(e)}")
            return None
    
    def _check_data_quality(self, df):
        """
        Check for data quality issues (zero flags for city-days)
        """
        flag_columns = [
            'simple_flag_when_SP_0',
            'simple_instances_when_SP_0',
            'simple_group_flag_when_SP_0',
            'simple_group_instances_when_SP_0'
        ]
        
        # Check if all flag columns exist
        existing_flag_cols = [col for col in flag_columns if col in df.columns]
        
        if not existing_flag_cols:
            return {
                'has_issues': False,
                'total_issues': 0,
                'zero_flags': {'count': 0, 'details': pd.DataFrame()}
            }
        
        # Aggregate by city and day
        df_copy = df.copy()
        df_copy['process_dt'] = pd.to_datetime(df_copy['process_dt'])
        df_copy['date'] = df_copy['process_dt'].dt.date
        
        city_day = df_copy.groupby(['city_name', 'date']).agg({
            col: 'sum' for col in existing_flag_cols
        }).reset_index()
        
        # Find city-days where ALL flags are zero
        zero_mask = (city_day[existing_flag_cols] == 0).all(axis=1)
        zero_city_days = city_day[zero_mask]
        
        if len(zero_city_days) > 0:
            # Add more details
            zero_city_days['date'] = pd.to_datetime(zero_city_days['date'])
            zero_city_days['day_name'] = zero_city_days['date'].dt.day_name()
            zero_city_days['week'] = zero_city_days['date'].dt.isocalendar().week
            
            # Reorder columns for display
            display_cols = ['city_name', 'date', 'day_name', 'week'] + existing_flag_cols
            zero_city_days = zero_city_days[display_cols]
            
            return {
                'has_issues': True,
                'total_issues': len(zero_city_days),
                'zero_flags': {
                    'count': len(zero_city_days),
                    'details': zero_city_days
                }
            }
        
        return {
            'has_issues': False,
            'total_issues': 0,
            'zero_flags': {'count': 0, 'details': pd.DataFrame()}
        }
    
    def configure_parameters(self):
        """Configure baseline generation parameters"""
        st.subheader("⚙️ Configure Baseline Parameters")

        _ACTIVE_DS = os.path.join("outputs", "active_dataset.parquet")
        if not st.session_state.get("raw_data_loaded") and not os.path.exists(_ACTIVE_DS):
            st.warning("⚠️ Please load raw data first — open **1. Load Raw Data** in the sidebar.")
            return

        # Load parameters from Google Sheets
        from planning_suite.services.google_sheets import GoogleSheetsManager
        sheets_manager = GoogleSheetsManager()
        sheets_params = sheets_manager.read_pipeline_params()
        
        # Load values from sheets_params
        val_clustering = sheets_params.get("use_clustering", True)
        val_outliers = sheets_params.get("remove_outliers", True)
        val_hub_changes = sheets_params.get("apply_hub_changes", True)
        val_availability = sheets_params.get("use_availability", True)
        val_stf = sheets_params.get("use_stf", True)
        val_percentile = sheets_params.get("use_percentile", True)
        val_weeks_back = int(sheets_params.get("weeks_back", 4))
        val_avail_threshold = float(sheets_params.get("avail_threshold", 0.20))

        # Render interactive parameter editors
        st.markdown("#### ⚙️ Edit Parameters (Syncs to Google Sheet)")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            new_clustering = st.toggle("Use Clustering Phase Mapping", value=val_clustering)
            new_outliers = st.toggle("Remove Outliers", value=val_outliers)
            new_hub_changes = st.toggle("Apply Hub & KML Changes", value=val_hub_changes)
            new_availability = st.toggle("Use Availability Corrections", value=val_availability)
        with col_f2:
            new_stf = st.toggle("Use Sell-Through Factor (STF)", value=val_stf)
            new_percentile = st.toggle("Use Percentile Plans", value=val_percentile)
            new_weeks_back = st.number_input("Weeks Back (History Size)", min_value=1, max_value=52, value=val_weeks_back, step=1)
            new_avail_threshold = st.number_input("Availability Threshold", min_value=0.0, max_value=1.0, value=val_avail_threshold, step=0.01)

        # Detect changes and write back to Google Sheet
        has_changes = (
            new_clustering != val_clustering or
            new_outliers != val_outliers or
            new_hub_changes != val_hub_changes or
            new_availability != val_availability or
            new_stf != val_stf or
            new_percentile != val_percentile or
            new_weeks_back != val_weeks_back or
            new_avail_threshold != val_avail_threshold
        )

        if has_changes:
            updated_dict = {
                "use_clustering": new_clustering,
                "remove_outliers": new_outliers,
                "apply_hub_changes": new_hub_changes,
                "use_availability": new_availability,
                "use_stf": new_stf,
                "use_percentile": new_percentile,
                "weeks_back": new_weeks_back,
                "avail_threshold": new_avail_threshold
            }
            with st.spinner("Syncing updated parameters to Google Sheet..."):
                sheets_manager.write_pipeline_params(updated_dict)
            st.success("✅ Parameters synced and saved to Google Sheet!")
            st.rerun()

        # Display current configuration parameters loaded from Google Sheet
        st.markdown("---")
        st.markdown("#### 📋 Parameters currently loaded from Google Sheet")
        config_rows = []
        for var_name, val in sheets_params.items():
            config_rows.append({
                "Parameter": var_name,
                "Value": str(val),
                "Data Type": type(val).__name__
            })
        if config_rows:
            st.dataframe(pd.DataFrame(config_rows), use_container_width=True, hide_index=True)

        # Store in session state for downstream logic compatibility
        st.session_state.baseline_params = {
            'use_clustering': new_clustering,
            'remove_outliers': new_outliers,
            'apply_hub_changes': new_hub_changes,
            'use_availability': new_availability,
            'use_stf': new_stf,
            'use_percentile': new_percentile,
            'weeks_back': new_weeks_back,
            'avail_threshold': new_avail_threshold
        }

        st.caption(
            "🚀 **Hub changes** (KML remapping / new hub volume) are configured under **Master Data** → Hub Changes."
        )

        param_tabs = st.tabs(["Configuration Masters"])
        
        # Tab: Configuration Masters — single Google Sheet, sync to Excel
        with param_tabs[0]:
            st.markdown("### 📋 Configuration Masters")

            # DP_LOGICS_FOLDER is already imported from config

            # Worksheets to sync from that Google Sheet
            WORKSHEETS_TO_SYNC = [
                "City_Cat",
                "SellThroughFactor",
                "City_drops",
                "Percentile",
                "Avl_Flag",
            ]

            col1, col2 = st.columns([3, 1])
            with col1:
                st.info(f"**Source:** [DP Logics Google Sheet]({DP_LOGICS_SHEET_URL})")
                st.caption(f"**Save to:** `{DP_LOGICS_FOLDER}`")
                st.caption(f"**Worksheets:** {', '.join(WORKSHEETS_TO_SYNC)}")
            with col2:
                st.link_button("📝 Open Sheet", DP_LOGICS_SHEET_URL, use_container_width=True)

            st.markdown("---")

            # Show status of already saved Excel files
            st.markdown("#### 📁 Local Excel Files Status")
            os.makedirs(DP_LOGICS_FOLDER, exist_ok=True)
            status_rows = []
            for ws in WORKSHEETS_TO_SYNC:
                fpath = os.path.join(DP_LOGICS_FOLDER, f"{ws}.xlsx")
                if os.path.exists(fpath):
                    mtime = pd.Timestamp(os.path.getmtime(fpath), unit="s").strftime("%Y-%m-%d %H:%M")
                    status_rows.append({"Worksheet": ws, "Status": "✅ Saved", "Last Updated": mtime})
                else:
                    status_rows.append({"Worksheet": ws, "Status": "⬜ Not synced", "Last Updated": "—"})
            st.dataframe(pd.DataFrame(status_rows), use_container_width=True, hide_index=True)

            st.markdown("---")

            # Sync button
            if st.button("🔄 Sync All & Save as Excel", type="primary", use_container_width=True):
                try:
                    sheets_manager = self.sheets_manager
                    os.makedirs(DP_LOGICS_FOLDER, exist_ok=True)

                    from planning_suite.config import FF_MASTERS_XLSX
                    from planning_suite.services.baseline_io import refresh_all_engine_sidecars

                    sync_results = sheets_manager.sync_dp_logics_worksheets_to_folder(
                        DP_LOGICS_FOLDER,
                        WORKSHEETS_TO_SYNC,
                        allow_local_fallback=False,
                    )
                    sidecar_status = refresh_all_engine_sidecars(DP_LOGICS_FOLDER, FF_MASTERS_XLSX)

                    for ws_name, info in sync_results.items():
                        if info.get("status") == "synced":
                            st.caption(
                                f"✅ {ws_name}: {info.get('rows', 0):,} rows saved "
                                f"({info.get('source', 'google_sheets')})"
                            )
                        else:
                            st.warning(f"⚠️ {ws_name}: not synced")

                    if sidecar_status:
                        st.caption(
                            "Engine sidecars: "
                            + ", ".join(f"{k}={v}" for k, v in sidecar_status.items())
                        )

                    st.success(f"✅ All worksheets synced and saved to `{DP_LOGICS_FOLDER}`")
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Sync failed: {e}")
                    import traceback
                    with st.expander("🔍 Error Details"):
                        st.code(traceback.format_exc())


    def generate_baseline(self, user_id, *, read_only: bool = False):
        """Generate baseline forecast"""
        st.subheader("🚀 Generate Baseline")
        
        _ACTIVE_DS = os.path.join("outputs", "active_dataset.parquet")
        if not st.session_state.get("raw_data_loaded") and not os.path.exists(_ACTIVE_DS):
            st.warning("⚠️ Please load raw data first — open **1. Load Raw Data** in the sidebar.")
            return

        if 'baseline_params' not in st.session_state:
            st.warning("⚠️ Please configure parameters first (Step 2)")
            return
        
        # Week Selection for Previous Baseline
        st.markdown("### 📅 Week Configuration")
        
        col1, col2 = st.columns([1, 2])
        
        # Load parameters from Google Sheets
        from planning_suite.services.google_sheets import GoogleSheetsManager
        sheets_manager = GoogleSheetsManager()
        sheets_params = sheets_manager.read_pipeline_params()

        with col1:
            # Get current week and year
            current_week = datetime.now().isocalendar()[1]
            current_year = datetime.now().year

            target_week_val = sheets_params.get("target_week", current_week)
            try:
                target_week_val = int(target_week_val)
            except ValueError:
                target_week_val = current_week

            target_week = st.number_input(
                "Select Week Number for Baseline",
                min_value=1,
                max_value=53,
                value=target_week_val,
                step=1,
                help="Enter the week number for which you want to generate baseline"
            )

            target_year_val = sheets_params.get("target_year", current_year)
            try:
                target_year_val = int(target_year_val)
            except ValueError:
                target_year_val = current_year

            target_year = st.number_input(
                "Select Year for Baseline",
                min_value=2020,
                max_value=current_year + 1,
                value=target_year_val,
                step=1,
                help="Filter baseline by year — avoids mixing 2025 and 2026 data for the same week"
            )

            # Auto-save changes back to the parameters Google Sheet
            if int(target_week) != target_week_val or int(target_year) != target_year_val:
                with st.spinner("Saving week/year parameters to Google Sheet..."):
                    sheets_manager.write_pipeline_params({
                        "target_week": int(target_week),
                        "target_year": int(target_year)
                    })
                st.success("✅ Saved week/year configurations to Google Sheet!")
                st.rerun()

            st.caption(f"Current week: {current_week}  ·  Year: {current_year}")

        _PB_CACHE_DIR  = os.path.join("outputs", "prev_baseline_cache")
        _pb_cache_path = os.path.join(_PB_CACHE_DIR, f"prev_baseline_wk{target_week}_yr{target_year}.parquet")
        os.makedirs(_PB_CACHE_DIR, exist_ok=True)

        with col2:
            st.info("**Previous Baseline Fetch**")
            st.caption("Fetches baseline from same 6w_v3.rds file for the selected week")
            st.caption("💾 Saves Excel: `Baseline Wk{N} 2026.xlsx`")

            # Only store the parquet PATH in session_state — not the DataFrame
            _baseline_key = (target_week, target_year)
            if st.session_state.get("baseline_week_year") != _baseline_key:
                if os.path.exists(_pb_cache_path):
                    st.session_state.pb_cache_path      = _pb_cache_path
                    st.session_state.baseline_week_year = _baseline_key
                    st.session_state["_pb_from_cache"]  = True

            if st.button("📥 Fetch Previous Baseline", type="secondary", use_container_width=True):
                with st.spinner(f"Fetching previous baseline for Week {target_week} / {target_year}..."):
                    try:
                        prev_baseline = self._fetch_previous_baseline(int(target_week), int(target_year))
                        
                        if prev_baseline is not None and not prev_baseline.empty:
                            # Save to disk, store only path in session_state
                            try:
                                prev_baseline.to_parquet(_pb_cache_path, index=False)
                                _latest_path = os.path.join("outputs", "prev_baseline_latest.parquet")
                                prev_baseline.to_parquet(_latest_path, index=False)
                            except Exception:
                                pass
                            st.session_state.pb_cache_path      = _pb_cache_path
                            st.session_state.baseline_week_year = _baseline_key
                            st.session_state["_pb_from_cache"]  = False
                            st.success(f"✅ Loaded Week {target_week} / {target_year}: {len(prev_baseline):,} records — cached for next time")
                            col_a, col_b, col_c = st.columns(3)
                            with col_a:
                                st.metric("Products", f"{prev_baseline['product_id'].nunique():,}")
                            with col_b:
                                st.metric("Hubs", f"{prev_baseline['hub_name'].nunique():,}")
                            with col_c:
                                if 'BasePlan' in prev_baseline.columns:
                                    st.metric("Total Plan", f"{prev_baseline['BasePlan'].sum():,.0f}")
                            del prev_baseline  # free memory
                        else:
                            st.warning(f"⚠️ No previous baseline found for Week {target_week}")
                    
                    except Exception as e:
                        st.error(f"❌ Error fetching previous baseline: {str(e)}")
                        import traceback
                        with st.expander("🔍 Error Details"):
                            st.code(traceback.format_exc())
        
        # Show loaded previous baseline — read from disk on demand
        _pb_path = st.session_state.get("pb_cache_path")
        if _pb_path and os.path.exists(_pb_path) and st.session_state.get("baseline_week_year") == _baseline_key:
            _from_cache = st.session_state.get("_pb_from_cache", False)
            _src_label  = "loaded from cache" if _from_cache else "freshly fetched"
            from datetime import datetime as _dt
            _mtime = _dt.fromtimestamp(os.path.getmtime(_pb_path)).strftime('%d %b %Y %H:%M')
            with st.expander(f"📊 Previous Baseline — Week {target_week} ({target_year})  ({_src_label})  ·  cache: {_mtime}"):
                with st.spinner("Loading..."):
                    _pb_df = pd.read_parquet(_pb_path)
                col_a, col_b, col_c = st.columns(3)
                with col_a: st.metric("Products", f"{_pb_df['product_id'].nunique():,}")
                with col_b: st.metric("Hubs",     f"{_pb_df['hub_name'].nunique():,}")
                with col_c:
                    if "BasePlan" in _pb_df.columns:
                        st.metric("Sum of BasePlan", f"{_pb_df['BasePlan'].sum():,.0f}")
                st.dataframe(_pb_df.head(200), use_container_width=True, height=320)
                st.caption(f"Showing first 200 of {len(_pb_df):,} records")
                del _pb_df  # free memory
        
        st.markdown("---")
        
        st.markdown("### Review Configuration")
        
        params = st.session_state.baseline_params
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Enabled Features:**")
            for key, value in params.items():
                if value:
                    st.write(f"✅ {key.replace('_', ' ').title()}")
        
        with col2:
            st.write("**Data Summary:**")
            st.write(f"📊 Raw Data Rows: {st.session_state.get('raw_data_rows', '—'):,}" if st.session_state.get('raw_data_rows') else "📊 Raw Data Rows: —")
            st.write(f"📁 File: {st.session_state.get('raw_data_file', 'N/A')}")
        
        st.markdown("---")

        # Run Baseline Script
        st.markdown("### ▶️ Run Baseline")
        st.info(
            "Clicking the button below will execute **optimized_baseline_avail_correction.py** end-to-end "
            "and save the output as `Summary_<timestamp>.xlsx` in the Outputs folder."
        )

        _OUTPUTS_FOLDER = BASELINE_OUTPUTS_FOLDER
        _SCRIPT_PATH = str(PROJECT_ROOT / "scripts" / "optimized_baseline_avail_correction.py")
        _ACTIVE_DS = os.path.abspath(os.path.join("outputs", "active_dataset.parquet"))

        # Visible confirmation: exactly what data baseline will use
        # Demo mode banner
        _demo_city  = st.session_state.get("demo_city_filter", "All Cities")
        _demo_hubs  = st.session_state.get("demo_hub_filter") or []
        _is_demo    = (_demo_city and _demo_city != "All Cities") or bool(_demo_hubs)

        if _is_demo:
            _demo_parts = []
            if _demo_city and _demo_city != "All Cities":
                _demo_parts.append(f"City: <strong>{_demo_city}</strong>")
            if _demo_hubs:
                _demo_parts.append(f"Hubs: <strong>{', '.join(_demo_hubs)}</strong>")
            st.markdown(
                f"""<div style="padding:0.6rem 1rem; background:rgba(234,179,8,0.12);
                                border:1px solid rgba(234,179,8,0.5); border-radius:6px;
                                font-size:0.85rem; font-weight:600; color:#92400E; margin-bottom:0.75rem;">
                        Demo Mode active — {' &nbsp;·&nbsp; '.join(_demo_parts)}
                    </div>""",
                unsafe_allow_html=True
            )

        if os.path.exists(_ACTIVE_DS):
            try:
                _meta_df = pd.read_parquet(_ACTIVE_DS, columns=["week"])
                _weeks = sorted(pd.to_numeric(_meta_df["week"], errors="coerce").dropna().astype(int).unique().tolist()) if "week" in _meta_df.columns else []
                _rows = st.session_state.get("raw_data_rows")
                if _rows is None:
                    _rows = len(pd.read_parquet(_ACTIVE_DS, columns=["process_dt"]))
                _weeks_txt  = ", ".join([f"Wk {w}" for w in _weeks]) if _weeks else "Unknown"
                _city_note  = f" | City: **{_demo_city}**" if (_demo_city and _demo_city != "All Cities") else ""
                _hub_note   = f" | Hubs: **{', '.join(_demo_hubs)}**" if _demo_hubs else ""
                st.success(
                    f"✅ This run will use only selected UI dataset: **{_rows:,} rows** | **Weeks: {_weeks_txt}**{_city_note}{_hub_note}"
                )
            except Exception:
                st.warning("⚠️ Active dataset found, but could not read week metadata. Run will still use UI-selected dataset only.")
        else:
            st.error("❌ Active dataset not found. Open **1. Load Raw Data** and click **Load Selected Weeks** before running baseline.")
            return

        # Show existing summaries
        if os.path.isdir(_OUTPUTS_FOLDER):
            _existing = sorted(
                [f for f in os.listdir(_OUTPUTS_FOLDER) if f.startswith("Summary_") and f.endswith(".xlsx")],
                reverse=True
            )
            if _existing:
                with st.expander(f"📂 Existing Summaries ({len(_existing)} files)", expanded=False):
                    for _f in _existing[:10]:
                        _fp = os.path.join(_OUTPUTS_FOLDER, _f)
                        _mtime = datetime.fromtimestamp(os.path.getmtime(_fp)).strftime("%Y-%m-%d %H:%M")
                        st.write(f"📄 `{_f}` — {_mtime}")

        if read_only:
            st.warning("You have read-only access. Baseline runs are disabled for your role.")
            st.button("🚀 Run Baseline & Save Summary", type="primary", use_container_width=True, disabled=True)
            return

        if st.button("🚀 Run Baseline & Save Summary", type="primary", use_container_width=True):
            import subprocess, sys, copy

            _env = copy.copy(os.environ)
            _env["PYTHONIOENCODING"] = "utf-8"
            _env["PYTHONUTF8"] = "1"

            _active_ds, _env, _filter_error = self._prepare_demo_filter_dataset(_env)
            if _filter_error:
                st.error(_filter_error)
                return

            run_id = generate_run_id("BL")
            st.session_state.baseline_run_id = run_id
            run_name = f"Baseline Wk{target_week} {target_year}"
            self.db.save_baseline_run(
                {
                    "run_id": run_id,
                    "run_name": run_name,
                    "user_id": user_id,
                    "status": "running",
                    "raw_data_file": st.session_state.get("raw_data_file", _ACTIVE_DS),
                    "parameters": st.session_state.get("baseline_params", {}),
                }
            )
            with st.spinner("Running Baseline(Avail_correction).py… this may take several minutes."):
                try:
                    _env["BASELINE_USE_ACTIVE_DATASET"] = "1"
                    _env["BASELINE_ACTIVE_DATASET_PATH"] = _active_ds
                    _env["PROJECT_ROOT"] = str(PROJECT_ROOT)

                    _result = subprocess.run(
                        [sys.executable, _SCRIPT_PATH],
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",   # never crash on undecodable bytes
                        timeout=3600,
                        env=_env
                    )
                    _stdout = _result.stdout or ""
                    _stderr = _result.stderr or ""

                    if _result.returncode == 0:
                        _saved = sorted(
                            [f for f in os.listdir(_OUTPUTS_FOLDER) if f.startswith("Summary_") and f.endswith(".xlsx")],
                            key=lambda f: os.path.getmtime(os.path.join(_OUTPUTS_FOLDER, f)),
                            reverse=True
                        ) if os.path.isdir(_OUTPUTS_FOLDER) else []
                        _latest = _saved[0] if _saved else None
                        _output_path = os.path.join(_OUTPUTS_FOLDER, _latest) if _latest else ""
                        self.db.update_baseline_run(
                            run_id,
                            status="completed",
                            output_file=_output_path,
                            summary_stats={"week": target_week, "year": target_year},
                        )
                        notify_baseline_run_finished(
                            run_id=run_id,
                            run_name=run_name,
                            status="completed",
                            user_id=user_id,
                            db=self.db,
                        )
                        st.success(f"✅ Baseline completed!" + (f" Saved as **{_latest}**" if _latest else ""))

                        # Surface any step-level errors even when returncode=0
                        _warn_lines = [ln for ln in _stdout.splitlines()
                                       if any(kw in ln for kw in ["ERROR", "Skipping", "failed", "WARN"])]
                        if _warn_lines:
                            st.warning("⚠️ Some steps had issues (see details below):")
                            for _wl in _warn_lines:
                                st.caption(_wl)
                    else:
                        _fail_params = {"stderr": _stderr[-2000:], "stdout_tail": _stdout[-2000:]}
                        self.db.update_baseline_run(
                            run_id,
                            status="failed",
                            parameters=_fail_params,
                        )
                        notify_baseline_run_finished(
                            run_id=run_id,
                            run_name=run_name,
                            status="failed",
                            user_id=user_id,
                            error_detail=format_error_from_parameters(_fail_params),
                            db=self.db,
                        )
                        st.error("❌ Baseline script exited with errors.")
                        if _stderr:
                            with st.expander("🔍 Error Details", expanded=True):
                                st.code(_stderr[-5000:])

                    # Always show full script output so errors are visible
                    if _stdout:
                        with st.expander("📋 Script Output (click to expand)", expanded=_result.returncode != 0):
                            st.code(_stdout[-8000:] if len(_stdout) > 8000 else _stdout)
                except subprocess.TimeoutExpired:
                    _fail_params = {"error": "timeout_60min"}
                    self.db.update_baseline_run(run_id, status="failed", parameters=_fail_params)
                    notify_baseline_run_finished(
                        run_id=run_id,
                        run_name=run_name,
                        status="failed",
                        user_id=user_id,
                        error_detail="Script timed out after 60 minutes.",
                        db=self.db,
                    )
                    st.error("❌ Script timed out after 60 minutes.")
                except Exception as _e:
                    _fail_params = {"error": str(_e)}
                    self.db.update_baseline_run(run_id, status="failed", parameters=_fail_params)
                    notify_baseline_run_finished(
                        run_id=run_id,
                        run_name=run_name,
                        status="failed",
                        user_id=user_id,
                        error_detail=str(_e),
                        db=self.db,
                    )
                    st.error(f"❌ Failed to run script: {_e}")
                    import traceback
                    with st.expander("🔍 Traceback"):
                        st.code(traceback.format_exc())
    
    def _run_baseline_generation(self, raw_df, masters, params):
        """Core baseline generation logic"""
        progress = st.progress(0)
        status = st.empty()
        
        # Step 1: Clustering (if enabled)
        status.text("Step 1/7: Processing hub clustering...")
        merged_df = raw_df.copy()
        
        if params['use_clustering'] and 'cluster_mapping' in masters:
            merged_df = self._apply_clustering(merged_df, masters['cluster_mapping'])
        progress.progress(15)
        
        # Step 2: Availability Flag
        status.text("Step 2/7: Calculating availability...")
        if params['use_availability'] and 'avl_flag' in masters:
            merged_df = self._calculate_availability(merged_df, masters['avl_flag'])
        progress.progress(30)
        
        # Step 3: Hub Changes
        status.text("Step 3/7: Applying hub changes...")
        if params['apply_hub_changes'] and 'hub_changes' in masters:
            merged_df = self._apply_hub_changes(merged_df, masters['hub_changes'])
        progress.progress(45)
        
        # Step 4: Outlier Removal
        status.text("Step 4/7: Removing outlier days...")
        if params['remove_outliers'] and 'outlier' in masters:
            merged_df = self._remove_outliers(merged_df, masters['outlier'])
        progress.progress(60)
        
        # Step 5: Availability Correction
        status.text("Step 5/7: Applying availability correction...")
        if params['use_availability']:
            pivot_df = self._apply_availability_correction(merged_df, masters)
        progress.progress(75)
        
        # Step 6: City Drops
        status.text("Step 6/7: Applying city drops...")
        if 'city_drops' in masters:
            pivot_df = self._apply_city_drops(pivot_df, masters['city_drops'])
        progress.progress(85)
        
        # Step 7: Percentile & Final Plan
        status.text("Step 7/7: Calculating final plan...")
        if params['use_percentile'] and 'percentile' in masters:
            final_df = self._calculate_percentile_plan(pivot_df, masters['percentile'])
        progress.progress(100)
        
        status.text("✅ Baseline generation complete!")
        
        # Calculate summary statistics
        summary_stats = self._calculate_summary_stats(final_df)
        
        return final_df, summary_stats
    
    def _apply_clustering(self, df, cluster_df):
        """Apply hub clustering logic"""
        # Implementation from Baseline.py lines 49-189
        cluster_df = cluster_df.copy()
        cluster_df["Cluster_Flag"] = cluster_df["Cluster_Flag"].astype(int)
        cluster_df = cluster_df[cluster_df["Cluster_Flag"] == 1]
        
        df = df.merge(
            cluster_df[["product_id", "Mother_hubid", "MotherHub_name", "childHub_name"]],
            left_on=["product_id", "hub_name"],
            right_on=["product_id", "childHub_name"],
            how="left"
        )
        
        # Aggregate child hubs to mother hubs
        # ... (simplified for brevity - full implementation would go here)
        
        return df
    
    def _calculate_availability(self, df, avl_flag_df):
        """Calculate simple availability"""
        # Implementation from Baseline.py lines 209-253
        df = df.merge(
            avl_flag_df[['product_id', 'Avl Flag']],
            how='left',
            on='product_id'
        )
        
        df['Avl Flag'] = df['Avl Flag'].fillna(0).astype(int)
        
        df['simple_avail_num'] = df.apply(
            lambda row: row['simple_flag_when_SP_0'] if row['Avl Flag'] == 1 
            else row['simple_grp_flag_when_SP_0'], 
            axis=1
        )
        
        df['simple_avail_den'] = df.apply(
            lambda row: row['simple_instances_when_SP_0'] if row['Avl Flag'] == 1 
            else row['simple_grp_instances_when_SP_0'],
            axis=1
        )
        
        df['simple_availability'] = np.where(
            (df['simple_avail_num'] == 0) & (df['simple_avail_den'] == 0),
            0,
            np.where(
                (df['simple_avail_num'] == 0) | (df['simple_avail_den'] == 0),
                0,
                (df['simple_avail_num'] / df['simple_avail_den']) * 100
            )
        )
        
        return df
    
    def _apply_hub_changes(self, df, hub_changes_df):
        """Apply new hub launches and KML remapping"""
        # Implementation from Baseline.py lines 286-570
        st.info("Applying hub changes...")
        return df  # Simplified for now
    
    def _remove_outliers(self, df, outlier_df):
        """Remove outlier days"""
        # Implementation from Baseline.py lines 574-613
        outlier_df['process_dt'] = pd.to_datetime(outlier_df['process_dt'], errors='coerce')
        df['process_dt'] = pd.to_datetime(df['process_dt'], errors='coerce')
        
        df = df.merge(
            outlier_df[['city_name', 'sub category', 'process_dt', 'Outlier_Flag']],
            on=['city_name', 'sub category', 'process_dt'],
            how='left'
        )
        
        df['Outlier_Flag'] = pd.to_numeric(df['Outlier_Flag'], errors='coerce').fillna(0).astype(int)
        df.loc[df['Outlier_Flag'] == 1, ['Sales (qty)', 'simple_avail_num']] = 0
        
        return df
    
    def _apply_availability_correction(self, df, masters):
        """Apply availability and sell-through correction"""
        # Simplified version - full implementation would include all the pivot logic
        return df
    
    def _apply_city_drops(self, df, city_drops_df):
        """Apply city-level percentage drops"""
        # Implementation from Baseline.py lines 813-890
        return df
    
    def _calculate_percentile_plan(self, df, percentile_df):
        """Calculate suggested plan using percentile logic"""
        # Implementation from Baseline.py lines 997-1351
        return df
    
    def _calculate_summary_stats(self, df):
        """Calculate summary statistics for the baseline"""
        stats = {
            'total_records': len(df),
            'unique_products': df['product_id'].nunique() if 'product_id' in df.columns else 0,
            'unique_hubs': df['hub_name'].nunique() if 'hub_name' in df.columns else 0,
            'unique_cities': df['city_name'].nunique() if 'city_name' in df.columns else 0,
            'total_plan_qty': df['sugg_plan'].sum() if 'sugg_plan' in df.columns else 0,
            'avg_plan_qty': df['sugg_plan'].mean() if 'sugg_plan' in df.columns else 0
        }
        return stats
    
    def review_baseline(self, user_id):
        """Review generated baseline"""
        st.subheader("📊 Review Baseline Output")

        _OUTPUTS_FOLDER = BASELINE_OUTPUTS_FOLDER
        _summary_files  = sorted(
            [f for f in os.listdir(_OUTPUTS_FOLDER) if f.startswith("Summary_") and f.endswith(".xlsx")],
            key=lambda f: os.path.getmtime(os.path.join(_OUTPUTS_FOLDER, f)),
            reverse=True
        ) if os.path.isdir(_OUTPUTS_FOLDER) else []

        if _summary_files:
            _latest = _summary_files[0]
            _latest_path = os.path.join(_OUTPUTS_FOLDER, _latest)
            st.success(f"✅ Latest baseline: **{_latest}**")

            with st.expander("📋 Preview Latest Summary File"):
                try:
                    _preview_df = pd.read_excel(_latest_path)
                    # Fix mixed-type object columns so Arrow can serialise them
                    for _col in _preview_df.select_dtypes(include="object").columns:
                        _preview_df[_col] = _preview_df[_col].astype(str)
                    st.dataframe(_preview_df.head(500), use_container_width=True, height=350)
                    st.caption(f"{len(_preview_df):,} rows × {len(_preview_df.columns)} columns")
                    _csv = _preview_df.to_csv(index=False).encode("utf-8")
                    st.download_button("📥 Download as CSV", _csv, file_name=_latest.replace(".xlsx", ".csv"), mime="text/csv")
                except Exception as _e:
                    st.error(f"Could not read summary file: {_e}")
        else:
            st.info("No summary file found yet. Run the baseline in Step 3 to generate one.")

        st.markdown("---")

        # ── Base Plan Comparison ──────────────────────────────────────────────
        st.subheader("📊 Base Plan: Previous vs Current")
        st.caption("Hub × SKU Class Prod × Day — pasted to 'Baseline' tab in Google Sheet.")

        _SUMMARY_FOLDER = BASELINE_OUTPUTS_FOLDER
        _LOG_FOLDER     = DP_LOGICS_FOLDER
        _VAL_URL = VALIDATION_SHEET_URL

        if st.button("🔄 Load Comparison", type="primary", use_container_width=True):
            try:
                from planning_suite.services.baseline_comparison import (
                    build_hub_sku_day_comparison,
                    resolve_hub_suggestion_previous,
                    resolve_latest_summary,
                )

                _CMP_CACHE_DIR = os.path.join("outputs", "cmp_cache")
                _sum_df, _sum_files_name, _sum_path = resolve_latest_summary(
                    _SUMMARY_FOLDER, cache_dir=_CMP_CACHE_DIR
                )
                _log_df, _log_source, _log_path = resolve_hub_suggestion_previous(
                    log_folder=_LOG_FOLDER, cache_dir=_CMP_CACHE_DIR
                )

                _log_label = os.path.basename(_log_path) if _log_path else _log_source
                st.info(f"**Current**: {_sum_files_name}  |  **Previous**: {_log_label}")

                _cmp = build_hub_sku_day_comparison(_sum_df, _log_df)

                st.session_state["_bpcmp_baseline"] = _cmp
                # Persist to disk so it survives refresh
                try:
                    os.makedirs("outputs", exist_ok=True)
                    _cmp.to_parquet(os.path.join("outputs", "cmp_baseline_latest.parquet"), index=False)
                except Exception:
                    pass

                # Write to "Baseline" tab in Google Sheet
                _gs_status = st.empty()
                try:
                    from planning_suite.services.google_sheets import GoogleSheetsManager
                    sheets_manager = GoogleSheetsManager()
                    _vss = sheets_manager.gc.open_by_url(_VAL_URL)
                    try:
                        _vws = _vss.worksheet("Baseline")
                    except gspread.exceptions.WorksheetNotFound:
                        _vws = _vss.add_worksheet(title="Baseline", rows=max(len(_cmp) + 100, 500), cols=10)
                    _vws.clear()
                    set_with_dataframe(_vws, _cmp)
                    _gs_status.empty()
                    st.session_state["_gs_write_ok"] = True
                except Exception as _gs_err:
                    _gs_status.empty()
                    st.session_state["_gs_write_err"] = str(_gs_err)

                st.rerun()
            except Exception as _e:
                st.error(f"❌ {_e}")
                import traceback
                st.code(traceback.format_exc())

        # Show Google Sheet write status after rerun
        if st.session_state.pop("_gs_write_ok", False):
            st.success("✅ Comparison written to 'Baseline' tab in Google Sheet!")
        _gs_err_msg = st.session_state.pop("_gs_write_err", None)
        if _gs_err_msg:
            st.warning(f"⚠️ Google Sheet write failed: {_gs_err_msg}")

        # Auto-restore comparison table from disk after refresh
        _CMP_CACHE = os.path.join("outputs", "cmp_baseline_latest.parquet")
        if "_bpcmp_baseline" not in st.session_state and os.path.exists(_CMP_CACHE):
            try:
                st.session_state["_bpcmp_baseline"] = pd.read_parquet(_CMP_CACHE)
            except Exception:
                pass

        # Display comparison table
        if "_bpcmp_baseline" in st.session_state and st.session_state["_bpcmp_baseline"] is not None:
            _df = st.session_state["_bpcmp_baseline"]
            _prev_tot = int(_df["Previous Baseline"].sum())
            _curr_tot = int(_df["Current Baseline"].sum())
            _pct_tot  = round((_curr_tot - _prev_tot) / _prev_tot * 100, 1) if _prev_tot != 0 else 0
            _mc1, _mc2, _mc3 = st.columns(3)
            with _mc1: st.metric("Previous Total", f"{_prev_tot:,.0f}")
            with _mc2: st.metric("Current Total",  f"{_curr_tot:,.0f}", delta=f"{_curr_tot - _prev_tot:+,.0f}")
            with _mc3: st.metric("Overall Delta %", f"{_pct_tot:+.1f}%")
            st.dataframe(
                _df,
                column_config={
                    "Previous Baseline": st.column_config.NumberColumn("Previous Baseline", format="%,.0f"),
                    "Current Baseline":  st.column_config.NumberColumn("Current Baseline",  format="%,.0f"),
                    "Delta %":           st.column_config.NumberColumn("Delta %",           format="%.1f%%"),
                },
                use_container_width=True,
                height=min(60 + len(_df) * 36, 600),
            )
            st.caption(f"{len(_df):,} rows")
        else:
            st.info("Click 'Load Comparison' to view data.")

        # ── Review Updated Baseline (Multi-level) ────────────────────────────
        st.markdown("---")
        st.subheader("📊 Review Updated Baseline")
        st.caption(
            "Compare current vs previous baseline at four granularity levels. "
            "Conditional formatting highlights meaningful changes."
        )

        # ── Threshold configuration ───────────────────────────────────────────
        with st.expander("⚙️ Configure Change Thresholds", expanded=False):
            _rc1, _rc2, _rc3, _rc4 = st.columns(4)
            with _rc1:
                st.number_input(
                    "Strong ↓ (%)", value=-20, max_value=-1, step=1,
                    key="rv_thresh_neg_strong",
                    help="Delta % below this threshold → Red background",
                )
            with _rc2:
                st.number_input(
                    "Moderate ↓ (%)", value=-10, max_value=-1, step=1,
                    key="rv_thresh_neg_mod",
                    help="Delta % below this (but above Strong ↓) → Orange background",
                )
            with _rc3:
                st.number_input(
                    "Moderate ↑ (%)", value=10, min_value=1, step=1,
                    key="rv_thresh_pos_mod",
                    help="Delta % above this (but below Strong ↑) → Amber background",
                )
            with _rc4:
                st.number_input(
                    "Strong ↑ (%)", value=20, min_value=1, step=1,
                    key="rv_thresh_pos_strong",
                    help="Delta % above this threshold → Green background",
                )

        _rv_t_ns = st.session_state.get("rv_thresh_neg_strong", -20)
        _rv_t_nm = st.session_state.get("rv_thresh_neg_mod",    -10)
        _rv_t_pm = st.session_state.get("rv_thresh_pos_mod",     10)
        _rv_t_ps = st.session_state.get("rv_thresh_pos_strong",  20)

        # ── Load button ───────────────────────────────────────────────────────
        if st.button(
            "🔄 Load Multi-level Comparison",
            key="rv_load_btn", type="primary", use_container_width=True,
        ):
            try:
                from planning_suite.services.baseline_comparison import (
                    CITY_ALIASES,
                    CURRENT_VALUE_ALIASES,
                    DAY_COLUMN_ALIASES,
                    HUB_ALIASES,
                    PREVIOUS_VALUE_ALIASES,
                    SKU_ALIASES,
                    build_comparison_view,
                    find_column as _rv_fc,
                    resolve_hub_suggestion_previous,
                    resolve_latest_summary,
                )

                _RV_SUM_FOLDER = BASELINE_OUTPUTS_FOLDER
                _RV_LOG_FOLDER = DP_LOGICS_FOLDER
                _RV_CACHE_DIR  = os.path.join("outputs", "cmp_cache")

                with st.spinner("Loading baseline files…"):
                    _rv_sum_df, _rv_sum_name, _rv_sum_path = resolve_latest_summary(
                        _RV_SUM_FOLDER, cache_dir=_RV_CACHE_DIR
                    )
                    _rv_log_df, _rv_log_source, _rv_log_path = resolve_hub_suggestion_previous(
                        log_folder=_RV_LOG_FOLDER, cache_dir=_RV_CACHE_DIR
                    )

                _rv_log_label = os.path.basename(_rv_log_path) if _rv_log_path else _rv_log_source
                st.info(f"**Current**: {_rv_sum_name}  |  **Previous**: {_rv_log_label}")

                _s_city = _rv_fc(_rv_sum_df, CITY_ALIASES)
                _s_hub  = _rv_fc(_rv_sum_df, HUB_ALIASES)
                _s_cat  = _rv_fc(_rv_sum_df, SKU_ALIASES)
                _s_day  = _rv_fc(_rv_sum_df, DAY_COLUMN_ALIASES)
                _s_fp   = _rv_fc(_rv_sum_df, CURRENT_VALUE_ALIASES)
                _l_city = _rv_fc(_rv_log_df, CITY_ALIASES)
                _l_hub  = _rv_fc(_rv_log_df, HUB_ALIASES)
                _l_cat  = _rv_fc(_rv_log_df, SKU_ALIASES)
                _l_day  = _rv_fc(_rv_log_df, DAY_COLUMN_ALIASES)
                _l_bp   = _rv_fc(_rv_log_df, PREVIOUS_VALUE_ALIASES)

                _missing_cols = [
                    n for n, v in [
                        ("Summary:hub_name", _s_hub), ("Summary:day", _s_day), ("Summary:Final_Plan", _s_fp),
                        ("Log:hub_name", _l_hub),     ("Log:day", _l_day),     ("Log:Base_plan", _l_bp),
                    ] if not v
                ]
                if _missing_cols:
                    st.error(f"Missing required columns: {', '.join(_missing_cols)}")
                    raise ValueError(f"Missing required columns: {_missing_cols}")

                # View 1 : City × Day
                _rv_v1 = (
                    build_comparison_view(
                        _rv_sum_df, _rv_log_df, _s_fp, _l_bp,
                        [_s_city, _s_day], [_l_city, _l_day],
                        ["City", "Day"],
                    ) if _s_city and _l_city else None
                )
                # View 2 : City × Category × Day
                _rv_v2 = (
                    build_comparison_view(
                        _rv_sum_df, _rv_log_df, _s_fp, _l_bp,
                        [_s_city, _s_cat, _s_day], [_l_city, _l_cat, _l_day],
                        ["City", "Category", "Day"],
                    ) if (_s_city and _l_city and _s_cat and _l_cat) else None
                )
                # View 3 : Hub × Category × Day
                _rv_v3 = (
                    build_comparison_view(
                        _rv_sum_df, _rv_log_df, _s_fp, _l_bp,
                        [_s_hub, _s_cat, _s_day], [_l_hub, _l_cat, _l_day],
                        ["Hub", "Category", "Day"],
                    ) if (_s_cat and _l_cat) else None
                )
                # View 4 : Hub × Day
                _rv_v4 = build_comparison_view(
                    _rv_sum_df, _rv_log_df, _s_fp, _l_bp,
                    [_s_hub, _s_day], [_l_hub, _l_day],
                    ["Hub", "Day"],
                )

                st.session_state["_rv_views"] = {
                    "v1": _rv_v1, "v2": _rv_v2, "v3": _rv_v3, "v4": _rv_v4,
                    "curr_file": _rv_sum_name,
                    "prev_file": _rv_log_label,
                }

                # Persist views to disk
                _rv_disk_dir = os.path.join("outputs", "rv_cache")
                os.makedirs(_rv_disk_dir, exist_ok=True)
                for _vname, _vdf in [("v1", _rv_v1), ("v2", _rv_v2), ("v3", _rv_v3), ("v4", _rv_v4)]:
                    if _vdf is not None:
                        try:
                            _vdf.to_parquet(os.path.join(_rv_disk_dir, f"{_vname}.parquet"), index=False)
                        except Exception:
                            pass

                st.rerun()
            except Exception as _rv_err:
                import traceback as _rv_tb
                st.error(f"❌ {_rv_err}")
                with st.expander("Error details"):
                    st.code(_rv_tb.format_exc())

        # Auto-restore views from disk on page refresh
        _rv_disk_dir = os.path.join("outputs", "rv_cache")
        if "_rv_views" not in st.session_state and os.path.isdir(_rv_disk_dir):
            _rv_restored = {}
            for _vname in ["v1", "v2", "v3", "v4"]:
                _vpath = os.path.join(_rv_disk_dir, f"{_vname}.parquet")
                _rv_restored[_vname] = (
                    pd.read_parquet(_vpath) if os.path.exists(_vpath) else None
                )
            if any(v is not None for v in _rv_restored.values()):
                st.session_state["_rv_views"] = _rv_restored

        # ── Display views ─────────────────────────────────────────────────────
        if "_rv_views" in st.session_state:
            _rv_views = st.session_state["_rv_views"]

            if _rv_views.get("curr_file"):
                st.info(
                    f"**Current Baseline:** {_rv_views['curr_file']}  "
                    f"|  **Previous Baseline:** {_rv_views.get('prev_file', '—')}"
                )

            # ── Color legend ──────────────────────────────────────────────────
            st.markdown(
                f"<div style='font-size:0.8rem; margin-bottom:0.5rem;'>"
                f"<span style='background:#FF4444;color:#fff;padding:2px 8px;border-radius:4px;margin-right:6px;'>Strong ↓ &lt; {_rv_t_ns}%</span>"
                f"<span style='background:#FF8C00;color:#fff;padding:2px 8px;border-radius:4px;margin-right:6px;'>Moderate ↓ &lt; {_rv_t_nm}%</span>"
                f"<span style='background:#F59E0B;color:#1a1a1a;padding:2px 8px;border-radius:4px;margin-right:6px;'>Moderate ↑ &gt; {_rv_t_pm}%</span>"
                f"<span style='background:#16A34A;color:#fff;padding:2px 8px;border-radius:4px;'>Strong ↑ &gt; {_rv_t_ps}%</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

            # Conditional-format helper
            def _rv_color_delta(val, t_ns, t_nm, t_pm, t_ps):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return ""
                v = float(val)
                if v < t_ns:
                    return "background-color: #FF4444; color: white"
                if v < t_nm:
                    return "background-color: #FF8C00; color: white"
                if v > t_ps:
                    return "background-color: #16A34A; color: white"
                if v > t_pm:
                    return "background-color: #F59E0B; color: #1a1a1a"
                return ""

            def _rv_apply_style(df, t_ns, t_nm, t_pm, t_ps):
                if "Delta %" not in df.columns:
                    return df.style
                try:
                    return df.style.map(
                        lambda v: _rv_color_delta(v, t_ns, t_nm, t_pm, t_ps),
                        subset=["Delta %"],
                    )
                except AttributeError:
                    return df.style.applymap(
                        lambda v: _rv_color_delta(v, t_ns, t_nm, t_pm, t_ps),
                        subset=["Delta %"],
                    )

            # Build tab list (only include views that exist)
            _rv_tab_spec = [
                ("v1", "City × Day"),
                ("v2", "City × Category × Day"),
                ("v3", "Hub × Category × Day"),
                ("v4", "Hub × Day"),
            ]
            _rv_available = [(k, lbl) for k, lbl in _rv_tab_spec if _rv_views.get(k) is not None]

            if not _rv_available:
                st.warning(
                    "No comparison views could be built. Verify that source files contain "
                    "city_name, hub_name, SKU Class Prod, day, and the baseline value columns."
                )
            else:
                _rv_tab_objs = st.tabs([lbl for _, lbl in _rv_available])

                for _rv_tab_obj, (_rv_vk, _rv_vlbl) in zip(_rv_tab_objs, _rv_available):
                    with _rv_tab_obj:
                        _rv_vdf = _rv_views[_rv_vk].copy()
                        _rv_vdf["Previous Baseline"] = pd.to_numeric(
                            _rv_vdf["Previous Baseline"], errors="coerce"
                        ).fillna(0)
                        _rv_vdf["Current Baseline"] = pd.to_numeric(
                            _rv_vdf["Current Baseline"], errors="coerce"
                        ).fillna(0)

                        # Summary metrics
                        _rv_prev_t = int(_rv_vdf["Previous Baseline"].sum())
                        _rv_curr_t = int(_rv_vdf["Current Baseline"].sum())
                        _rv_pct_t  = (
                            round((_rv_curr_t - _rv_prev_t) / _rv_prev_t * 100, 1)
                            if _rv_prev_t != 0 else 0.0
                        )
                        _rm1, _rm2, _rm3 = st.columns(3)
                        _rm1.metric("Previous Total",  f"{_rv_prev_t:,}")
                        _rm2.metric("Current Total",   f"{_rv_curr_t:,}",
                                    delta=f"{_rv_curr_t - _rv_prev_t:+,.0f}")
                        _rm3.metric("Overall Delta %", f"{_rv_pct_t:+.1f}%")

                        # Dimension filter on first key column
                        _rv_dim_cols = [
                            c for c in _rv_vdf.columns
                            if c not in {"Previous Baseline", "Current Baseline", "Delta", "Delta %"}
                        ]
                        _rv_filter_col = _rv_dim_cols[0] if _rv_dim_cols else None
                        _rv_display_df = _rv_vdf.copy()
                        if _rv_filter_col:
                            _rv_f_opts = ["All"] + sorted(
                                _rv_vdf[_rv_filter_col].dropna().astype(str).unique().tolist()
                            )
                            _rv_f_sel = st.selectbox(
                                f"Filter by {_rv_filter_col}", _rv_f_opts,
                                key=f"rv_filter_{_rv_vk}",
                            )
                            if _rv_f_sel != "All":
                                _rv_display_df = _rv_vdf[
                                    _rv_vdf[_rv_filter_col].astype(str) == _rv_f_sel
                                ].copy()

                        # Styled dataframe — fall back to plain display if too large for Styler
                        _rv_cell_count = _rv_display_df.shape[0] * _rv_display_df.shape[1]
                        _RV_STYLE_LIMIT = 200_000  # cells
                        if _rv_cell_count <= _RV_STYLE_LIMIT:
                            pd.set_option("styler.render.max_elements", _rv_cell_count + 1000)
                            _rv_styled = _rv_apply_style(
                                _rv_display_df, _rv_t_ns, _rv_t_nm, _rv_t_pm, _rv_t_ps
                            ).format(
                                {
                                    "Previous Baseline": "{:,.0f}",
                                    "Current Baseline":  "{:,.0f}",
                                    "Delta":             "{:+,.0f}",
                                    "Delta %": lambda x: (
                                        f"{x:+.1f}%"
                                        if x is not None and not (isinstance(x, float) and pd.isna(x))
                                        else "—"
                                    ),
                                },
                                na_rep="—",
                            )
                            st.dataframe(
                                _rv_styled,
                                use_container_width=True,
                                height=min(60 + len(_rv_display_df) * 36, 600),
                            )
                        else:
                            # Too large to style — render plain with column formatting
                            st.info(
                                f"ℹ️ Table has {_rv_cell_count:,} cells — "
                                "use the filter above to narrow down before conditional formatting is applied. "
                                "Showing plain view. Colour-coded Excel download still works for all rows."
                            )
                            _rv_fmt_df = _rv_display_df.copy()
                            if "Delta %" in _rv_fmt_df.columns:
                                _rv_fmt_df["Delta %"] = _rv_fmt_df["Delta %"].apply(
                                    lambda x: f"{x:+.1f}%" if x is not None and not (
                                        isinstance(x, float) and pd.isna(x)
                                    ) else "—"
                                )
                            st.dataframe(
                                _rv_fmt_df,
                                column_config={
                                    "Previous Baseline": st.column_config.NumberColumn(
                                        "Previous Baseline", format="%,.0f"
                                    ),
                                    "Current Baseline": st.column_config.NumberColumn(
                                        "Current Baseline", format="%,.0f"
                                    ),
                                    "Delta": st.column_config.NumberColumn(
                                        "Delta", format="%+,.0f"
                                    ),
                                },
                                use_container_width=True,
                                height=min(60 + len(_rv_display_df) * 36, 600),
                            )
                        st.caption(f"{len(_rv_display_df):,} rows")

                # ── Excel Download ────────────────────────────────────────────
                st.markdown("---")

                def _rv_generate_excel(views_dict, t_ns, t_nm, t_pm, t_ps):
                    import io
                    import openpyxl
                    from openpyxl.styles import PatternFill, Font, Alignment
                    from openpyxl.utils import get_column_letter

                    _fill_red    = PatternFill("solid", fgColor="FF4444")
                    _fill_orange = PatternFill("solid", fgColor="FF8C00")
                    _fill_green  = PatternFill("solid", fgColor="16A34A")
                    _fill_amber  = PatternFill("solid", fgColor="F59E0B")
                    _font_white  = Font(color="FFFFFF", bold=True)
                    _font_dark   = Font(color="1A1A1A", bold=True)
                    _hdr_fill    = PatternFill("solid", fgColor="0D1B2A")
                    _hdr_font    = Font(color="FFFFFF", bold=True)

                    sheet_defs = [
                        ("City×Day",          views_dict.get("v1")),
                        ("City×Cat×Day",       views_dict.get("v2")),
                        ("Hub×Cat×Day",        views_dict.get("v3")),
                        ("Hub×Day",            views_dict.get("v4")),
                    ]

                    wb  = openpyxl.Workbook()
                    wb.remove(wb.active)

                    for sname, sdf in sheet_defs:
                        if sdf is None or sdf.empty:
                            continue
                        ws   = wb.create_sheet(title=sname)
                        cols = sdf.columns.tolist()
                        delta_idx = (cols.index("Delta %") + 1) if "Delta %" in cols else None

                        # Header row
                        for ci, col in enumerate(cols, 1):
                            cell       = ws.cell(row=1, column=ci, value=col)
                            cell.fill  = _hdr_fill
                            cell.font  = _hdr_font
                            cell.alignment = Alignment(horizontal="center")

                        # Data rows
                        num_start = max(
                            (cols.index(c) + 1)
                            for c in ["Previous Baseline", "Current Baseline"]
                            if c in cols
                        ) - 1 if any(c in cols for c in ["Previous Baseline", "Current Baseline"]) else len(cols)

                        for ri, row_vals in enumerate(sdf.itertuples(index=False), 2):
                            for ci, val in enumerate(row_vals, 1):
                                cell = ws.cell(row=ri, column=ci, value=val)
                                cell.alignment = Alignment(
                                    horizontal="right" if ci > num_start else "left"
                                )
                                if ci == delta_idx and val is not None:
                                    try:
                                        fv = float(val)
                                        if fv < t_ns:
                                            cell.fill = _fill_red;    cell.font = _font_white
                                        elif fv < t_nm:
                                            cell.fill = _fill_orange; cell.font = _font_white
                                        elif fv > t_ps:
                                            cell.fill = _fill_green;  cell.font = _font_white
                                        elif fv > t_pm:
                                            cell.fill = _fill_amber;  cell.font = _font_dark
                                    except (TypeError, ValueError):
                                        pass

                        # Auto-column widths (sample first 200 rows)
                        for ci, col in enumerate(cols, 1):
                            sample_vals = sdf[col].head(200).astype(str).tolist() if col in sdf.columns else []
                            max_w = max([len(col)] + [len(v) for v in sample_vals], default=10)
                            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 40)

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf.getvalue()

                _rv_excel_bytes = _rv_generate_excel(
                    _rv_views, _rv_t_ns, _rv_t_nm, _rv_t_pm, _rv_t_ps
                )
                st.download_button(
                    label="📥 Download Comparison as Excel (4 sheets)",
                    data=_rv_excel_bytes,
                    file_name=f"Baseline_Comparison_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="rv_excel_download",
                    use_container_width=False,
                )
        else:
            st.info("Click 'Load Multi-level Comparison' to view comparison across granularity levels.")

    def approve_baseline(self, user_id, *, user: dict | None = None, read_only: bool = False):
        """Approve baseline — city × category × day pivot view then approve to unlock Final Plan."""
        st.subheader("✅ Approve Baseline")

        _HUB_SHEET_URL = DP_LOGICS_SHEET_URL
        _HUB_SHEET_KEY = DP_LOGICS_SHEET_KEY
        _HUB_WS_NAME   = "Hub level Suggestion"
        _CREDS_PATH = GOOGLE_CREDENTIALS_PATH
        _role = (user or {}).get("role", "")
        _may_approve = can_approve(_role) and not read_only

        # ── Approval banner ────────────────────────────────────────────────────
        if st.session_state.get("baseline_approved"):
            st.success(
                f"✅ Baseline **Approved** by `{st.session_state.get('baseline_approved_by', '—')}` "
                f"at {st.session_state.get('baseline_approved_at', '—')}  |  "
                "You can now proceed to **🎯 Final Plan** from the sidebar."
            )
            if st.button("↺ Revoke Approval", key="revoke_approve_top", type="secondary", disabled=not _may_approve):
                if not _may_approve:
                    st.error("Only administrators can revoke baseline approval.")
                    return
                self._clear_baseline_approval()
                st.rerun()
            st.markdown("---")

        # ── Data load ──────────────────────────────────────────────────────────
        st.markdown("### 📊 City × Category × Day — Base Plan View")
        st.caption(f"Source: **{_HUB_WS_NAME}** · [Open Sheet]({_HUB_SHEET_URL})")

        if st.button("🔄 Load / Refresh", key="approve_load_btn", type="secondary"):
            with st.spinner("Loading Hub level Suggestion from Google Sheets…"):
                try:
                    from planning_suite.services.google_sheets import GoogleSheetsManager
                    sheets_manager = GoogleSheetsManager()
                    _ws    = sheets_manager.gc.open_by_key(_HUB_SHEET_KEY).worksheet(_HUB_WS_NAME)
                    _raw   = _ws.get_all_values()
                    if len(_raw) < 2:
                        st.error("Hub level Suggestion sheet is empty — run the baseline first.")
                    else:
                        _loaded = pd.DataFrame(_raw[1:], columns=_raw[0])
                        _loaded.columns = [c.strip() for c in _loaded.columns]
                        st.session_state["_approve_hub_df"] = _loaded
                        # Persist to disk so it survives refresh
                        try:
                            os.makedirs("outputs", exist_ok=True)
                            _loaded.to_parquet(
                                os.path.join("outputs", "hub_suggestion_latest.parquet"), index=False
                            )
                        except Exception:
                            pass
                        st.success(f"✅ {len(_loaded):,} rows loaded from **{_HUB_WS_NAME}**")
                except Exception as _e:
                    st.error(f"❌ Failed to load sheet: {_e}")

        # ── Pivot table view ───────────────────────────────────────────────────
        # Auto-restore Hub Suggestion data from disk after refresh
        _HS_CACHE = os.path.join("outputs", "hub_suggestion_latest.parquet")
        if "_approve_hub_df" not in st.session_state and os.path.exists(_HS_CACHE):
            try:
                st.session_state["_approve_hub_df"] = pd.read_parquet(_HS_CACHE)
                st.caption("⚡ Base plan data restored from local cache.")
            except Exception:
                pass

        # Auto-restore approval flags from disk after refresh
        _APPROVAL_JSON = str(BASELINE_APPROVAL_JSON)
        if not st.session_state.get("baseline_approved") and os.path.exists(_APPROVAL_JSON):
            try:
                import json as _json
                with open(_APPROVAL_JSON) as _jf:
                    _saved = _json.load(_jf)
                st.session_state.baseline_approved    = _saved.get("approved", False)
                st.session_state.baseline_approved_at = _saved.get("approved_at", "")
                st.session_state.baseline_approved_by = _saved.get("approved_by", "")
            except Exception:
                pass

        if "_approve_hub_df" not in st.session_state:
            st.info("Click **Load / Refresh** to view the current base plan before approving.")
            return

        _df = st.session_state["_approve_hub_df"].copy()

        def _find_col(df, candidates):
            for c in df.columns:
                if c.strip().lower() in [x.lower() for x in candidates]:
                    return c
            return None

        _city_col = _find_col(_df, ["city_name", "city"])
        _hub_col  = _find_col(_df, ["hub_name", "hub"])
        _sku_col  = _find_col(_df, ["sku class prod", "SKU Class Prod", "sku_class_prod", "category"])
        _day_col  = _find_col(_df, ["day"])
        _bp_col   = _find_col(_df, ["Base_plan", "base_plan", "base plan", "BasePlan"])

        if not all([_hub_col, _sku_col, _day_col, _bp_col]):
            st.error(
                f"Sheet is missing required columns. Expected: hub_name, sku class prod, day, Base_plan.  \n"
                f"Found: {_df.columns.tolist()}"
            )
            return

        # Coerce Base_plan to numeric
        _df[_bp_col] = pd.to_numeric(_df[_bp_col], errors="coerce").fillna(0)

        # Group column: city if available, else hub
        _grp_label = "City" if _city_col else "Hub"
        _grp_col   = _city_col if _city_col else _hub_col

        # Summary metrics
        _mc1, _mc2, _mc3, _mc4 = st.columns(4)
        _mc1.metric("Total Base Plan",  f"{int(_df[_bp_col].sum()):,}")
        _mc2.metric(f"Unique {_grp_label}s", _df[_grp_col].nunique())
        _mc3.metric("SKU Classes",      _df[_sku_col].nunique())
        _mc4.metric("Unique Hubs",      _df[_hub_col].nunique())

        # Filters
        _f1, _f2 = st.columns(2)
        with _f1:
            _city_opts = ["All"] + sorted(_df[_grp_col].dropna().astype(str).unique().tolist())
            _sel_city  = st.selectbox(f"Filter by {_grp_label}", _city_opts, key="approve_city_filter")
        with _f2:
            _sku_opts = ["All"] + sorted(_df[_sku_col].dropna().astype(str).unique().tolist())
            _sel_sku  = st.selectbox("Filter by SKU Class Prod", _sku_opts, key="approve_sku_filter")

        _vis = _df.copy()
        if _sel_city != "All":
            _vis = _vis[_vis[_grp_col].astype(str) == _sel_city]
        if _sel_sku != "All":
            _vis = _vis[_vis[_sku_col].astype(str) == _sel_sku]

        st.caption(f"Showing {len(_vis):,} of {len(_df):,} rows")

        # Build pivot: rows = city × sku, columns = day, values = Base_plan
        _pivot_idx = [c for c in [_city_col, _hub_col, _sku_col] if c]
        _agg = (
            _vis.groupby(_pivot_idx + [_day_col], as_index=False)[_bp_col]
            .sum()
        )

        if _agg.empty:
            st.info("No data for the selected filters.")
        else:
            _DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            _pivot = _agg.pivot_table(
                index=_pivot_idx,
                columns=_day_col,
                values=_bp_col,
                aggfunc="sum",
                fill_value=0,
            ).reset_index()
            _pivot.columns.name = None

            # Sort day columns Mon → Sun
            _day_cols   = [d for d in _DAY_ORDER if d in _pivot.columns]
            _extra_days = [d for d in _pivot.columns if d not in _pivot_idx and d not in _day_cols]
            _pivot      = _pivot[_pivot_idx + _day_cols + _extra_days]

            # Row total
            _num_cols       = _day_cols + _extra_days
            _pivot["Total"] = _pivot[_num_cols].sum(axis=1)

            _col_cfg = {
                c: st.column_config.NumberColumn(c, format="%,.0f")
                for c in _num_cols + ["Total"]
            }
            st.dataframe(
                _pivot,
                column_config=_col_cfg,
                use_container_width=True,
                height=min(60 + len(_pivot) * 36, 560),
            )

            # Day-wise totals as metrics
            if _day_cols:
                _day_metric_cols = st.columns(min(len(_day_cols), 7))
                for _i, _d in enumerate(_day_cols[:7]):
                    _day_metric_cols[_i].metric(_d[:3], f"{int(_pivot[_d].sum()):,}")

        # ── Approve button ─────────────────────────────────────────────────────
        st.markdown("---")
        if not st.session_state.get("baseline_approved"):
            st.markdown("### 🔐 Approve & Unlock Final Plan")
            st.info(
                "Once you approve, the baseline is locked and **🎯 Final Plan Generation** "
                "becomes accessible from the sidebar."
            )
            if not _may_approve:
                st.warning("Only administrators can approve baselines.")
            if st.button(
                "✅ Approve Baseline", key="approve_btn_main",
                type="primary", use_container_width=True,
                disabled=not _may_approve,
            ):
                from datetime import datetime as _dt
                _now = _dt.now().strftime("%Y-%m-%d %H:%M")
                run_id = self.db.resolve_baseline_run_id_for_approval(
                    st.session_state.get("baseline_run_id")
                )
                if not run_id:
                    st.error(
                        "No completed baseline run found in history. "
                        "Run baseline generation successfully before approving."
                    )
                else:
                    try:
                        self.db.approve_baseline_run(run_id, user_id)
                        import json as _json
                        os.makedirs("outputs", exist_ok=True)
                        with open(str(BASELINE_APPROVAL_JSON), "w") as _jf:
                            _json.dump(
                                {"approved": True, "approved_at": _now, "approved_by": str(user_id)},
                                _jf,
                            )
                    except Exception as exc:
                        st.error(f"Approval failed — could not persist to database: {exc}")
                    else:
                        _br = self.db.get_baseline_run(run_id)
                        _run_name = run_id
                        if _br is not None and hasattr(_br, "_mapping"):
                            _run_name = _br._mapping.get("run_name") or run_id
                        _approver = (user or {}).get("full_name") or (user or {}).get("username") or ""
                        notify_baseline_approved(
                            run_id=run_id,
                            run_name=_run_name,
                            approver_id=user_id,
                            approver_name=_approver,
                            db=self.db,
                        )
                        st.session_state.baseline_approved = True
                        st.session_state.baseline_approved_at = _now
                        st.session_state.baseline_approved_by = user_id
                        st.session_state.baseline_run_id = run_id
                        clear_baseline_approval_cache()
                        st.success("🎉 Baseline Approved! Navigate to **🎯 Final Plan** from the sidebar.")
                        st.balloons()
                        st.rerun()
        else:
            st.success(
                f"✅ Already Approved — navigate to **🎯 Final Plan** from the sidebar."
            )
            if st.button("↺ Revoke Approval", key="revoke_approve_bottom", type="secondary", disabled=not _may_approve):
                if not _may_approve:
                    st.error("Only administrators can revoke baseline approval.")
                    return
                self._clear_baseline_approval()
                st.rerun()
