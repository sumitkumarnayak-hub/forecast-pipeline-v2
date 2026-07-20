###############################################################
# newlaunchv2_full.py  —  New Product Launch Module
#
# Flow (per launch type):
#   Stage 1 – "input"        : select plan level, pick cities/hubs,
#                              download template, upload filled file
#   Stage 2 – "split_review" : (city-level upload) editable hub split
#                              derived from hub_suggestion_latest.parquet
#   Stage 3 – "set_date"     : Monday ≥ T+4 launch date picker
#   Stage 4 – "confirm"      : review + submit  →  mail placeholder
#
# Duplicate rule:
#   • After city-level upload  → check city + product_id
#   • After hub-level upload   → check city + hub + product_id
#   → Duplicate found: show existing submission (from Submission_Log)
#     in editable mode; user updates and re-submits
###############################################################

import pandas as pd
import numpy as np
import gspread
import io
import os
import logging
from datetime import datetime, date, timedelta
from google.oauth2.service_account import Credentials
import uuid
import pandera as pa

from app.config import (
    GOOGLE_CREDENTIALS_PATH,
    HUB_LEVEL_PLANNING_SHEET_KEY,
    OUTPUT_PATH,
    FF_AUTOMATION_SHEET_KEY,
)

logger = logging.getLogger(__name__)

WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

CITY_UPLOAD_SCHEMA = pa.DataFrameSchema({
    # Support both "City" and internal "city_name"
    "City": pa.Column(str, coerce=True, required=False),
    "city_name": pa.Column(str, coerce=True, required=False),
    # Product ID / PRODUCT_ID / product_id
    "Product ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "PRODUCT_ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "product_id": pa.Column(str, coerce=True, nullable=True, required=False),
    # Product Name / PRODUCT_NAME / product_name
    "Product Name": pa.Column(str, coerce=True, nullable=True, required=False),
    "PRODUCT_NAME": pa.Column(str, coerce=True, nullable=True, required=False),
    "product_name": pa.Column(str, coerce=True, nullable=True, required=False),
    # Sub Category / SUB_CATEGORY / category
    "Sub Category": pa.Column(str, coerce=True, nullable=True, required=False),
    "SUB_CATEGORY": pa.Column(str, coerce=True, nullable=True, required=False),
    "category": pa.Column(str, coerce=True, nullable=True, required=False),
    # Anchor ID
    "Anchor ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "anchor_id": pa.Column(str, coerce=True, nullable=True, required=False),
    # Channel
    "Channel": pa.Column(str, coerce=True, nullable=True, required=False),
    # MRP / MRP\n(Before KVi Discount)
    "MRP": pa.Column(float, coerce=True, nullable=True, required=False),
    "MRP\n(Before KVi Discount)": pa.Column(float, coerce=True, nullable=True, required=False),
    # Optional columns
    "UOM": pa.Column(str, coerce=True, nullable=True, required=False),
    "Yield": pa.Column(str, coerce=True, nullable=True, required=False),
    "RM": pa.Column(str, coerce=True, nullable=True, required=False),
    "Meat Ratio": pa.Column(str, coerce=True, nullable=True, required=False),
    "Meat Ratio (for VA)": pa.Column(str, coerce=True, nullable=True, required=False),
    "Total Shelf Life": pa.Column(str, coerce=True, nullable=True, required=False),
    "Hub Shelf Life": pa.Column(str, coerce=True, nullable=True, required=False),
    "PLU Code": pa.Column(str, coerce=True, nullable=True, required=False),
    "PLU_CODE": pa.Column(str, coerce=True, nullable=True, required=False),
    "Production PC": pa.Column(str, coerce=True, nullable=True, required=False),
    "Old Product ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "Old Product Name": pa.Column(str, coerce=True, nullable=True, required=False),
    "old_product_id": pa.Column(str, coerce=True, nullable=True, required=False),
    "old_product_name": pa.Column(str, coerce=True, nullable=True, required=False),
    "Replacement Percentage": pa.Column(str, coerce=True, nullable=True, required=False),
    "replacement_percentage": pa.Column(str, coerce=True, nullable=True, required=False),
    **{day: pa.Column(int, coerce=True, nullable=True, required=False) for day in WEEKDAYS}
}, strict=False)

HUB_UPLOAD_SCHEMA = pa.DataFrameSchema({
    # Support both "City" and internal "city_name"
    "City": pa.Column(str, coerce=True, required=False),
    "city_name": pa.Column(str, coerce=True, required=False),
    # Support "Hub Name" and internal "hub_name"
    "Hub Name": pa.Column(str, coerce=True, required=False),
    "hub_name": pa.Column(str, coerce=True, required=False),
    # Product ID / PRODUCT_ID / product_id
    "Product ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "PRODUCT_ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "product_id": pa.Column(str, coerce=True, nullable=True, required=False),
    # Product Name / PRODUCT_NAME / product_name
    "Product Name": pa.Column(str, coerce=True, nullable=True, required=False),
    "PRODUCT_NAME": pa.Column(str, coerce=True, nullable=True, required=False),
    "product_name": pa.Column(str, coerce=True, nullable=True, required=False),
    # Sub Category / SUB_CATEGORY / category
    "Sub Category": pa.Column(str, coerce=True, nullable=True, required=False),
    "SUB_CATEGORY": pa.Column(str, coerce=True, nullable=True, required=False),
    "category": pa.Column(str, coerce=True, nullable=True, required=False),
    # Anchor ID
    "Anchor ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "anchor_id": pa.Column(str, coerce=True, nullable=True, required=False),
    # Channel
    "Channel": pa.Column(str, coerce=True, nullable=True, required=False),
    # MRP / MRP\n(Before KVi Discount)
    "MRP": pa.Column(float, coerce=True, nullable=True, required=False),
    "MRP\n(Before KVi Discount)": pa.Column(float, coerce=True, nullable=True, required=False),
    # Optional columns
    "UOM": pa.Column(str, coerce=True, nullable=True, required=False),
    "Yield": pa.Column(str, coerce=True, nullable=True, required=False),
    "RM": pa.Column(str, coerce=True, nullable=True, required=False),
    "Meat Ratio": pa.Column(str, coerce=True, nullable=True, required=False),
    "Meat Ratio (for VA)": pa.Column(str, coerce=True, nullable=True, required=False),
    "Total Shelf Life": pa.Column(str, coerce=True, nullable=True, required=False),
    "Hub Shelf Life": pa.Column(str, coerce=True, nullable=True, required=False),
    "PLU Code": pa.Column(str, coerce=True, nullable=True, required=False),
    "PLU_CODE": pa.Column(str, coerce=True, nullable=True, required=False),
    "Production PC": pa.Column(str, coerce=True, nullable=True, required=False),
    "Old Product ID": pa.Column(str, coerce=True, nullable=True, required=False),
    "Old Product Name": pa.Column(str, coerce=True, nullable=True, required=False),
    "old_product_id": pa.Column(str, coerce=True, nullable=True, required=False),
    "old_product_name": pa.Column(str, coerce=True, nullable=True, required=False),
    "Replacement Percentage": pa.Column(str, coerce=True, nullable=True, required=False),
    "replacement_percentage": pa.Column(str, coerce=True, nullable=True, required=False),
    **{day: pa.Column(int, coerce=True, nullable=True, required=False) for day in WEEKDAYS}
}, strict=False)

SERVICE_ACCOUNT_FILE = GOOGLE_CREDENTIALS_PATH
SPREADSHEET_ID = HUB_LEVEL_PLANNING_SHEET_KEY
MASTER_FILE_ID = FF_AUTOMATION_SHEET_KEY
MASTER_SHEET_NAME    = "P-L Master"
OUTPUT_SHEET_NAME    = "Launch_Output"
LOG_SHEET_NAME       = "Submission_Log"

WEEKDAYS    = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
OUTPUTS_DIR = str(OUTPUT_PATH)
SALIENCE_SHEET_NAME = "Hub level Suggestion"
HUB_SKU_MASTER_SHEET = "Hub Sku Master"

LOG_HEADERS = [
    "Timestamp", "Submission_ID", "Submission_Type",
    "Product ID", "Product Name", "Category",
    "City", "Hub", "MRP", "Start Date",
    "Status", "Rejection_Reason", "Submitted_By",
    "Old Product ID", "Old Product Name", "Replacement Percentage",
] + WEEKDAYS + [
    "PLU Code", "PLU_CODE", "UOM", "Yield", "RM", "Meat Ratio", "Meat Ratio (for VA)", "Total Shelf Life", "Hub Shelf Life"
]

LAUNCH_OUTPUT_HEADERS = [
    "Product ID", "Product Name", "Category", "City", "Hub", "Start Date", "Day", "Plan",
]


# ──────────────────────────────────────────────────────────────────
# COLUMN-NAME NORMALISER  (handles any casing / spacing from sheets)
# ──────────────────────────────────────────────────────────────────
def _canon(s: str) -> str:
    return s.strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Drop duplicate columns first (e.g. from Google Sheets blank columns)
    df = df.loc[:, ~df.columns.duplicated(keep="first")].copy()
    
    rename = {}
    for col in df.columns:
        k = _canon(str(col))
        if k == "subcategory":   rename[col] = "sub_category"
        elif k == "baseplan":    rename[col] = "Base_plan"
        elif k == "planflag":    rename[col] = "Plan Flag"
        elif k == "productid":   rename[col] = "Product id"
        elif k == "cityname":    rename[col] = "city_name"
        elif k == "hubname":     rename[col] = "hub_name"
    if rename:
        df = df.rename(columns=rename)
    for c in df.select_dtypes(include="object").columns:
        df[c] = df[c].astype(str).str.strip()
    return df


def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep first occurrence when headers are duplicated (common in sheet uploads)."""
    if df is None or df.empty:
        return df
    if not df.columns.duplicated().any():
        return df
    return df.loc[:, ~df.columns.duplicated(keep="first")].copy()


def _resolve_col(df: pd.DataFrame, *candidates: str) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    lookup = {_canon(c): c for c in df.columns}
    for cand in candidates:
        hit = lookup.get(_canon(cand))
        if hit:
            return hit
    return None


def _subcat_col(df: pd.DataFrame) -> str:
    for c in ["sub_category", "sub category", "Sub-category", "Sub Category"]:
        if c in df.columns:
            return c
    for c in df.columns:
        if _canon(c) == "subcategory":
            return c
    return "sub_category"


# ──────────────────────────────────────────────────────────────────
# GOOGLE SHEETS  CONNECTORS
# ──────────────────────────────────────────────────────────────────
def _get_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    return gspread.authorize(creds)


def _open_sheet(spreadsheet_id: str, sheet_name: str):
    client = _get_client()
    sh = client.open_by_key(spreadsheet_id)
    try:
        return sh.worksheet(sheet_name)
    except Exception:
        ws = sh.add_worksheet(sheet_name, rows="5000", cols="50")
        return ws


# ──────────────────────────────────────────────────────────────────
# PRODUCT MASTER  (Google Sheet)
# ──────────────────────────────────────────────────────────────────

def load_product_master() -> pd.DataFrame:
    from features.product_launch.ff_masters import load_product_master_df

    return load_product_master_df()


def _subcat_col_master(df: pd.DataFrame) -> str:
    return _subcat_col(df)


def get_categories(df_master: pd.DataFrame) -> list:
    c = _subcat_col_master(df_master)
    return sorted(df_master[c].dropna().unique().tolist()) if c in df_master.columns else []


def get_products_by_category(df_master: pd.DataFrame, category: str) -> list:
    c = _subcat_col_master(df_master)
    sub = df_master[df_master[c] == category] if c in df_master.columns else df_master
    return sorted(sub["Product Name"].dropna().unique().tolist())


def get_product_id(df_master: pd.DataFrame, product_name: str) -> str:
    row = df_master[df_master["Product Name"] == product_name]
    if row.empty:
        return ""
    for c in ["Product id", "Product ID"]:
        if c in row.columns:
            return str(row[c].iloc[0])
    return ""


def get_product_info(df_master: pd.DataFrame, product_id: str) -> dict:
    for pid_col in ["Product id", "Product ID"]:
        if pid_col not in df_master.columns:
            continue
        row = df_master[df_master[pid_col].astype(str) == str(product_id)]
        if not row.empty:
            sc = _subcat_col_master(df_master)
            return {
                "name":     row["Product Name"].iloc[0] if "Product Name" in row.columns else "",
                "category": row[sc].iloc[0] if sc in row.columns else "",
            }
    return {"name": "", "category": ""}


# ──────────────────────────────────────────────────────────────────
# HUB SALIENCE  (Hub level Suggestion sheet — same as Product)
# ──────────────────────────────────────────────────────────────────

def _salience_spreadsheet_key() -> str | None:
    """Spreadsheet ID for Hub level Suggestion (DP Logics preferred, Hub Level Planning fallback)."""
    from app import config as cfg

    for key in (getattr(cfg, "DP_LOGICS_SHEET_KEY", None), SPREADSHEET_ID):
        if key:
            return key
    return None


def _normalize_weekday(val) -> str:
    """Map day labels to Mon–Sun."""
    text = str(val).strip()
    if text in WEEKDAYS:
        return text
    aliases = {
        "monday": "Mon", "mon": "Mon",
        "tuesday": "Tue", "tues": "Tue", "tue": "Tue",
        "wednesday": "Wed", "wed": "Wed",
        "thursday": "Thu", "thur": "Thu", "thu": "Thu",
        "friday": "Fri", "fri": "Fri",
        "saturday": "Sat", "sat": "Sat",
        "sunday": "Sun", "sun": "Sun",
    }
    return aliases.get(text.lower(), text)


def load_salience_source() -> pd.DataFrame:
    """Load hub-level Base_plan rows from Hub level Suggestion (Product parity)."""
    from features.product_launch.sheet_reads import read_sheet_values_cached

    sheet_key = _salience_spreadsheet_key()
    if not sheet_key:
        return pd.DataFrame()

    def _fetch():
        from core.shared.sheets_throttle import sheets_slot

        with sheets_slot():
            sheet = _open_sheet(sheet_key, SALIENCE_SHEET_NAME)
            return sheet.get_all_values()

    data = read_sheet_values_cached(
        sheet_key,
        SALIENCE_SHEET_NAME,
        "all",
        sheet_category="dp_logics",
        fetcher=_fetch,
    )
    if not data or len(data) < 2:
        return pd.DataFrame()

    df = pd.DataFrame(data[1:], columns=data[0])
    df = _normalize_columns(df)

    day_col = _resolve_col(df, "day", "Day")
    if day_col and day_col != "day":
        df = df.rename(columns={day_col: "day"})
    if "day" in df.columns:
        df["day"] = df["day"].map(_normalize_weekday)

    if "Base_plan" not in df.columns:
        bp = _resolve_col(df, "Base_plan", "Base plan", "base_plan", "BasePlan")
        if bp:
            df = df.rename(columns={bp: "Base_plan"})
    if "Base_plan" not in df.columns:
        return pd.DataFrame()

    df["Base_plan"] = pd.to_numeric(df["Base_plan"], errors="coerce").fillna(0).round()
    df = df[df["Base_plan"] > 0]
    return df.reset_index(drop=True)


def _truthy_flag(val) -> bool:
    """1 / TRUE / Y / A / Active all count as an "on" flag; blank / 0 / N / I do not."""
    s = str(val).strip().lower()
    return s in ("1", "1.0", "true", "yes", "y", "a", "active")


def load_hub_sku_master_flags() -> pd.DataFrame:
    """
    Raw "Hub Sku Master" tab (hub_level_planning sheet) normalized to
    city_name / hub_name / hub_active / active_plan (booleans).

    Returns an empty DataFrame if the tab is missing/unreadable so callers can
    skip the active-hub filter rather than break the split.
    """
    from features.product_launch.sheet_reads import read_sheet_values_cached

    if not SPREADSHEET_ID:
        return pd.DataFrame()

    def _fetch():
        from core.shared.sheets_throttle import sheets_slot

        with sheets_slot():
            sheet = _open_sheet(SPREADSHEET_ID, HUB_SKU_MASTER_SHEET)
            return sheet.get_all_values()

    try:
        data = read_sheet_values_cached(
            SPREADSHEET_ID,
            HUB_SKU_MASTER_SHEET,
            "all",
            sheet_category="hub_level_planning",
            fetcher=_fetch,
        )
    except Exception as exc:
        logger.warning("Could not load Hub Sku Master tab: %s", exc)
        return pd.DataFrame()

    if not data or len(data) < 2:
        return pd.DataFrame()

    df = pd.DataFrame(data[1:], columns=data[0])
    df = _normalize_columns(df)

    city_col = _resolve_col(df, "city_name", "City Name", "city")
    hub_col = _resolve_col(df, "hub_name", "Hub Name", "hub")
    if not city_col or not hub_col:
        return pd.DataFrame()

    active_col = _resolve_col(df, "hub active", "hub_active", "Hub Active", "hubactive")
    plan_col = _resolve_col(
        df, "active plan", "active_plan", "Active Plan", "plan active", "planactive", "activeplan",
    )

    out = pd.DataFrame({
        "city_name": df[city_col].astype(str).str.strip(),
        "hub_name": df[hub_col].astype(str).str.strip(),
    })
    out["hub_active"] = df[active_col].apply(_truthy_flag) if active_col else True
    out["active_plan"] = df[plan_col].apply(_truthy_flag) if plan_col else True
    return out[(out["city_name"] != "") & (out["hub_name"] != "")].reset_index(drop=True)


def get_active_hub_sku_master_pairs() -> set[tuple[str, str]]:
    """
    (city_name, hub_name) pairs (lowercased) where Hub Sku Master marks
    Hub Active = 1 AND Active Plan = 1.

    Empty set means the filter is unavailable — callers should skip it
    rather than treat "no active hubs" as "no hubs at all".
    """
    df = load_hub_sku_master_flags()
    if df.empty:
        return set()
    active = df[df["active_plan"]]
    return {
        (str(c).strip().lower(), str(h).strip().lower())
        for c, h in zip(active["city_name"], active["hub_name"])
        if str(c).strip() and str(h).strip()
    }


def _filter_rows_by_active_hub_pairs(
    df: pd.DataFrame,
    pairs: set[tuple[str, str]],
    *,
    city_col: str | None = None,
    hub_col: str | None = None,
) -> pd.DataFrame:
    """Keep only rows whose (city, hub) is in the active-hub whitelist.

    If nothing matches (e.g. naming drift between sheets) the original
    frame is returned unchanged so the split never silently drops to zero.
    """
    if df is None or df.empty or not pairs:
        return df
    city_col = city_col or _resolve_col(df, "city_name", "City Name", "city")
    hub_col = hub_col or _resolve_col(df, "hub_name", "Hub Name", "hub")
    if not city_col or not hub_col:
        return df
    mask = [
        (str(c).strip().lower(), str(h).strip().lower()) in pairs
        for c, h in zip(df[city_col], df[hub_col])
    ]
    filtered = df[pd.Series(mask, index=df.index)]
    if filtered.empty:
        logger.warning(
            "[NPL] Hub Sku Master active-hub filter matched 0 rows — "
            "keeping unfiltered salience source (check city/hub name alignment)."
        )
        return df
    return filtered


def load_hub_sku_master() -> pd.DataFrame:
    """City/hub catalog from FF Automation Hub Mapping (cached parquet)."""
    from features.product_launch.ff_masters import hub_mapping_as_catalog

    return hub_mapping_as_catalog()


def get_active_hubs_for_city(hub_sku_df: pd.DataFrame, city: str, category: str | None = None) -> list[str]:
    """Return sorted active hub names for a city and optional category."""
    if hub_sku_df is None or hub_sku_df.empty:
        return []
    mask = (
        hub_sku_df["city_name"].astype(str).str.strip().str.lower() == str(city).strip().lower()
        ) & (
        hub_sku_df["Plan Flag"].astype(str).str.strip().str.upper() != "I"
        )
    # Ignore category since Hub_Mapping is category-agnostic
    return sorted(hub_sku_df.loc[mask, "hub_name"].dropna().astype(str).str.strip().unique().tolist())


def get_expansion_hubs_for_city(
    hub_sku_df: pd.DataFrame,
    city: str,
    category: str,
    all_hubs_in_city: list[str],
) -> list[str]:
    """Return hubs eligible for expansion in a city+category."""
    if hub_sku_df is None or hub_sku_df.empty:
        return sorted(set(all_hubs_in_city))

    city_lower = str(city).strip().lower()
    
    city_mask = (
        hub_sku_df["city_name"].astype(str).str.strip().str.lower() == city_lower
    )
    city_df = hub_sku_df.loc[city_mask]

    inactive_hubs = city_df[
        city_df["Plan Flag"].astype(str).str.strip().str.upper() == "I"
    ]["hub_name"].dropna().astype(str).str.strip().unique().tolist()

    present_hubs = {
        str(h).strip().lower()
        for h in city_df["hub_name"].dropna().astype(str).tolist()
    }
    not_present_hubs = [
        hub for hub in all_hubs_in_city
        if str(hub).strip().lower() not in present_hubs
    ]

    return sorted(set(inactive_hubs + not_present_hubs))


def get_expansion_cities(hub_sku_df: pd.DataFrame, category: str, all_cities: list[str]) -> list[str]:
    """Return cities that have at least one expansion-eligible hub for the category."""
    if hub_sku_df is None or hub_sku_df.empty:
        return []
    result = []

    for city in all_cities:
        city_lower = str(city).strip().lower()
        city_mask = (
            hub_sku_df["city_name"].astype(str).str.strip().str.lower() == city_lower
        )
        city_df = hub_sku_df.loc[city_mask]

        has_inactive = (
            city_df["Plan Flag"].astype(str).str.strip().str.upper() == "I"
        ).any()
        if has_inactive:
            result.append(city)
            continue

        present_hubs = {
            str(h).strip().lower()
            for h in city_df["hub_name"].dropna().astype(str).tolist()
        }
        all_hubs_city = {
            str(h).strip().lower()
            for h in hub_sku_df.loc[city_mask, "hub_name"].dropna().astype(str).tolist()
        }
        if all_hubs_city - present_hubs:
            result.append(city)

    return sorted(result)


def compute_salience_category(df: pd.DataFrame) -> pd.DataFrame:
    """
    Salience at city × hub × category (× day when day column is present).
    salience = hub Base_plan / city Base_plan for the same category (and day).
    """
    empty = pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])
    if df is None or df.empty:
        return empty

    work = _normalize_columns(df.copy())
    city_col = _resolve_col(work, "city_name", "City Name", "city")
    hub_col = _resolve_col(work, "hub_name", "Hub Name", "hub")
    sc_col = _subcat_col(work)
    if not city_col or not hub_col or sc_col not in work.columns or "Base_plan" not in work.columns:
        return empty

    if city_col != "city_name":
        work = work.rename(columns={city_col: "city_name"})
    if hub_col != "hub_name":
        work = work.rename(columns={hub_col: "hub_name"})
    if sc_col != "sub_category":
        work = work.rename(columns={sc_col: "sub_category"})

    day_col = _resolve_col(work, "day", "Day")
    if day_col and day_col != "day":
        work = work.rename(columns={day_col: "day"})
    if "day" in work.columns:
        work["day"] = work["day"].map(_normalize_weekday)

    hub_group = ["city_name", "hub_name", "sub_category"]
    city_group = ["city_name", "sub_category"]
    if "day" in work.columns:
        hub_group.append("day")
        city_group.append("day")

    hub_tot = (
        work.groupby(hub_group, as_index=False)["Base_plan"]
        .sum()
        .rename(columns={"Base_plan": "hub_total"})
    )
    city_tot = (
        work.groupby(city_group, as_index=False)["Base_plan"]
        .sum()
        .rename(columns={"Base_plan": "city_total"})
    )
    merged = pd.merge(hub_tot, city_tot, on=city_group, how="left")
    merged["salience"] = np.where(
        merged["city_total"] > 0,
        merged["hub_total"] / merged["city_total"],
        0.0,
    )

    out_cols = ["city_name", "hub_name", "sub_category", "salience"]
    if "day" in merged.columns:
        out_cols.append("day")
    return merged[out_cols]


def _load_equal_weight_hub_salience() -> pd.DataFrame:
    """Fallback salience: equal weight across active hubs per city from Hub Mapping."""
    hub_df = load_hub_sku_master()
    if hub_df.empty:
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    active_df = hub_df[hub_df["Plan Flag"].astype(str).str.strip().str.upper() == "A"].copy()
    if active_df.empty:
        return pd.DataFrame(columns=["city_name", "hub_name", "sub_category", "salience"])

    active_pairs = get_active_hub_sku_master_pairs()
    active_df = _filter_rows_by_active_hub_pairs(active_df, active_pairs)

    city_counts = active_df.groupby("city_name")["hub_name"].count().to_dict()
    master_df = load_product_master()
    categories = get_categories(master_df)
    if not categories:
        categories = ["default"]

    rows = []
    for _, row in active_df.iterrows():
        city = str(row["city_name"]).strip()
        hub = str(row["hub_name"]).strip()
        count = city_counts.get(city, 1)
        salience = 1.0 / count if count > 0 else 1.0
        for cat in categories:
            rows.append({
                "city_name": city,
                "hub_name": hub,
                "sub_category": cat,
                "salience": salience,
            })

    return pd.DataFrame(rows)


def load_hub_salience() -> pd.DataFrame:
    """
    Returns city_name, hub_name, sub_category, salience (+ day when available).

    Hub universe: only (city, hub) pairs where Hub Sku Master marks
    Hub Active = 1 AND Active Plan = 1 (see get_active_hub_sku_master_pairs).
    That whitelist is applied to whichever Base_plan source is used below,
    so salience is only ever split across currently-active, actively-planned hubs.

    Primary: Hub level Suggestion Base_plan (hub × category × day).
    Fallback: outputs/hub_suggestion_latest.parquet, then equal-weight Hub Mapping.
    """
    active_pairs = get_active_hub_sku_master_pairs()

    raw = load_salience_source()
    if not raw.empty:
        raw = _filter_rows_by_active_hub_pairs(raw, active_pairs)
        computed = compute_salience_category(raw)
        if not computed.empty:
            return computed

    path = os.path.join(OUTPUTS_DIR, "hub_suggestion_latest.parquet")
    if os.path.exists(path):
        try:
            df = pd.read_parquet(path)
            df = _normalize_columns(df)
            sc = _subcat_col(df)
            if sc != "sub_category":
                df = df.rename(columns={sc: "sub_category"})
            city_col = _resolve_col(df, "city_name", "City Name", "city")
            hub_col = _resolve_col(df, "hub_name", "Hub Name", "hub")
            if city_col and hub_col and "sub_category" in df.columns:
                if city_col != "city_name":
                    df = df.rename(columns={city_col: "city_name"})
                if hub_col != "hub_name":
                    df = df.rename(columns={hub_col: "hub_name"})
                day_col = _resolve_col(df, "day", "Day")
                if day_col and day_col != "day":
                    df = df.rename(columns={day_col: "day"})
                if "day" in df.columns:
                    df["day"] = df["day"].map(_normalize_weekday)
                df = _filter_rows_by_active_hub_pairs(df, active_pairs)
                if "Base_plan" in df.columns and "salience" not in df.columns:
                    return compute_salience_category(df)
                if "salience" in df.columns:
                    cols = ["city_name", "hub_name", "sub_category", "salience"]
                    if "day" in df.columns:
                        cols.append("day")
                    return df[[c for c in cols if c in df.columns]]
        except Exception as exc:
            logger.warning("Could not load hub salience parquet fallback: %s", exc)

    logger.warning(
        "Hub level Suggestion unavailable — using equal-weight salience from Hub Mapping"
    )
    return _load_equal_weight_hub_salience()


def get_cities_from_salience(sal_df: pd.DataFrame) -> list:
    if sal_df is None or sal_df.empty:
        return []
    city_col = _resolve_col(sal_df, "city_name", "City Name", "city")
    if not city_col:
        return []
    return sorted(sal_df[city_col].dropna().astype(str).str.strip().unique().tolist())


def _require_salience(sal_df: pd.DataFrame) -> bool:
    return sal_df is not None and not sal_df.empty


def _get_hubs_for_city_raw(sal_df: pd.DataFrame, city: str, category: str = None) -> list:
    if sal_df is None:
        sal_df = pd.DataFrame()

    city_col = _resolve_col(sal_df, "city_name", "City Name", "city")
    hub_col = _resolve_col(sal_df, "hub_name", "Hub Name", "hub")

    # 1. Prefer hubs from Hub level Suggestion salience (Product city-plan parity)
    if city_col and hub_col and not sal_df.empty:
        mask = sal_df[city_col].astype(str).str.strip().str.lower() == str(city).strip().lower()
        if category:
            sc = _subcat_col(sal_df)
            if sc in sal_df.columns:
                cat_mask = mask & (sal_df[sc].astype(str).str.strip().str.lower() == category.strip().lower())
                hubs = sorted(sal_df.loc[cat_mask, hub_col].dropna().astype(str).str.strip().unique().tolist())
                if hubs:
                    return hubs
        hubs = sorted(sal_df.loc[mask, hub_col].dropna().astype(str).str.strip().unique().tolist())
        if hubs:
            return hubs

    # 2. Active hubs from Hub Mapping / Sku Master
    hub_sku_df = load_hub_sku_master()
    if category and not hub_sku_df.empty:
        active_hubs = get_active_hubs_for_city(hub_sku_df, city, category)
        if active_hubs:
            return active_hubs

    if not hub_sku_df.empty:
        active_hubs = get_active_hubs_for_city(hub_sku_df, city, category=None)
        if active_hubs:
            return active_hubs

    return []


def get_hubs_for_city(sal_df: pd.DataFrame, city: str, category: str = None) -> list:
    """Hubs eligible for a city, restricted to Hub Sku Master Hub Active=1 & Active Plan=1.

    Falls back to the unfiltered list when the active-hub whitelist is
    unavailable or matches nothing (naming drift) so the split never breaks.
    """
    hubs = _get_hubs_for_city_raw(sal_df, city, category)
    if not hubs:
        return hubs

    active_pairs = get_active_hub_sku_master_pairs()
    if not active_pairs:
        return hubs

    city_l = str(city).strip().lower()
    filtered = [h for h in hubs if (city_l, str(h).strip().lower()) in active_pairs]
    return filtered if filtered else hubs


# ──────────────────────────────────────────────────────────────────
# CITY → HUB  SPLIT  (per-day salience allocation)
# ──────────────────────────────────────────────────────────────────
def _hub_day_salience(
    sal_df: pd.DataFrame,
    city: str,
    hub: str,
    category: str,
    day: str,
    *,
    per_day_salience: bool,
) -> float:
    """Salience for one hub on one day; falls back to city-wide (all categories) when category match is zero."""
    sc_col = _subcat_col(sal_df)

    def _lookup(use_category: bool) -> float:
        mask = (
            (sal_df["city_name"].astype(str).str.strip() == city)
            & (sal_df["hub_name"].astype(str).str.strip() == hub)
        )
        if use_category and sc_col in sal_df.columns and category:
            mask &= sal_df[sc_col].astype(str).str.strip().str.lower() == category.strip().lower()
        if per_day_salience:
            mask &= sal_df["day"].astype(str).str.strip() == day
        rows_match = sal_df[mask]
        if rows_match.empty:
            return 0.0
        return float(rows_match["salience"].sum())

    val = _lookup(use_category=True)
    if val == 0.0 and category:
        val = _lookup(use_category=False)
    return val


def _weighted_alloc(total: int, hub_sal: dict) -> dict:
    """Allocate 'total' units across hubs by salience; every hub gets ≥ 1."""
    hubs = list(hub_sal.keys())
    n    = len(hubs)
    if n == 0:
        return {}
    if total <= 0:
        return {h: 0 for h in hubs}
    if total < n:
        alloc = {h: 0 for h in hubs}
        for h in hubs[:total]:
            alloc[h] = 1
        return alloc
    alloc     = {h: 1 for h in hubs}
    remaining = total - n
    total_sal = sum(hub_sal.values())
    if total_sal == 0:
        base, rem = divmod(remaining, n)
        for h in hubs:
            alloc[h] += base
        for h in hubs[:rem]:
            alloc[h] += 1
        return alloc
    raw = {h: remaining * hub_sal[h] / total_sal for h in hubs}
    for h in hubs:
        alloc[h] += int(raw[h])
    leftover = remaining - (sum(alloc.values()) - n)
    for h in sorted(hubs, key=lambda x: raw[x] - int(raw[x]), reverse=True)[:leftover]:
        alloc[h] += 1
    return alloc


def split_city_to_hubs(
    city_df: pd.DataFrame,
    sal_df: pd.DataFrame,
    forced_hubs: dict = None,   # {city: [hub, ...]} override from user selection
) -> tuple:
    """
    city_df : rows with city_name, product_id, product_name, category, MRP, Mon-Sun
    sal_df  : hub salience DataFrame from load_hub_salience()
    forced_hubs : if provided, only split to these hubs per city

    Returns (hub_split_df, zero_sal_info)
      hub_split_df : city_name, hub_name, product_id, product_name, category, MRP, Mon-Sun
      zero_sal_info: {city: [hubs with zero salience]}
    """
    rows          = []
    zero_sal_info = {}

    for _, r in city_df.iterrows():
        city     = str(r.get("city_name", "")).strip()
        category = str(r.get("category", "")).strip()
        if not city:
            continue

        # Hubs to split into
        if forced_hubs and city in forced_hubs:
            hubs = forced_hubs[city]
        else:
            hubs = get_hubs_for_city(sal_df, city, category)

        if not hubs:
            continue

        zero_hubs: list[str] = []
        per_day_salience = "day" in sal_df.columns
        # Pre-compute per-day allocations
        day_alloc = {}
        for day in WEEKDAYS:
            day_sal = {
                hub: _hub_day_salience(
                    sal_df, city, hub, category, day, per_day_salience=per_day_salience,
                )
                for hub in hubs
            }

            # Track hubs still at zero after category + city-wide fallback (Mon only)
            if day == WEEKDAYS[0]:
                zero_hubs = [h for h, s in day_sal.items() if s == 0.0]

            total = int(float(r.get(day, 0) or 0))
            day_alloc[day] = _weighted_alloc(total, day_sal)

        if zero_hubs:
            existing = zero_sal_info.setdefault(city, [])
            for hub in zero_hubs:
                if hub not in existing:
                    existing.append(hub)

        for hub in hubs:
            hub_row = {
                "city_name":    city,
                "hub_name":     hub,
                "product_id":   str(r.get("product_id", "")).strip(),
                "product_name": str(r.get("product_name", "")).strip(),
                "anchor_id":    str(r.get("anchor_id", r.get("product_id", ""))).strip() or str(r.get("product_id", "")).strip(),
                "category":     category,
                "Channel":      str(r.get("Channel", DEFAULT_NPL_CHANNEL)).strip() or DEFAULT_NPL_CHANNEL,
                "MRP":          r.get("MRP", ""),
            }
            for opt in ALL_POSSIBLE_OPTIONAL_COLS:
                if opt in r:
                    hub_row[opt] = r[opt]
            meat_ratio = _meat_ratio_from_source(r.to_dict() if hasattr(r, "to_dict") else dict(r))
            if meat_ratio:
                hub_row["Meat Ratio (for VA)"] = meat_ratio
                hub_row["Meat Ratio"] = meat_ratio
            for day in WEEKDAYS:
                hub_row[day] = day_alloc[day].get(hub, 0)
            rows.append(hub_row)

    cols = ["city_name", "hub_name", "product_id", "product_name", "anchor_id", "category", "Channel", "MRP"] + WEEKDAYS + ALL_POSSIBLE_OPTIONAL_COLS
    hub_df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)

    if not hub_df.empty and not city_df.empty:
        for day in WEEKDAYS:
            if day not in city_df.columns or day not in hub_df.columns:
                continue
            for _, city_row in city_df.iterrows():
                city = str(city_row.get("city_name", "")).strip()
                pid = str(city_row.get("product_id", "")).strip()
                expected = int(float(city_row.get(day, 0) or 0))
                if not city:
                    continue
                mask = (
                    (hub_df["city_name"].astype(str).str.strip() == city)
                    & (hub_df["product_id"].astype(str).str.strip() == pid)
                )
                actual = int(hub_df.loc[mask, day].sum()) if mask.any() else 0
                if actual != expected:
                    logger.warning(
                        "[NPL] split total mismatch %s %s %s: expected=%s actual=%s",
                        city, pid, day, expected, actual,
                    )

    return (hub_df, zero_sal_info)


# ──────────────────────────────────────────────────────────────────
# EXCEL  TEMPLATE  BUILDERS
# ──────────────────────────────────────────────────────────────────
DEFAULT_NPL_CHANNEL = "Online"

CITY_COLS = ["city_name", "product_id", "product_name", "anchor_id", "category", "Channel", "MRP"] + WEEKDAYS
HUB_COLS  = ["city_name", "hub_name", "product_id", "product_name", "anchor_id", "category", "Channel", "MRP"] + WEEKDAYS

ALL_POSSIBLE_OPTIONAL_COLS = [
    "UOM", "Yield", "RM", "Meat Ratio (for VA)", "Meat Ratio",
    "Total Shelf Life", "Hub Shelf Life", "PLU Code", "Start Date",
    "Production PC",
    "old_product_id", "old_product_name", "replacement_percentage"
]

OPTIONAL_PLAN_PROP_COLS = [
    "UOM", "Yield", "RM", "Meat Ratio", "Meat Ratio (for VA)",
    "Total Shelf Life", "Hub Shelf Life", "PLU Code", "Start Date",
    "Production PC",
]


def _optional_str(val) -> str:
    """Preserve optional template fields as plain strings (any user-entered value)."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "nat", "null") else s


def _parse_meat_ratio_percent(val) -> float | None:
    """Parse meat ratio as a percentage in 0–100. Returns None when blank."""
    s = _optional_str(val)
    if not s:
        return None
    raw = str(val).strip()
    has_pct = "%" in raw
    num_str = s.replace("%", "").strip()
    try:
        num = float(num_str)
    except ValueError as exc:
        raise ValueError(f"Meat Ratio must be numeric up to 100% (got {raw!r})") from exc
    if not has_pct and 0 < num <= 1:
        num *= 100
    if num < 0 or num > 100:
        raise ValueError(f"Meat Ratio must be between 0 and 100% (got {raw!r})")
    return num


def _validate_meat_ratio_column(df: pd.DataFrame) -> list[str]:
    """Validate Meat Ratio (for VA) / Meat Ratio values are <= 100%."""
    errors: list[str] = []
    candidates = [
        c for c in df.columns
        if "meat ratio" in str(c).strip().lower()
    ]
    for col in candidates:
        for idx, val in df[col].items():
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            if str(val).strip() == "":
                continue
            try:
                _parse_meat_ratio_percent(val)
            except ValueError as exc:
                errors.append(f"Row {int(idx) + 2}, column '{col}': {exc}")
    return errors


def _normalize_upload_mrp_column(df: pd.DataFrame) -> pd.DataFrame:
    """Map template MRP header to internal MRP and drop duplicate column."""
    mrp_header = "MRP\n(Before KVi Discount)"
    if mrp_header in df.columns:
        if "MRP" not in df.columns:
            df["MRP"] = df[mrp_header]
        df = df.drop(columns=[mrp_header])
    return df


def _validate_required_text_columns(
    df: pd.DataFrame,
    columns: list[tuple[str, str]],
) -> list[str]:
    """Row-level validation for required text columns. columns = [(field, label), ...]."""
    errors: list[str] = []
    for col, label in columns:
        if col not in df.columns:
            errors.append(f"Column '{label}' is missing from the uploaded sheet.")
            continue
        for idx, val in df[col].items():
            if val is None or (isinstance(val, float) and pd.isna(val)):
                errors.append(f"Row {int(idx) + 2}: {label} is required.")
                continue
            text = str(val).strip()
            if not text or text.lower() in ("nan", "none", "nat"):
                errors.append(f"Row {int(idx) + 2}: {label} is required.")
    return errors


def _validate_production_pc_column(df: pd.DataFrame) -> list[str]:
    """Production PC is mandatory on every data row."""
    errors: list[str] = []
    pc_col = next((c for c in df.columns if str(c).strip().lower() == "production pc"), None)
    if not pc_col:
        return ["Column 'Production PC' is missing from the uploaded sheet."]
    for idx, val in df[pc_col].items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            errors.append(f"Row {int(idx) + 2}: Production PC is required.")
            continue
        if not str(val).strip() or str(val).strip().lower() in ("nan", "none", "nat"):
            errors.append(f"Row {int(idx) + 2}: Production PC is required.")
    return errors


def _meat_ratio_from_source(source: dict) -> str:
    for key in ("Meat Ratio (for VA)", "Meat Ratio", "meat_ratio_for_va", "meat_ratio"):
        if key in source and source[key] is not None:
            return _optional_str(source[key])
    for key, val in source.items():
        if val is None:
            continue
        kn = str(key).strip().lower().replace("_", " ")
        if "meat ratio" in kn:
            return _optional_str(val)
    return ""


def _normalize_meat_ratio_column(df: pd.DataFrame) -> pd.DataFrame:
    """Map any meat-ratio header variant to the canonical Meat Ratio (for VA) column."""
    if "Meat Ratio" in df.columns and "Meat Ratio (for VA)" not in df.columns:
        df["Meat Ratio (for VA)"] = df["Meat Ratio"]
    if "Meat Ratio (for VA)" not in df.columns:
        for col in list(df.columns):
            cn = str(col).strip().lower().replace("_", " ")
            if "meat ratio" in cn:
                df["Meat Ratio (for VA)"] = df[col]
                break
    return df


def get_template_columns(plan_level: str, sub_type: str) -> tuple[list[str], list[str]]:
    """Returns (mandatory_cols, optional_cols) dynamically based on plan level and sub_type."""
    if sub_type == "Replacement":
        # Replacement headers
        if plan_level == "city":
            mandatory = ["City", "Product ID", "Product Name", "Anchor ID", "Sub Category", "Channel", "MRP", "Old Product ID", "Old Product Name", "Replacement Percentage", "Production PC"] + WEEKDAYS
        else:
            mandatory = ["City", "Hub Name", "Product ID", "Product Name", "Anchor ID", "Sub Category", "Channel", "MRP", "Old Product ID", "Old Product Name", "Replacement Percentage", "Production PC"] + WEEKDAYS
        optional = ["PLU Code", "UOM", "Yield", "RM", "Meat Ratio", "Total Shelf Life", "Hub Shelf Life"]
    else:
        # New Launch & Expansion headers
        if plan_level == "city":
            mandatory = ["City", "PRODUCT_ID", "PRODUCT_NAME", "Anchor ID", "SUB_CATEGORY", "Channel", "MRP\n(Before KVi Discount)", "Production PC"] + WEEKDAYS
        else:
            mandatory = ["City", "hub_name", "PRODUCT_ID", "PRODUCT_NAME", "Anchor ID", "SUB_CATEGORY", "Channel", "MRP\n(Before KVi Discount)", "Production PC"] + WEEKDAYS
        optional = ["PLU_CODE", "UOM", "Yield", "RM", "Meat Ratio (for VA)", "Total Shelf Life", "Hub Shelf Life"]

    return mandatory + optional, []


def _style_ws(ws, mandatory_cols):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        blue = PatternFill("solid", fgColor="1A73E8")  # Mandatory header
        slate = PatternFill("solid", fgColor="475569")  # Optional header
        grey = PatternFill("solid", fgColor="F8FAFC")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            if cell.value in mandatory_cols:
                cell.fill = blue
            else:
                cell.fill = slate
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                if i % 2 == 0:
                    cell.fill = grey
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max(13, min(mx + 3, 40))
        ws.row_dimensions[1].height = 20
    except ImportError:
        pass


def build_city_template(cities: list, category: str,
                         product_id: str = "", product_name: str = "", mrp: str = "",
                         sub_type: str = "New Launch",
                         old_product_id: str = "", old_product_name: str = "",
                         replacement_percentage: str = "") -> bytes:
    """One blank row per city; prefilled with category, product_id, product_name, and mrp if provided."""
    mandatory_cols, optional_cols = get_template_columns("city", sub_type)
    rows = []
    for c in cities:
        row = {}
        for col in mandatory_cols + optional_cols:
            if col == "City":
                row[col] = c
            elif col in ["PRODUCT_ID", "Product ID"]:
                row[col] = product_id
            elif col in ["PRODUCT_NAME", "Product Name"]:
                row[col] = product_name
            elif col == "Anchor ID":
                row[col] = product_id
            elif col in ["SUB_CATEGORY", "Sub Category"]:
                row[col] = category
            elif col == "Channel":
                row[col] = DEFAULT_NPL_CHANNEL
            elif col in ["MRP", "MRP\n(Before KVi Discount)"]:
                row[col] = mrp
            elif col in ["Old Product ID", "old_product_id"]:
                row[col] = old_product_id
            elif col in ["Old Product Name", "old_product_name"]:
                row[col] = old_product_name
            elif col in ["Replacement Percentage", "replacement_percentage"]:
                row[col] = replacement_percentage
            elif col in WEEKDAYS:
                row[col] = 0
            else:
                row[col] = ""
        rows.append(row)

    df  = pd.DataFrame(rows, columns=mandatory_cols + optional_cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="City Plan")
        _style_ws(w.sheets["City Plan"], mandatory_cols)
    buf.seek(0)
    return buf.getvalue()


def build_hub_template(cities_hubs: dict, category: str,
                        product_id: str = "", product_name: str = "", mrp: str = "",
                        sub_type: str = "New Launch",
                        old_product_id: str = "", old_product_name: str = "",
                        replacement_percentage: str = "") -> bytes:
    """cities_hubs = {city: [hub, ...]}
    Rows pre-filled with city_name / hub_name / category; product_id, product_name, and mrp if provided."""
    mandatory_cols, optional_cols = get_template_columns("hub", sub_type)
    rows = []
    for city, hubs in cities_hubs.items():
        for hub in sorted(hubs):
            row = {}
            for col in mandatory_cols + optional_cols:
                if col == "City":
                    row[col] = city
                elif col in ["hub_name", "Hub Name"]:
                    row[col] = hub
                elif col in ["PRODUCT_ID", "Product ID"]:
                    row[col] = product_id
                elif col in ["PRODUCT_NAME", "Product Name"]:
                    row[col] = product_name
                elif col == "Anchor ID":
                    row[col] = product_id
                elif col in ["SUB_CATEGORY", "Sub Category"]:
                    row[col] = category
                elif col == "Channel":
                    row[col] = DEFAULT_NPL_CHANNEL
                elif col in ["MRP", "MRP\n(Before KVi Discount)"]:
                    row[col] = mrp
                elif col in ["Old Product ID", "old_product_id"]:
                    row[col] = old_product_id
                elif col in ["Old Product Name", "old_product_name"]:
                    row[col] = old_product_name
                elif col in ["Replacement Percentage", "replacement_percentage"]:
                    row[col] = replacement_percentage
                elif col in WEEKDAYS:
                    row[col] = 0
                else:
                    row[col] = ""
            rows.append(row)

    df  = pd.DataFrame(rows, columns=mandatory_cols + optional_cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Hub Plan")
        _style_ws(w.sheets["Hub Plan"], mandatory_cols)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# EXCEL  UPLOAD  PARSERS
# ──────────────────────────────────────────────────────────────────
def _safe_int(val) -> int:
    try:
        return int(round(float(val))) if pd.notna(val) else 0
    except Exception:
        return 0


def _parse_percent_to_decimal(val) -> float | None:
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("na", "none", "null", ""):
        return None
    try:
        if val_str.endswith("%"):
            return float(val_str.rstrip("%")) / 100.0
        num = float(val_str)
        if num > 1.0:
            return num / 100.0
        return num
    except ValueError:
        return None


def parse_city_upload(file) -> tuple:
    """Returns (df, errors).  df columns = CITY_COLS."""
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return pd.DataFrame(), [f"Cannot read file: {e}"]
    df.columns = [str(c).strip() for c in df.columns]
    
    # Normalize naming mapping to standard internal fields
    if "City" in df.columns and "city_name" not in df.columns:
        df["city_name"] = df["City"]
    elif "city_name" in df.columns and "City" in df.columns:
        df = df.drop(columns=["City"])
    if "PRODUCT_ID" in df.columns and "product_id" not in df.columns:
        df["product_id"] = df["PRODUCT_ID"]
    elif "Product ID" in df.columns and "product_id" not in df.columns:
        df["product_id"] = df["Product ID"]
    if "PRODUCT_NAME" in df.columns and "product_name" not in df.columns:
        df["product_name"] = df["PRODUCT_NAME"]
    elif "Product Name" in df.columns and "product_name" not in df.columns:
        df["product_name"] = df["Product Name"]
    if "SUB_CATEGORY" in df.columns and "category" not in df.columns:
        df["category"] = df["SUB_CATEGORY"]
    elif "Sub Category" in df.columns and "category" not in df.columns:
        df["category"] = df["Sub Category"]
    df = _normalize_upload_mrp_column(df)
    if "PLU_CODE" in df.columns and "PLU Code" not in df.columns:
        df["PLU Code"] = df["PLU_CODE"]
    df = _normalize_meat_ratio_column(df)
    if "Old Product ID" in df.columns and "old_product_id" not in df.columns:
        df["old_product_id"] = df["Old Product ID"]
    if "Old Product Name" in df.columns and "old_product_name" not in df.columns:
        df["old_product_name"] = df["Old Product Name"]
    if "Replacement Percentage" in df.columns and "replacement_percentage" not in df.columns:
        df["replacement_percentage"] = df["Replacement Percentage"]
    if "Anchor ID" in df.columns and "anchor_id" not in df.columns:
        df["anchor_id"] = df["Anchor ID"]
    if "Channel" not in df.columns:
        df["Channel"] = DEFAULT_NPL_CHANNEL

    missing = [c for c in ["city_name", "product_id"] + WEEKDAYS if c not in df.columns]
    if missing:
        return pd.DataFrame(), [f"Missing columns: {missing}"]
    
    errors = _validate_required_text_columns(
        df,
        [
            ("city_name", "City"),
            ("product_id", "Product ID"),
            ("product_name", "Product Name"),
            ("category", "Sub Category"),
        ],
    )

    # Validate MRP per row
    if "MRP" not in df.columns:
        errors.append("Column 'MRP' is missing from the uploaded sheet.")
    else:
        for idx, val in df["MRP"].items():
            mrp = pd.to_numeric(val, errors="coerce")
            if pd.isna(mrp):
                errors.append(f"Row {int(idx) + 2}: MRP must be a number greater than 0.")
            elif float(mrp) <= 0:
                errors.append(f"Row {int(idx) + 2}: MRP must be greater than 0.")

    for day in WEEKDAYS:
        if day not in df.columns:
            errors.append(f"Column '{day}' is missing from the uploaded sheet.")
            continue
        for idx, val in df[day].items():
            day_val = pd.to_numeric(val, errors="coerce")
            if pd.isna(day_val):
                errors.append(f"Row {int(idx) + 2}: {day} must be a number.")
            elif float(day_val) < 0:
                errors.append(f"Row {int(idx) + 2}: {day} must be non-negative.")

    # 3. Validate cities list against official list
    if "city_name" in df.columns and not errors:
        try:
            from features.product_launch.core import get_cities_from_salience, load_hub_salience

            valid_cities = {c.strip().lower() for c in get_cities_from_salience(load_hub_salience())}
            invalid_rows = []
            for idx, row in df.iterrows():
                city = str(row["city_name"]).strip()
                if city.lower() not in valid_cities:
                    invalid_rows.append(f"Row {idx+2}: '{city}' is not a valid planning city.")
            if invalid_rows:
                errors.extend(invalid_rows[:5]) # limit to first 5 errors to keep it clean
        except Exception as e:
            logger.warning(f"Error checking valid cities list: {e}")

    if errors:
        return pd.DataFrame(), errors[:12]

    meat_errors = _validate_meat_ratio_column(df)
    if meat_errors:
        return pd.DataFrame(), meat_errors[:8]

    pc_errors = _validate_production_pc_column(df)
    if pc_errors:
        return pd.DataFrame(), pc_errors[:8]

    try:
        df = CITY_UPLOAD_SCHEMA.validate(df)
    except Exception as err:
        return pd.DataFrame(), [f"Validation failed: {err}"]

    for day in WEEKDAYS:
        df[day] = df[day].apply(_safe_int)
    df["city_name"]    = df["city_name"].astype(str).str.strip()
    df["product_id"]   = df["product_id"].astype(str).str.strip()
    df["product_name"] = df.get("product_name", "").astype(str).str.strip()
    df["category"]     = df.get("category", "").astype(str).str.strip()
    df["anchor_id"]    = df.get("anchor_id", df["product_id"]).astype(str).str.strip()
    df.loc[df["anchor_id"].eq("") | df["anchor_id"].str.lower().eq("nan"), "anchor_id"] = df["product_id"]
    df["Channel"]      = df.get("Channel", DEFAULT_NPL_CHANNEL).astype(str).str.strip()
    df.loc[df["Channel"].eq("") | df["Channel"].str.lower().eq("nan"), "Channel"] = DEFAULT_NPL_CHANNEL
    df["MRP"]          = pd.to_numeric(df.get("MRP", 0), errors="coerce").fillna(0)

    # Clean and standardize optional columns (Accept all as strings)
    for col in ALL_POSSIBLE_OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = None
        else:
            df[col] = df[col].astype(str).str.strip().replace("nan", "").replace("None", "")

    df = df[df["city_name"].str.lower() != "nan"]
    if df.empty:
        errors.append("No valid data rows found.")
    df = _dedupe_columns(df)
    return df[CITY_COLS + ALL_POSSIBLE_OPTIONAL_COLS], errors


def parse_hub_upload(file) -> tuple:
    """Returns (df, errors).  df columns = HUB_COLS."""
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return pd.DataFrame(), [f"Cannot read file: {e}"]
    df.columns = [str(c).strip() for c in df.columns]
    
    # Normalize naming mapping to standard internal fields
    if "City" in df.columns and "city_name" not in df.columns:
        df["city_name"] = df["City"]
    elif "city_name" in df.columns and "City" in df.columns:
        df = df.drop(columns=["City"])
    if "Hub Name" in df.columns and "hub_name" not in df.columns:
        df["hub_name"] = df["Hub Name"]
    if "PRODUCT_ID" in df.columns and "product_id" not in df.columns:
        df["product_id"] = df["PRODUCT_ID"]
    elif "Product ID" in df.columns and "product_id" not in df.columns:
        df["product_id"] = df["Product ID"]
    if "PRODUCT_NAME" in df.columns and "product_name" not in df.columns:
        df["product_name"] = df["PRODUCT_NAME"]
    elif "Product Name" in df.columns and "product_name" not in df.columns:
        df["product_name"] = df["Product Name"]
    if "SUB_CATEGORY" in df.columns and "category" not in df.columns:
        df["category"] = df["SUB_CATEGORY"]
    elif "Sub Category" in df.columns and "category" not in df.columns:
        df["category"] = df["Sub Category"]
    df = _normalize_upload_mrp_column(df)
    if "PLU_CODE" in df.columns and "PLU Code" not in df.columns:
        df["PLU Code"] = df["PLU_CODE"]
    df = _normalize_meat_ratio_column(df)
    if "Old Product ID" in df.columns and "old_product_id" not in df.columns:
        df["old_product_id"] = df["Old Product ID"]
    if "Old Product Name" in df.columns and "old_product_name" not in df.columns:
        df["old_product_name"] = df["Old Product Name"]
    if "Replacement Percentage" in df.columns and "replacement_percentage" not in df.columns:
        df["replacement_percentage"] = df["Replacement Percentage"]
    if "Anchor ID" in df.columns and "anchor_id" not in df.columns:
        df["anchor_id"] = df["Anchor ID"]
    if "Channel" not in df.columns:
        df["Channel"] = DEFAULT_NPL_CHANNEL

    missing = [c for c in ["city_name", "hub_name", "product_id"] + WEEKDAYS if c not in df.columns]
    if missing:
        return pd.DataFrame(), [f"Missing columns: {missing}"]
    
    errors = _validate_required_text_columns(
        df,
        [
            ("city_name", "City"),
            ("hub_name", "Hub Name"),
            ("product_id", "Product ID"),
            ("product_name", "Product Name"),
            ("category", "Sub Category"),
        ],
    )

    if "MRP" not in df.columns:
        errors.append("Column 'MRP' is missing from the uploaded sheet.")
    else:
        for idx, val in df["MRP"].items():
            mrp = pd.to_numeric(val, errors="coerce")
            if pd.isna(mrp):
                errors.append(f"Row {int(idx) + 2}: MRP must be a number greater than 0.")
            elif float(mrp) <= 0:
                errors.append(f"Row {int(idx) + 2}: MRP must be greater than 0.")

    for day in WEEKDAYS:
        if day not in df.columns:
            errors.append(f"Column '{day}' is missing from the uploaded sheet.")
            continue
        for idx, val in df[day].items():
            day_val = pd.to_numeric(val, errors="coerce")
            if pd.isna(day_val):
                errors.append(f"Row {int(idx) + 2}: {day} must be a number.")
            elif float(day_val) < 0:
                errors.append(f"Row {int(idx) + 2}: {day} must be non-negative.")

    # 3. Validate cities list against official list
    if "city_name" in df.columns and not errors:
        try:
            from features.product_launch.core import get_cities_from_salience, load_hub_salience

            valid_cities = {c.strip().lower() for c in get_cities_from_salience(load_hub_salience())}
            invalid_rows = []
            for idx, row in df.iterrows():
                city = str(row["city_name"]).strip()
                if city.lower() not in valid_cities:
                    invalid_rows.append(f"Row {idx+2}: '{city}' is not a valid planning city.")
            if invalid_rows:
                errors.extend(invalid_rows[:5]) # limit to first 5 errors to keep it clean
        except Exception as e:
            logger.warning(f"Error checking valid cities list: {e}")

    if errors:
        return pd.DataFrame(), errors[:12]

    meat_errors = _validate_meat_ratio_column(df)
    if meat_errors:
        return pd.DataFrame(), meat_errors[:8]

    pc_errors = _validate_production_pc_column(df)
    if pc_errors:
        return pd.DataFrame(), pc_errors[:8]

    try:
        df = HUB_UPLOAD_SCHEMA.validate(df)
    except Exception as err:
        return pd.DataFrame(), [f"Validation failed: {err}"]

    for day in WEEKDAYS:
        df[day] = df[day].apply(_safe_int)
    for col in ["city_name", "hub_name", "product_id", "product_name", "category"]:
        df[col] = df.get(col, "").astype(str).str.strip()
    df["anchor_id"] = df.get("anchor_id", df["product_id"]).astype(str).str.strip()
    df.loc[df["anchor_id"].eq("") | df["anchor_id"].str.lower().eq("nan"), "anchor_id"] = df["product_id"]
    df["Channel"] = df.get("Channel", DEFAULT_NPL_CHANNEL).astype(str).str.strip()
    df.loc[df["Channel"].eq("") | df["Channel"].str.lower().eq("nan"), "Channel"] = DEFAULT_NPL_CHANNEL
    df["MRP"] = pd.to_numeric(df.get("MRP", 0), errors="coerce").fillna(0)

    # Clean and standardize optional columns (Accept all as strings)
    for col in ALL_POSSIBLE_OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = None
        else:
            df[col] = df[col].astype(str).str.strip().replace("nan", "").replace("None", "")

    df = df[df["city_name"].str.lower() != "nan"]
    if df.empty:
        errors.append("No valid data rows found.")
    df = _dedupe_columns(df)
    return df[HUB_COLS + ALL_POSSIBLE_OPTIONAL_COLS], errors


# ──────────────────────────────────────────────────────────────────
# SUBMISSION  LOG  (Google Sheet)
# ──────────────────────────────────────────────────────────────────
def _ensure_log():
    client = _get_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    try:
        ws = sh.worksheet(LOG_SHEET_NAME)
        existing = ws.row_values(1)
        new_cols = ["PLU Code", "PLU_CODE", "UOM", "Yield", "RM", "Meat Ratio", "Meat Ratio (for VA)", "Total Shelf Life", "Hub Shelf Life"]
        for col in ["Status", "Rejection_Reason", "Submitted_By", "MRP", "Old Product ID", "Old Product Name", "Replacement Percentage"] + new_cols:
            if col not in existing:
                ws.update_cell(1, len(existing) + 1, col)
                existing.append(col)
    except Exception:
        ws = sh.add_worksheet(LOG_SHEET_NAME, rows="5000", cols="50")
        ws.append_row(LOG_HEADERS)


def load_log() -> pd.DataFrame:
    from features.product_launch.sheet_reads import read_sheet_values_cached


    _ensure_log()

    def _fetch():
        from core.shared.sheets_throttle import sheets_slot


        with sheets_slot():
            sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
            return sheet.get_all_values()

    data = read_sheet_values_cached(
        SPREADSHEET_ID,
        LOG_SHEET_NAME,
        "all",
        sheet_category="hub_level_planning",
        fetcher=_fetch,
    )
    if len(data) <= 1:
        return pd.DataFrame(columns=LOG_HEADERS)
    df = pd.DataFrame(data[1:], columns=data[0])
    for col, default in [("Status", "Pending"), ("Rejection_Reason", ""),
                          ("Submitted_By", ""), ("MRP", "")]:
        if col not in df.columns:
            df[col] = default
    return df


def _sanitize(df: pd.DataFrame) -> pd.DataFrame:
    df = _dedupe_columns(df.copy().replace([np.nan, np.inf, -np.inf], ""))
    for c in df.columns:
        series = df[c]
        if isinstance(series, pd.DataFrame):
            series = series.iloc[:, 0]
        df[c] = series.map(lambda x: "" if pd.isna(x) else x)
    return df


def _overwrite_rows(sheet_id: str, sheet_name: str,
                    df_new: pd.DataFrame, key_cols: list):
    client = _get_client()
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet(sheet_name)
    data = ws.get_all_values()
    if len(data) <= 1:
        ws.clear()
        ws.update("A1", [df_new.columns.tolist()] + df_new.values.tolist(), value_input_option="USER_ENTERED")
        return
    headers = data[0]
    df_old = pd.DataFrame(data[1:], columns=headers)
    missing = [k for k in key_cols if k not in df_old.columns]
    if missing:
        ws.clear()
        ws.update("A1", [df_new.columns.tolist()] + df_new.values.tolist(), value_input_option="USER_ENTERED")
        return
    merged = df_old.merge(df_new[key_cols], on=key_cols, how="left", indicator=True)
    keep = merged[merged["_merge"] == "left_only"][headers]
    final_df = pd.concat([keep, df_new], ignore_index=True)
    ws.clear()
    ws.update("A1", [headers] + final_df.values.tolist(), value_input_option="USER_ENTERED")


def _ensure_sheet_headers(sheet_id: str, sheet_name: str, headers: list[str]):
    """Open worksheet (create if missing) and ensure row 1 contains headers."""
    sheet = _open_sheet(sheet_id, sheet_name)
    existing = sheet.row_values(1)
    if not any(str(cell).strip() for cell in existing):
        sheet.update("A1", [headers], value_input_option="USER_ENTERED")
    return sheet


def _append_sheet_rows(
    sheet_id: str,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    headers: list[str],
    chunk_size: int = 500,
) -> int:
    """Append rows without rewriting the whole worksheet (production-safe)."""
    if df is None or df.empty:
        return 0
    sheet = _ensure_sheet_headers(sheet_id, sheet_name, headers)
    cols = [c for c in headers if c in df.columns]
    if not cols:
        cols = df.columns.tolist()
    values = _sanitize(df[cols]).values.tolist()
    return _append_sheet_row_values(sheet, values, chunk_size=chunk_size)


def _append_sheet_row_values(
    sheet,
    rows: list[list],
    *,
    chunk_size: int = 500,
    table_range: str | None = None,
    row_width: int | None = None,
) -> int:
    """Append pre-built row lists directly (preserves duplicate headers / column order).

    table_range: anchor append to column A (e.g. ``A2`` when headers are on row 2).
    Without this, Google Sheets may append starting at a wrong column when the sheet
    has sparse / misaligned historical rows.
    """
    if not rows:
        return 0
    written = 0
    append_kwargs: dict = {"value_input_option": "USER_ENTERED"}
    if table_range:
        append_kwargs["table_range"] = table_range
    for start in range(0, len(rows), chunk_size):
        chunk = []
        for row in rows[start : start + chunk_size]:
            if row_width is not None:
                if len(row) < row_width:
                    row = row + [""] * (row_width - len(row))
                elif len(row) > row_width:
                    row = row[:row_width]
            chunk.append(row)
        sheet.append_rows(chunk, **append_kwargs)
        written += len(chunk)
    return written


def save_to_log(rows_df: pd.DataFrame):
    """Append submission rows to Submission_Log (no full-sheet rewrite)."""
    _ensure_log()
    df = _sanitize(rows_df)
    log_cols = [c for c in LOG_HEADERS if c in df.columns]
    if not log_cols:
        log_cols = df.columns.tolist()
    _append_sheet_rows(SPREADSHEET_ID, LOG_SHEET_NAME, df, headers=LOG_HEADERS)
    from features.product_launch.sheet_reads import invalidate_npl_sheet_cache

    invalidate_npl_sheet_cache(SPREADSHEET_ID, LOG_SHEET_NAME, "all")
    try:
        from core.shared.api_cache import CacheNS, cache_invalidate

        cache_invalidate(CacheNS.NPL_WIZARD)
    except Exception:
        pass


def update_submission_status(sub_id: str, status: str, reason: str = "", *, tail_rows: int = 600):
    """Update Status/Rejection_Reason for a submission (bounded tail read for speed)."""
    sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
    row_offset = 0
    try:
        header = sheet.row_values(1)
        if not header:
            return
        col_count = max(len(header), 52)
        end_col = _col_letter(col_count)
        row_count = sheet.row_count
        start_row = max(2, row_count - tail_rows + 1)
        body = sheet.get_values(f"A{start_row}:{end_col}{row_count}") or []
        data = [header] + body
        row_offset = start_row - 2
    except Exception:
        data = sheet.get_all_values()
        row_offset = 0
    if len(data) <= 1:
        return
    headers = data[0]
    sid_idx = headers.index("Submission_ID") + 1 if "Submission_ID" in headers else None
    stat_idx = headers.index("Status") + 1 if "Status" in headers else None
    reason_idx = headers.index("Rejection_Reason") + 1 if "Rejection_Reason" in headers else None
    if not sid_idx:
        return
    batch_updates = []
    for i, row in enumerate(data[1:], start=2 + row_offset):
        if len(row) >= sid_idx and row[sid_idx - 1] == sub_id:
            if stat_idx:
                batch_updates.append({
                    "range": f"{_col_letter(stat_idx)}{i}",
                    "values": [[status]],
                })
            if reason_idx and reason:
                batch_updates.append({
                    "range": f"{_col_letter(reason_idx)}{i}",
                    "values": [[reason]],
                })
    if batch_updates:
        sheet.batch_update(batch_updates)
        from features.product_launch.sheet_reads import invalidate_npl_sheet_cache


        invalidate_npl_sheet_cache(SPREADSHEET_ID, LOG_SHEET_NAME, "all")
        try:
            from core.shared.api_cache import CacheNS, cache_invalidate


            cache_invalidate(CacheNS.NPL_WIZARD)
        except Exception:
            pass


def _col_letter(col_idx: int) -> str:
    """1-based column index to A1 letter(s)."""
    letters = ""
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def gen_sub_id(sub_type: str) -> str:
    prefix = {"New Launch": "NEW", "Expansion": "EXP", "Replacement": "REP"}.get(sub_type, "UNK")
    return f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"


# ──────────────────────────────────────────────────────────────────
# ROW-LEVEL SHEET OPERATIONS  (position-safe, duplicate-aware)
# ──────────────────────────────────────────────────────────────────

def get_submission_rows_with_indices(submission_id: str) -> list[dict]:
    """
    Return all rows in Submission_Log that belong to submission_id,
    each annotated with '_sheet_row_index' (1-based, header = row 1).
    Using row positions instead of values makes duplicate-row deletion safe.
    """
    sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
    data = sheet.get_all_values()
    if len(data) <= 1:
        return []
    headers = data[0]
    sid_col = next((i for i, h in enumerate(headers) if h == "Submission_ID"), None)
    if sid_col is None:
        return []
    result = []
    for sheet_row_idx, row in enumerate(data[1:], start=2):  # row 1 = header
        if len(row) > sid_col and row[sid_col] == submission_id:
            row_dict = dict(zip(headers, row))
            row_dict["_sheet_row_index"] = sheet_row_idx
            result.append(row_dict)
    return result


def delete_submission_rows_by_index(
    submission_id: str,
    row_indices: list[int],
    reason: str = "",
) -> int:
    """
    Delete exactly the specified 1-based sheet row indices from Submission_Log,
    and clean up matching rows from Launch_Output, City_Plan, and Hub_Plan.
    """
    import gspread.utils as gu
    from app import config as cfg

    sheet = _open_sheet(SPREADSHEET_ID, LOG_SHEET_NAME)
    data = sheet.get_all_values()
    if len(data) <= 1:
        return 0

    headers = data[0]
    headers_upper = [str(h).strip().upper() for h in headers]
    
    sid_col = next((i for i, h in enumerate(headers_upper) if h == "SUBMISSION_ID"), None)
    pid_col = next((i for i, h in enumerate(headers_upper) if h in ("PRODUCT ID", "PRODUCT_ID")), None)
    city_col = next((i for i, h in enumerate(headers_upper) if h == "CITY"), None)
    hub_col = next((i for i, h in enumerate(headers_upper) if h == "HUB"), None)
    status_col = next((i for i, h in enumerate(headers_upper) if h == "STATUS"), None)
    reason_col = next((i for i, h in enumerate(headers_upper) if h in ("REJECTION_REASON", "REJECTION REASON")), None)

    if sid_col is None:
        return 0

    # Build the set of valid row indices for this submission (1-based, skip header)
    valid_indices: set[int] = set()
    for sheet_row_idx, row in enumerate(data[1:], start=2):
        if len(row) > sid_col and row[sid_col] == submission_id:
            valid_indices.add(sheet_row_idx)

    # Filter to only indices that belong to this submission
    to_delete = sorted(set(row_indices) & valid_indices, reverse=True)
    if not to_delete:
        return 0

    # Gather rows details to delete from other sheets
    deleted_items = []
    for idx in to_delete:
        row_vals = data[idx - 1]
        pid = row_vals[pid_col] if pid_col is not None and len(row_vals) > pid_col else ""
        city = row_vals[city_col] if city_col is not None and len(row_vals) > city_col else ""
        hub = row_vals[hub_col] if hub_col is not None and len(row_vals) > hub_col else ""
        status = row_vals[status_col] if status_col is not None and len(row_vals) > status_col else ""
        deleted_items.append({
            "product_id": str(pid).strip(),
            "city": str(city).strip(),
            "hub": str(hub).strip(),
            "status": str(status).strip(),
        })

    # Update matched row statuses to "Deleted" and add deletion reason in Submission_Log instead of physical deletion
    client = _get_client()
    sh = client.open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(LOG_SHEET_NAME)

    from gspread.cell import Cell
    cell_list = []
    for idx in to_delete:
        if status_col is not None:
            cell_list.append(Cell(row=idx, col=status_col + 1, value="Deleted"))
        if reason_col is not None:
            cell_list.append(Cell(row=idx, col=reason_col + 1, value=reason))

    if cell_list:
        ws.update_cells(cell_list, value_input_option="USER_ENTERED")

    # Clean up Launch_Output (long format)
    try:
        out_sheet = _open_sheet(SPREADSHEET_ID, OUTPUT_SHEET_NAME)
        out_data = out_sheet.get_all_values()
        if len(out_data) > 1:
            out_headers = out_data[0]
            out_headers_upper = [str(h).strip().upper() for h in out_headers]
            o_pid_col = next((i for i, h in enumerate(out_headers_upper) if h in ("PRODUCT ID", "PRODUCT_ID")), None)
            o_city_col = next((i for i, h in enumerate(out_headers_upper) if h == "CITY"), None)
            o_hub_col = next((i for i, h in enumerate(out_headers_upper) if h == "HUB"), None)
            
            out_to_delete = []
            for sheet_row_idx, row in enumerate(out_data[1:], start=2):
                r_pid = str(row[o_pid_col]).strip() if o_pid_col is not None and len(row) > o_pid_col else ""
                r_city = str(row[o_city_col]).strip() if o_city_col is not None and len(row) > o_city_col else ""
                r_hub = str(row[o_hub_col]).strip() if o_hub_col is not None and len(row) > o_hub_col else ""
                
                # Check if this row matches any deleted items
                for item in deleted_items:
                    if r_pid == item["product_id"] and r_city == item["city"] and r_hub == item["hub"]:
                        out_to_delete.append(sheet_row_idx)
                        break
            
            if out_to_delete:
                out_sh = client.open_by_key(SPREADSHEET_ID)
                out_ws = out_sh.worksheet(OUTPUT_SHEET_NAME)
                out_requests = [
                    {
                        "deleteDimension": {
                            "range": {
                                "sheetId": out_ws.id,
                                "dimension": "ROWS",
                                "startIndex": idx - 1,
                                "endIndex": idx,
                            }
                        }
                    }
                    for idx in sorted(set(out_to_delete), reverse=True)
                ]
                out_sh.batch_update({"requests": out_requests})
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("[NPL] Cleanup of Launch_Output failed: %s", exc)

    # Clean up City_Plan and Hub_Plan in NEW_PRODUCT_LAUNCH_SHEET_KEY if they were Approved
    npl_sheet_key = cfg.NEW_PRODUCT_LAUNCH_SHEET_KEY
    if npl_sheet_key:
        # Check and delete from City_Plan
        try:
            city_items = [item for item in deleted_items if not item["hub"]]
            if city_items:
                city_sheet = _open_sheet(npl_sheet_key, "City_Plan")
                city_data = city_sheet.get_all_values()
                if len(city_data) > 1:
                    city_headers = city_data[0]
                    city_headers_upper = [str(h).strip().upper() for h in city_headers]
                    c_pid_col = next((i for i, h in enumerate(city_headers_upper) if h in ("PRODUCT ID", "PRODUCT_ID", "SKU", "ANCHOR ID", "ANCHOR_ID")), None)
                    c_city_col = next((i for i, h in enumerate(city_headers_upper) if h in ("CITY", "CITY NAME", "CITY_NAME")), None)
                    
                    city_to_delete = []
                    for sheet_row_idx, row in enumerate(city_data[1:], start=2):
                        r_pid = str(row[c_pid_col]).strip() if c_pid_col is not None and len(row) > c_pid_col else ""
                        r_city = str(row[c_city_col]).strip() if c_city_col is not None and len(row) > c_city_col else ""
                        
                        for item in city_items:
                            if r_pid == item["product_id"] and r_city == item["city"]:
                                city_to_delete.append(sheet_row_idx)
                                break
                    
                    if city_to_delete:
                        city_sh = client.open_by_key(npl_sheet_key)
                        city_ws = city_sh.worksheet("City_Plan")
                        city_requests = [
                            {
                                "deleteDimension": {
                                    "range": {
                                        "sheetId": city_ws.id,
                                        "dimension": "ROWS",
                                        "startIndex": idx - 1,
                                        "endIndex": idx,
                                    }
                                }
                            }
                            for idx in sorted(set(city_to_delete), reverse=True)
                        ]
                        city_sh.batch_update({"requests": city_requests})
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning("[NPL] Cleanup of City_Plan failed: %s", exc)

        # Check and delete from Hub_Plan
        try:
            hub_items = [item for item in deleted_items if item["hub"]]
            if hub_items:
                hub_sheet = _open_sheet(npl_sheet_key, "Hub_Plan")
                hub_data = hub_sheet.get_all_values()
                if len(hub_data) > 1:
                    hub_headers = hub_data[0]
                    hub_headers_upper = [str(h).strip().upper() for h in hub_headers]
                    h_pid_col = next((i for i, h in enumerate(hub_headers_upper) if h in ("PRODUCT ID", "PRODUCT_ID", "SKU", "ANCHOR ID", "ANCHOR_ID")), None)
                    h_city_col = next((i for i, h in enumerate(hub_headers_upper) if h in ("CITY", "CITY NAME", "CITY_NAME")), None)
                    h_hub_col = next((i for i, h in enumerate(hub_headers_upper) if h in ("HUB", "HUB NAME", "HUB_NAME", "HUB_NAME_", "HUB_NAME_COL", "HUB_NAME_HEADER")), None)
                    
                    hub_to_delete = []
                    for sheet_row_idx, row in enumerate(hub_data[1:], start=2):
                        r_pid = str(row[h_pid_col]).strip() if h_pid_col is not None and len(row) > h_pid_col else ""
                        r_city = str(row[h_city_col]).strip() if h_city_col is not None and len(row) > h_city_col else ""
                        r_hub = str(row[h_hub_col]).strip() if h_hub_col is not None and len(row) > h_hub_col else ""
                        
                        for item in hub_items:
                            if r_pid == item["product_id"] and r_city == item["city"] and r_hub == item["hub"]:
                                hub_to_delete.append(sheet_row_idx)
                                break
                    
                    if hub_to_delete:
                        hub_sh = client.open_by_key(npl_sheet_key)
                        hub_ws = hub_sh.worksheet("Hub_Plan")
                        hub_requests = [
                            {
                                "deleteDimension": {
                                    "range": {
                                        "sheetId": hub_ws.id,
                                        "dimension": "ROWS",
                                        "startIndex": idx - 1,
                                        "endIndex": idx,
                                    }
                                }
                            }
                            for idx in sorted(set(hub_to_delete), reverse=True)
                        ]
                        hub_sh.batch_update({"requests": hub_requests})
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning("[NPL] Cleanup of Hub_Plan failed: %s", exc)

    # Invalidate caches
    from features.product_launch.sheet_reads import invalidate_npl_sheet_cache
    try:
        invalidate_npl_sheet_cache(SPREADSHEET_ID, LOG_SHEET_NAME, "all")
        from core.shared.api_cache import CacheNS, cache_invalidate
        cache_invalidate(CacheNS.NPL_WIZARD)
    except Exception:
        pass

    return len(to_delete)




# ──────────────────────────────────────────────────────────────────
# DUPLICATE  DETECTION
# ──────────────────────────────────────────────────────────────────
def check_duplicates_city(df_log: pd.DataFrame, sub_type: str,
                           product_id: str, cities: list) -> pd.DataFrame:
    """Check city + product_id duplicates (after city-level upload)."""
    if df_log.empty:
        return pd.DataFrame()
    status_col = "Status" if "Status" in df_log.columns else None
    active = (
        df_log[~df_log[status_col].isin(["Withdrawn", "Voided", "Expired"])]
        if status_col
        else df_log
    )
    mask = (
        (active["Submission_Type"] == sub_type) &
        (active["Product ID"].astype(str) == str(product_id)) &
        (active["City"].isin(cities))
    )
    return active[mask]


def check_duplicates_hub(df_log: pd.DataFrame, sub_type: str,
                          hub_df: pd.DataFrame) -> pd.DataFrame:
    """Check city + hub + product_id duplicates (after hub-level upload) using vectorized inner merge."""
    if df_log.empty or hub_df.empty:
        return pd.DataFrame()
    status_col = "Status" if "Status" in df_log.columns else None
    active = (
        df_log[~df_log[status_col].isin(["Withdrawn", "Voided", "Expired"])].copy()
        if status_col
        else df_log.copy()
    )
    if active.empty:
        return pd.DataFrame()

    # Standardize types and strings for key columns to ensure robust matches
    active["Product ID"] = active["Product ID"].astype(str).str.strip()
    active["City"] = active["City"].astype(str).str.strip()
    active["Hub"] = active["Hub"].astype(str).str.strip()
    active["Submission_Type"] = active["Submission_Type"].astype(str).str.strip()

    temp_hub = hub_df.copy()
    temp_hub["Product ID"] = temp_hub["product_id"].astype(str).str.strip()
    temp_hub["City"] = temp_hub["city_name"].astype(str).str.strip()
    temp_hub["Hub"] = temp_hub["hub_name"].astype(str).str.strip()
    temp_hub["Submission_Type"] = str(sub_type).strip()

    merged = pd.merge(
        active,
        temp_hub[["Product ID", "City", "Hub", "Submission_Type"]],
        on=["Product ID", "City", "Hub", "Submission_Type"],
        how="inner"
    )
    return merged.drop_duplicates()


# ──────────────────────────────────────────────────────────────────
# DATE  HELPERS
# ──────────────────────────────────────────────────────────────────
def get_earliest_monday(min_days: int = 4) -> date:
    earliest = date.today() + timedelta(days=min_days)
    offset   = (7 - earliest.weekday()) % 7
    return earliest if earliest.weekday() == 0 else earliest + timedelta(days=offset)


# ──────────────────────────────────────────────────────────────────
# EMAIL  PLACEHOLDER
# ──────────────────────────────────────────────────────────────────
# Streamlit UI code removed. This module now provides only backend helpers for FastAPI and Next.js.


# ──────────────────────────────────────────────────────────────────
# WIDE → LONG  CONVERTER
# ──────────────────────────────────────────────────────────────────
def wide_to_long(df: pd.DataFrame) -> pd.DataFrame:
    id_cols  = [c for c in df.columns if c not in WEEKDAYS]
    day_cols = [c for c in WEEKDAYS if c in df.columns]
    if not day_cols:
        return df
    long_df = df.melt(id_vars=id_cols, value_vars=day_cols,
                      var_name="Day", value_name="Plan")
    return long_df.reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────
# SHARED:  SUBMIT  HUB-LEVEL  DATA  TO  SHEETS
# ──────────────────────────────────────────────────────────────────
def _submit_hub_df(
    hub_df: pd.DataFrame,
    sub_type: str,
    username: str = "",
    *,
    status: str = "Pending",
    rejection_reason: str = "",
) -> str:
    """
    Normalise, attach metadata, write to Submission_Log + Launch_Output.
    Launch Date is read per-row from the 'Launch Date' column in hub_df.
    Returns the generated Submission_ID.
    """
    if hub_df is None or hub_df.empty:
        raise ValueError("Cannot submit an empty hub plan.")

    df = _dedupe_columns(hub_df.copy())

    # Drop alias columns before rename to avoid duplicate headers (City + city_name, etc.)
    alias_drop = []
    if "city_name" in df.columns and "City" in df.columns:
        alias_drop.append("City")
    if "hub_name" in df.columns and "Hub" in df.columns:
        alias_drop.append("Hub")
    if "product_id" in df.columns and "Product ID" in df.columns:
        alias_drop.append("Product ID")
    if "product_name" in df.columns and "Product Name" in df.columns:
        alias_drop.append("Product Name")
    if "category" in df.columns and "Category" in df.columns:
        alias_drop.append("Category")
    if alias_drop:
        df = df.drop(columns=alias_drop, errors="ignore")

    # If Start Date is already present (e.g. from optional template columns),
    # prioritize it and drop Launch Date to avoid duplicate column name errors.
    if "Start Date" in df.columns and "Launch Date" in df.columns:
        df = df.drop(columns=["Launch Date"])

    # Rename to sheet canonical column names
    df = df.rename(columns={
        "city_name":    "City",
        "hub_name":     "Hub",
        "product_id":   "Product ID",
        "product_name": "Product Name",
        "category":     "Category",
        "Launch Date":  "Start Date",
        "old_product_id": "Old Product ID",
        "old_product_name": "Old Product Name",
        "replacement_percentage": "Replacement Percentage",
    })
    df = _dedupe_columns(df)

    if "Hub" not in df.columns:
        df["Hub"] = ""

    # Convert Start Date to string (handles date objects and strings)
    if "Start Date" in df.columns:
        df["Start Date"] = df["Start Date"].apply(
            lambda x: x.strftime("%Y-%m-%d") if hasattr(x, "strftime") else str(x)
        )
    else:
        df["Start Date"] = ""

    for day in WEEKDAYS:
        if day not in df.columns:
            df[day] = 0
    df[WEEKDAYS] = df[WEEKDAYS].fillna(0).astype(int)
    
    # Format percentage fields to decimal floats (Meat Ratio stays free-text)
    pct_cols = ["Yield", "Replacement Percentage", "replacement_percentage"]
    for col in pct_cols:
        if col in df.columns:
            df[col] = df[col].apply(_parse_percent_to_decimal)
    for col in ["Meat Ratio", "Meat Ratio (for VA)", "UOM", "RM", "Total Shelf Life", "Hub Shelf Life", "PLU Code", "PLU_CODE"]:
        if col in df.columns:
            df[col] = df[col].apply(_optional_str)

    submitted_by = username or ""
    sub_id = gen_sub_id(sub_type)
    df["Timestamp"]        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df["Submission_ID"]    = sub_id
    df["Submission_Type"]  = sub_type
    df["Status"]           = status
    df["Rejection_Reason"] = rejection_reason
    df["Submitted_By"]     = submitted_by

    # Launch_Output: append long format (creates tab if missing)
    out_wide = df[["Product ID", "Product Name", "Category",
                   "City", "Hub", "Start Date"] + WEEKDAYS]
    out_long = wide_to_long(out_wide)
    out_long = _sanitize(out_long)
    _append_sheet_rows(
        SPREADSHEET_ID,
        OUTPUT_SHEET_NAME,
        out_long,
        headers=LAUNCH_OUTPUT_HEADERS,
    )

    # Submission_Log: append only (no full-sheet rewrite)
    log_cols = ["Timestamp", "Submission_ID", "Submission_Type",
                "Product ID", "Product Name", "Category",
                "City", "Hub", "MRP", "Start Date",
                "Status", "Rejection_Reason", "Submitted_By",
                "Old Product ID", "Old Product Name", "Replacement Percentage"] + WEEKDAYS + [
                "PLU Code", "PLU_CODE", "UOM", "Yield", "RM", "Meat Ratio", "Meat Ratio (for VA)",
                "Total Shelf Life", "Hub Shelf Life", "Channel", "anchor_id"]
    log_df = df[[c for c in log_cols if c in df.columns]]
    save_to_log(_sanitize(log_df))

    return sub_id


# ──────────────────────────────────────────────────────────────────
# UI  HELPERS
# ──────────────────────────────────────────────────────────────────
def _stage_bar(stage: str, stages: list):
    """Simple text progress breadcrumb."""
    parts = []
    for s in stages:
        if s == stage:
            parts.append(f"**→ {s}**")
        else:
            parts.append(s)
    print("  ›  ".join(parts))
    print("---")


def _show_zero_sal_warning(zero_sal_info: dict):
    if zero_sal_info:
        hubs_list = ", ".join(
            f"{city}: {', '.join(hubs)}" for city, hubs in zero_sal_info.items()
        )
        print(
            f"⚠️ These hubs have **no salience data** for this category and received "
            f"an **equal split**: {hubs_list}. You can edit the values in the table below."
        )


# Streamlit UI handlers were removed from this module.
# This file now provides backend-only helpers for FastAPI/Next.js services.


def validate_npl_upload(df: pd.DataFrame) -> dict:

    """Validate an uploaded NPL workbook (city-level or hub-level schema)."""
    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]
    errors: list[str] = []

    optional_cols = ["product_name", "category", "MRP"]
    for col in optional_cols:
        if col not in work.columns:
            work[col] = None

    if "hub_name" in work.columns:
        try:
            validated = HUB_UPLOAD_SCHEMA.validate(work)
            return {"valid": True, "type": "hub", "rows": len(validated)}
        except Exception as exc:
            errors.append(f"Hub schema: {exc}")

    try:
        validated = CITY_UPLOAD_SCHEMA.validate(work)
        return {"valid": True, "type": "city", "rows": len(validated)}
    except Exception as exc:
        errors.append(f"City schema: {exc}")

    return {"valid": False, "errors": errors}
